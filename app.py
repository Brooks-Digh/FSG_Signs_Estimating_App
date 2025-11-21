from flask import Flask, session, render_template_string, url_for, request, redirect, flash, jsonify
import pyodbc
import math
from decimal import Decimal
import base64
from functools import wraps

app = Flask(__name__)
app.secret_key = 'super_secret_key_123'

########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
""""""
""""""
"""DATABASE CONNECTION."""
""""""
""""""
########################################################################################################################

DB_PATH = r"C:\Users\Brooks\OneDrive\Desktop\Sign_App1 - Step 3.accdb"
CONN_STR = (
    r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"
    rf"DBQ={DB_PATH};"
)

########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
""""""
""""""
"""DATABASE/HELPER FUNCTIONS"""
""""""
""""""
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################

"""HELPER FUNCTION TO GET CUSTOMER DATA"""

def get_customers():
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    query = """
        SELECT customer_ID, customer_name, customer_email, billing_address, billing_city, billing_state, billing_zip,
        contact_first_name, contact_last_name, contact_phone
        FROM Customers
        ORDER BY customer_name
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO CREATE A NEW CUSTOMER"""

def add_customer(customer_name, customer_email, billing_address, billing_city, billing_state, billing_zip,
                 contact_first_name, contact_last_name, contact_phone):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    query = """
        INSERT INTO Customers (customer_name, customer_email, billing_address, billing_city, billing_state, billing_zip,
        contact_first_name, contact_last_name, contact_phone)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """
    cursor.execute(query, (customer_name, customer_email, billing_address, billing_city, billing_state,
                           billing_zip, contact_first_name, contact_last_name, contact_phone))
    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO GET OPPORTUNITY DATA."""

def get_opportunities():
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    query = """
        SELECT o.opportunity_ID, o.opportunity_name,
               c.customer_ID, c.customer_name, c.customer_email
        FROM Opportunities o
        INNER JOIN Customers c ON CStr(o.customer_ID) = CStr(c.customer_ID)
        ORDER BY o.opportunity_ID;
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO CREATE A NEW OPPORTUNITY"""

def add_opportunity(customer_id, opportunity_name, tax_rate, site_address, site_city, site_state, site_zip):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO Opportunities (customer_ID, opportunity_name, tax_rate, opportunity_price, site_address, site_city, 
        site_state, site_zip)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (customer_id, opportunity_name, tax_rate, 0, site_address, site_city, site_state, site_zip))
    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO UPDATE OPPORTUNITY TOTAL PRICE"""

def update_opportunity_price(opportunity_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Sum all line item totals (unit_price * quantity)
    cursor.execute("""
        SELECT SUM(CDbl(quantity) * CDbl(unit_price))
        FROM Line_Items
        WHERE opportunity_ID = ?
    """, (opportunity_id,))
    total = cursor.fetchone()[0] or 0

    # Update the Opportunities table
    cursor.execute("""
        UPDATE Opportunities
        SET opportunity_price = ?
        WHERE opportunity_ID = ?
    """, (total, opportunity_id))
    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO GET LINE ITEM DATA."""

def get_line_items(opportunity_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT line_ID, line_item_description, quantity, unit_price, activation_status, line_item_sequence
        FROM Line_Items
        WHERE opportunity_ID = ?
        ORDER BY line_item_sequence
    """, (opportunity_id,))
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO CREATE A NEW LINE ITEM"""

def add_line_item(opportunity_id, description, quantity, sequence_number, activation_status="ACTIVE"):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
         INSERT INTO Line_Items (opportunity_ID, line_item_description, quantity, activation_status, line_item_sequence)
         VALUES (?, ?, ?, ?, ?)
     """, (opportunity_id, description, quantity, activation_status, sequence_number))
    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO UPDATE LINE ITEM DATA"""

def update_line_item(line_id, description, quantity, activation_status, sequence_number):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE Line_Items
        SET line_item_description = ?, 
            quantity = ?, 
            activation_status = ?, 
            line_item_sequence = ?
        WHERE line_ID = ?
    """, (description, quantity, activation_status, sequence_number, line_id))
    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO UPDATE LINE ITEM TOTALS (unit_cost & unit_price)"""

def update_line_item_totals(line_id):
    """
    Recalculate unit_cost and unit_price for a line item based on its components
    """
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Sum all component costs
    cursor.execute("""
        SELECT SUM(unit_cost), SUM(unit_price)
        FROM Components
        WHERE line_ID = ?
    """, (line_id,))
    row = cursor.fetchone()
    total_cost = row[0] or 0
    total_price = row[1] or 0

    # Update Line_Items table
    cursor.execute("""
        UPDATE Line_Items
        SET unit_cost = ?, unit_price = ?
        WHERE line_ID = ?
    """, (total_cost, total_price, line_id))
    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO GET THE OPPORTUNITY_ID FROM A LINE_ID"""

def get_opportunity_id_by_line(line_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT opportunity_ID
        FROM Line_Items
        WHERE line_ID = ?
    """, (line_id,))
    row = cursor.fetchone()
    conn.close()
    return row[0] if row else None

########################################################################################################################

"""HELPER FUNCTION TO GET COMPONENT DATA"""

def get_components(line_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    query = """
        SELECT component_ID, line_ID, component_type_ID, quantity, 
               unit_cost, unit_price, factor1, factor2, factor3, factor4, factor5
        FROM Components
        WHERE line_ID = ?
        ORDER BY component_ID
    """
    cursor.execute(query, (line_id,))
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO ADD A NEW COMPONENT"""

def insert_component(line_id, component_type_id):
    try:
        conn = pyodbc.connect(CONN_STR)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO Components
            (line_ID, component_type_ID, quantity, unit_cost, unit_price,
             factor1, factor2, factor3, factor4, factor5, factor6)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (line_id, component_type_id, 1, 0, 0, 0, 0, 0, 0, 0, 0))
        conn.commit()
    except Exception as e:
        print("Error inserting component:", e)
    finally:
        conn.close()

########################################################################################################################

"""HELPER FUNCTION TO GET COMPONENT TYPES DATA"""

def get_component_types():
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    query = """
        SELECT [component_type_ID], [component_types_description]
        FROM [Component_Types]
        ORDER BY [component_types_description]
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO UPDATE COMPONENT TOTALS (unit_cost & unit_price)"""

def update_component_totals(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # --- Sum MFG materials ---
    cursor.execute("""
        SELECT SUM(cm.quantity * m.material_price)
        FROM component_MFG_Materials cm
        INNER JOIN Materials m ON cm.material_ID = m.material_ID
        WHERE cm.component_ID = ?
    """, (component_id,))
    material_total = cursor.fetchone()[0] or 0

    # --- Sum MFG labor ---
    cursor.execute("""
        SELECT SUM(cl.quantity * l.burden_rate)
        FROM component_MFG_Labor cl
        INNER JOIN Labor_Types l ON cl.labor_ID = l.labor_ID
        WHERE cl.component_ID = ?
    """, (component_id,))
    labor_total = cursor.fetchone()[0] or 0

    unit_cost = float(material_total) + float(labor_total)

    # ⚡ You decide how unit_price is set:
    # Example: cost + 20% markup
    unit_price = float(material_total) * float(2.23) + float(labor_total) * float(3.62)

    # --- Save back into Components table ---
    cursor.execute("""
        UPDATE Components
        SET unit_cost = ?, unit_price = ?
        WHERE component_ID = ?
    """, (unit_cost, unit_price, component_id))
    conn.commit()
    conn.close()

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""MANUFACTURING SPECIFIC HELPER FUNCTIONS"""

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""HELPER FUNCTION TO GET MATERIALS DATA"""

def get_materials():
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    query = """
        SELECT material_ID, material_description, material_price, material_unit, stock
        FROM Materials
        WHERE stock='YES'
        ORDER BY material_description
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO GET COMPONENT MATERIAL DATA"""

def get_component_materials(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            cm.[ID],
            cm.[component_ID],
            cm.[material_ID],
            cm.[quantity],
            m.[material_description],
            m.[material_unit],
            m.[material_price]
        FROM [component_MFG_Materials] AS cm
        INNER JOIN [Materials] AS m
            ON cm.[material_ID] = m.[material_ID]
        WHERE cm.[component_ID] = ?
        ORDER BY cm.[ID]
    """, (component_id,))
    rows = cursor.fetchall()
    conn.close()

    results = []
    for r in rows:
        results.append({
            "ID": r.ID,
            "material_ID": r.material_ID,
            "material_description": r.material_description,
            "material_unit": r.material_unit,
            "material_price": float(r.material_price or 0),
            "quantity": float(r.quantity or 0)
        })
    return results

########################################################################################################################

"""HELPER FUNCTION TO ADD MATERIALS TO A COMPONENT"""

def add_component_material(component_id, material_id, quantity):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    query = """
        INSERT INTO Component_Materials (component_ID, material_ID, quantity)
        VALUES (?, ?, ?)
    """
    cursor.execute(query, (component_id, material_id, quantity))
    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO GET LABOR TYPES DATA"""

def get_labor_types():
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    query = """
        SELECT labor_ID, labor_type, burden_rate
        FROM Labor_Types
        ORDER BY labor_ID
    """
    cursor.execute(query)
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO GET COMPONENT LABOR DATA"""

def get_component_labor(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            cl.[line_item_labor_ID],
            cl.[component_ID],
            cl.[labor_ID],
            cl.[quantity],
            l.[labor_type],
            l.[burden_rate]
        FROM [component_MFG_Labor] AS cl
        INNER JOIN [Labor_Types] AS l
            ON cl.[labor_ID] = l.[labor_ID]
        WHERE cl.[component_ID] = ?
        ORDER BY cl.[line_item_labor_ID]
    """, (component_id,))
    rows = cursor.fetchall()
    conn.close()

    results = []
    for r in rows:
        results.append({
            "line_item_labor_ID": r.line_item_labor_ID,
            "labor_ID": r.labor_ID,
            "labor_type": r.labor_type,
            "burden_rate": float(r.burden_rate or 0),
            "quantity": float(r.quantity or 0)
        })
    return results

########################################################################################################################

"""HELPER FUNCTION TO ADD LABOR TO A COMPONENT"""

def add_component_labor(component_id, labor_id, quantity):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    query = """
        INSERT INTO Component_Labor (component_ID, labor_ID, quantity)
        VALUES (?, ?, ?)
    """
    cursor.execute(query, (component_id, labor_id, quantity))
    conn.commit()
    conn.close()

########################################################################################################################

""" HELPER FUNCTION THAT UPDATES THE LINE ITEM UNIT PRICE AS THE SUM OF THE COMPONENT TOTALS"""

def update_line_item_totals_from_components(line_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Sum all component subtotals for this line item
    cursor.execute("""
        SELECT SUM(c.unit_price * c.quantity) AS total_price
        FROM Components c
        WHERE c.line_ID = ?
    """, (line_id,))
    row = cursor.fetchone()
    total_price = float(row.total_price or 0)

    # Update line item
    cursor.execute("""
        UPDATE Line_Items
        SET unit_price = ?
        WHERE line_ID = ?
    """, (total_price, line_id))
    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO GET ALL COMPONENTS THAT BELONG TO A SPECIFIC LINE ITEM"""

def get_line_id_by_component(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("SELECT line_ID FROM Components WHERE component_ID = ?", (component_id,))
    row = cursor.fetchone()
    conn.close()
    return row.line_ID if row else None

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""INSTALL SPECIFIC HELPER FUNCTIONS"""

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""HELPER FUNCTION TO GET INSTALL LABOR TYPES DATA"""

def get_install_labor_types():
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT install_labor_ID, install_labor_type, burden_rate FROM Install_Labor_Types ORDER BY install_labor_type")
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO GET COMPONENT INSTALL LABOR DATA"""

def get_component_install_materials(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            cim.[component_install_materials_ID] AS ID,
            cim.[component_ID],
            cim.[quantity],
            cim.[unit_cost],
            cim.[material_description],
            cim.[material_unit]
        FROM [component_install_Materials] AS cim
        WHERE cim.[component_ID] = ?
        ORDER BY cim.[component_install_materials_ID]
    """, (component_id,))
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO GET COMPONENT INSTALL MATERIAL DATA"""

def get_component_install_labor(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT 
            cil.[component_install_labor_ID] AS ID,
            cil.[component_ID],
            cil.[install_labor_ID],
            cil.[quantity],
            ilt.[install_labor_type],
            ilt.[burden_rate]
        FROM [component_install_Labor] AS cil
        INNER JOIN [Install_Labor_Types] AS ilt
            ON cil.[install_labor_ID] = ilt.[install_labor_ID]
        WHERE cil.[component_ID] = ?
        ORDER BY cil.[component_install_labor_ID];
    """, (component_id,))
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO UPDATE THE COMPONENT UNIT COST"""

def update_component_unit_cost(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT SUM(quantity * unit_price)
        FROM component_install_materials
        WHERE component_ID = ?
    """, (component_id,))

    total_material_cost = cursor.fetchone()[0] or 0

    cursor.execute("""
        UPDATE Components
        SET unit_cost = ?
        WHERE component_ID = ?
    """, (total_material_cost, component_id))

    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO UPDATE INSTALL COMPONENT TOTAL COST AND PRICE"""

def update_install_component_totals(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # --- Sum INSTALL materials ---
    cursor.execute("""
        SELECT SUM(quantity * unit_cost)
        FROM component_install_Materials
        WHERE component_ID = ?
    """, (component_id,))
    material_total = cursor.fetchone()[0] or 0

    # --- Sum INSTALL labor ---
    cursor.execute("""
        SELECT SUM(ci.quantity * ilt.burden_rate)
        FROM component_install_Labor ci
        INNER JOIN Install_Labor_Types ilt ON ci.install_labor_ID = ilt.install_labor_ID
        WHERE ci.component_ID = ?
    """, (component_id,))
    labor_total = cursor.fetchone()[0] or 0

    # --- Sum SUBCONTRACTOR install cost ---
    cursor.execute("""
        SELECT SUM(subcontractor_cost)
        FROM subcontractor_install_cost
        WHERE component_ID = ?
    """, (component_id,))
    subcontract_total = cursor.fetchone()[0] or 0

    # --- Compute new unit_cost ---
    unit_cost = float(material_total) + float(labor_total) + float(subcontract_total)

    # --- Compute new unit_price ---
    # Materials use 1.45 mark-up
    # Labor uses 1.32 mark-up
    # Subcontractor cost ALSO uses 1.32 mark-up (per your requirement)
    unit_price = (
        float(material_total) * 1.45 +
        float(labor_total) * 1.32 +
        float(subcontract_total) * 1.32
    )

    # --- Update the Components table ---
    cursor.execute("""
        UPDATE Components
        SET unit_cost = ?, unit_price = ?
        WHERE component_ID = ?
    """, (unit_cost, unit_price, component_id))

    conn.commit()
    conn.close()

########################################################################################################################

def get_subcontract_install_costs(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT subcontracted_install_labor_ID, component_ID, subcontractor_cost
        FROM subcontractor_install_cost
        WHERE component_ID = ?
    """, (component_id,))
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

def add_subcontract_install_cost(component_id, cost):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO subcontractor_install_cost (component_ID, subcontractor_cost)
        VALUES (?, ?)
    """, (component_id, cost))
    conn.commit()
    conn.close()

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""EMC SPECIFIC HELPER FUNCTIONS"""

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""HELPER FUNCTION TO GET EMC DATA"""

def get_component_emc(component_id):
    """Return all EMC units for a given component as dicts for Jinja rendering."""
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            EMC_unit_ID AS ID,
            component_ID,
            quantity,
            unit_cost,
            EMC_description
        FROM Component_EMC
        WHERE component_ID = ?
        ORDER BY EMC_unit_ID
    """, (component_id,))

    columns = [col[0] for col in cursor.description]
    rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO UPDATE EMC PRICE"""

def update_emc_component_totals(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Sum EMC rows
    cursor.execute("""
        SELECT SUM(quantity * unit_cost)
        FROM component_EMC
        WHERE component_ID = ?
    """, (component_id,))
    total_cost = cursor.fetchone()[0] or 0

    unit_cost = float(total_cost)
    unit_price = unit_cost * 1.315

    cursor.execute("""
        UPDATE Components
        SET unit_cost = ?, unit_price = ?
        WHERE component_ID = ?
    """, (unit_cost, unit_price, component_id))

    conn.commit()
    conn.close()

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""PIPE AND FOUNDATION SPECIFIC HELPER FUNCTIONS"""

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""HELPER FUNCTION TO GET PIPE AND FOUNDATION DATA"""

def get_component_pipe_foundation(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT component_pipe_and_foundation_ID, component_ID,
               base_pipe_diameter, base_pipe_footage,
               stack_pipe1_diameter, stack_pipe1_footage,
               stack_pipe2_diameter, stack_pipe2_footage,
               stack_pipe3_diameter, stack_pipe3_footage,
               stack_pipe4_diameter, stack_pipe4_footage,
               pier_diameter, pier_depth, pier_quantity,
               rectangular_footer_length, rectangular_footer_width, rectangular_footer_depth,
               digging_cost, concrete_cost, additional_footer_cost, pipe_cost
        FROM component_pipe_and_foundation
        WHERE component_ID = ?
    """, (component_id,))
    row = cursor.fetchone()
    conn.close()
    return row

########################################################################################################################

"""HELPER FUNCTION TO UPDATE PIPE AND FOUNDATION PRICE"""

def update_pipe_foundation_totals(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT digging_cost, concrete_cost, additional_footer_costs, pipe_cost
        FROM component_pipe_and_foundation
        WHERE component_ID = ?
    """, (component_id,))
    row = cursor.fetchone()

    if row:
        digging, concrete, additional, pipe = [float(v or 0) for v in row]
        unit_cost = digging + concrete + additional + pipe
        unit_price = unit_cost * 1.35

        cursor.execute("""
            UPDATE Components
            SET unit_cost = ?, unit_price = ?
            WHERE component_ID = ?
        """, (unit_cost, unit_price, component_id))

    conn.commit()
    conn.close()

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""MASONRY SPECIFIC FUNCTIONS"""

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""HELPER FUNCTION TO GET MASONRY DATA"""

def get_component_masonry(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT component_masonry_ID, component_ID, quantity, unit_cost, masonry_description
        FROM component_Masonry
        WHERE component_ID = ?
        ORDER BY component_masonry_ID
    """, (component_id,))
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO UPDATE MASONRY PRICE"""

def update_masonry_component_totals(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT SUM(quantity * unit_cost)
        FROM component_Masonry
        WHERE component_ID = ?
    """, (component_id,))
    total_cost = cursor.fetchone()[0] or 0

    unit_cost = float(total_cost)
    unit_price = unit_cost * 1.35

    cursor.execute("""
        UPDATE Components
        SET unit_cost = ?, unit_price = ?
        WHERE component_ID = ?
    """, (unit_cost, unit_price, component_id))

    conn.commit()
    conn.close()

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""RENTAL EQUIPMENT SPECIFIC FUNCTIONS"""

########################################################################################################################
########################################################################################################################
########################################################################################################################

"""HELPER FUNCTION TO GET RENTAL EQUIPMENT DATA"""

def get_component_rental_equipment(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT component_equipment_ID, component_ID, quantity, unit_cost, equipment_description
        FROM component_Rental_Equipment
        WHERE component_ID = ?
        ORDER BY component_equipment_ID
    """, (component_id,))
    rows = cursor.fetchall()
    conn.close()
    return rows

########################################################################################################################

"""HELPER FUNCTION TO UPDATE RENTAL EQUIPMENT PRICE"""

def update_rental_equipment_component_totals(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT SUM(quantity * unit_cost)
        FROM component_Rental_Equipment
        WHERE component_ID = ?
    """, (component_id,))
    total_cost = cursor.fetchone()[0] or 0

    unit_cost = float(total_cost)
    unit_price = unit_cost * 1.35

    cursor.execute("""
        UPDATE Components
        SET unit_cost = ?, unit_price = ?
        WHERE component_ID = ?
    """, (unit_cost, unit_price, component_id))

    conn.commit()
    conn.close()

########################################################################################################################

"""HELPER FUNCTION TO WRAP ROUTES THAT SALES TEAM IS NOT ALLOWED TO SEE"""

def requires_role(*allowed_roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if session.get("employee_type") not in allowed_roles:
                return "Access denied", 403
            return f(*args, **kwargs)

        return wrapper

    return decorator

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "employee_ID" not in session:
            return redirect(url_for("login_route"))
        return f(*args, **kwargs)
    return wrapper

########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
""""""
""""""
"""HTML TEMPLATES"""
""""""
""""""
########################################################################################################################

LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Login</title>
    <style>
        body { font-family: Arial; margin: 40px; }
        .container {
            width: 350px; margin: 80px auto; padding: 20px;
            border: 1px solid #ccc; border-radius: 8px;
        }
        input[type=text], input[type=password] {
            width: 100%; padding: 8px; margin-top: 5px;
        }
        input[type=submit] {
            margin-top: 15px; padding: 8px 12px;
            background: #0275d8; color: white; border: none; cursor: pointer;
        }
        .error { color: red; margin-top: 10px; }
    </style>
</head>
<body>

<div class="container">
    <h2>Employee Login</h2>

    <form method="POST">
        <label>Email:</label>
        <input type="text" name="email" required>

        <label>Password:</label>
        <input type="password" name="password" required>

        <input type="submit" value="Login">
    </form>

    {% if error %}
        <p class="error">{{ error }}</p>
    {% endif %}
</div>

</body>
</html>
"""

########################################################################################################################

OPPORTUNITIES_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Opportunities</title>
    <link rel="stylesheet" href="https://cdn.datatables.net/1.13.7/css/jquery.dataTables.min.css">
    <link href="https://cdn.jsdelivr.net/npm/tom-select/dist/css/tom-select.css" rel="stylesheet">
    <style>
        body {
            font-family: Arial, sans-serif;
        }

        table {
            width: 100%;
            border-collapse: collapse;
        }

        th, td {
            border: 1px solid #333;
            padding: 6px;
            text-align: left;
        }

        h1, h2 {
            margin-top: 30px;
        }

        /* Layout */
        .form-grid {
            display: flex;
            gap: 40px;
            align-items: flex-start;
        }

        .form-col {
            flex: 1;
        }

        label {
            display: inline-block;
            width: 140px;
            margin-bottom: 6px;
        }

        input, select {
            width: 200px;
        }

        /* DataTables layout */
        div.dataTables_filter {
            text-align: center !important;
            float: none !important;
        }

        div.dataTables_filter label {
            width: 100%;
            display: flex;
            justify-content: center;
            align-items: center;
            gap: 8px;
        }

        div.dataTables_filter input {
            width: 250px;
        }

        /* Fix duplicate dropdowns from Tom Select */
        .ts-wrapper + select {
            display: none !important;
        }
    </style>
</head>
<body>
    <h1>Opportunities</h1>
    <hr>

    <!-- ======================================== -->
    <!-- Create New Customer -->
    <!-- ======================================== -->
    <h2>Create New Customer</h2>
    <form method="POST" action="{{ url_for('add_customer_route') }}">
        <div class="form-grid">
            <div class="form-col">
                <label>Customer Name:</label>
                <input type="text" name="customer_name" required><br><br>

                <label>Customer Email:</label>
                <input type="email" name="customer_email" required><br><br>

                <label>Contact First Name:</label>
                <input type="text" name="contact_first_name"><br><br>

                <label>Contact Last Name:</label>
                <input type="text" name="contact_last_name"><br><br>

                <label>Phone:</label>
                <input type="tel" name="phone" placeholder="555-555-5555"><br><br>
            </div>

            <div class="form-col">
                <label>Billing Address:</label>
                <input type="text" name="billing_address"><br><br>

                <label>Billing City:</label>
                <input type="text" name="billing_city"><br><br>

                <label>Billing State:</label>
                <input type="text" name="billing_state" maxlength="2" placeholder="TX"><br><br>

                <label>Billing Zip:</label>
                <input type="text" name="billing_zip" maxlength="10" placeholder="75001"><br><br>
            </div>
        </div>
        <br>
        <button type="submit">Create Customer</button>
    </form>

    <hr>

    <!-- ======================================== -->
    <!-- Create New Opportunity -->
    <!-- ======================================== -->
    <h2>Create New Opportunity</h2>
    <form method="POST" action="{{ url_for('add_opportunity_route') }}">
        <div class="form-grid">
            <div class="form-col">
                <label>Customer:</label>
                <select id="customer_id" name="customer_id" required>
                    <option value="" disabled selected>Select a customer...</option>
                    {% for customer in customers %}
                        <option value="{{ customer.customer_ID }}">{{ customer.customer_name }}</option>
                    {% endfor %}
                </select>
                <br><br>

                <label>Opportunity Name:</label>
                <input type="text" name="opportunity_name" required><br><br>

                <label>Tax Rate (%):</label>
                <input type="number" name="tax_rate" id="tax_rate" step="0.01" min="0" value="0.00" required><br><br>

                <label>Tax Type:</label>
                <select name="tax_type" required>
                <option value="Standard" selected>Standard</option>
                <option value="Exempt">Exempt</option>
                <option value="New Construction">New Construction</option>
                </select>
                <br><br>
            </div>

            <div class="form-col">
                <label>Site Address:</label>
                <input type="text" name="site_address"><br><br>

                <label>Site City:</label>
                <input type="text" name="site_city"><br><br>

                <label>Site State:</label>
                <input type="text" name="site_state" maxlength="2" placeholder="TX"><br><br>

                <label>Site Zip:</label>
                <input type="text" name="site_zip" maxlength="10" placeholder="75001"><br><br>
            </div>
        </div>
        <br>
        <button type="submit">Create Opportunity</button>
    </form>

    <hr>

    <!-- ======================================== -->
    <!-- Existing Opportunities -->
    <!-- ======================================== -->
    <h2>Existing Opportunities</h2>
    <table id="opportunities_table" class="display">
        <thead>
            <tr>
                <th>Opportunity ID</th>
                <th>Opportunity Name</th>
                <th>Customer Name</th>
                <th>Customer Email</th>
            </tr>
            <tr>
                <th><input type="text" placeholder="Search ID" /></th>
                <th><input type="text" placeholder="Search Name" /></th>
                <th><input type="text" placeholder="Search Customer" /></th>
                <th><input type="text" placeholder="Search Email" /></th>
            </tr>
        </thead>
        <tbody>
            {% for opp in opportunities %}
            <tr>
                <td>{{ opp.opportunity_ID }}</td>
                <td>
                    <a href="{{ url_for('show_opportunity_route', opportunity_id=opp.opportunity_ID) }}">
                        {{ opp.opportunity_name }}
                    </a>
                </td>
                <td>
                    <a href="{{ url_for('customer_detail_route', customer_id=opp.customer_ID) }}">
                        {{ opp.customer_name }}
                    </a>
                </td>
                <td>{{ opp.customer_email }}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>

    <hr>

    <!-- ======================================== -->
    <!-- All Customers Table -->
    <!-- ======================================== -->
    <h2>All Customers</h2>
    <table id="customers_table" class="display">
        <thead>
            <tr>
                <th>Customer ID</th>
                <th>Customer Name</th>
                <th>Email</th>
                <th>City</th>
                <th>State</th>
                <th>Zip</th>
            </tr>
            <tr>
                <th><input type="text" placeholder="Search ID" /></th>
                <th><input type="text" placeholder="Search Name" /></th>
                <th><input type="text" placeholder="Search Email" /></th>
                <th><input type="text" placeholder="Search City" /></th>
                <th><input type="text" placeholder="Search State" /></th>
                <th><input type="text" placeholder="Search Zip" /></th>
            </tr>
        </thead>
        <tbody>
            {% for cust in customers %}
            <tr>
                <td>{{ cust.customer_ID }}</td>
                <td><a href="{{ url_for('customer_detail_route', customer_id=cust.customer_ID) }}">{{ cust.customer_name }}</a></td>
                <td>{{ cust.customer_email }}</td>
                <td>{{ cust.billing_city }}</td>
                <td>{{ cust.billing_state }}</td>
                <td>{{ cust.billing_zip }}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>

    <!-- ======================================== -->
    <!-- JS -->
    <!-- ======================================== -->
    <script src="https://cdn.jsdelivr.net/npm/tom-select/dist/js/tom-select.complete.min.js"></script>
    <script>
        new TomSelect("#customer_id", {
            create: false,
            sortField: {
                field: "text",
                direction: "asc"
            }
        });
    </script>

    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <script src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
    <script>
        $(document).ready(function() {
            // Initialize DataTables for Opportunities
            var oppTable = $('#opportunities_table').DataTable({
                paging: true,
                ordering: false
            });
            $('#opportunities_table thead tr:eq(1) th input').on('keyup change clear', function () {
                let colIndex = $(this).parent().index();
                if (oppTable.column(colIndex).search() !== this.value) {
                    oppTable.column(colIndex).search(this.value).draw();
                }
            });

            // Initialize DataTables for Customers
            var custTable = $('#customers_table').DataTable({
                paging: true,
                ordering: false
            });
            $('#customers_table thead tr:eq(1) th input').on('keyup change clear', function () {
                let colIndex = $(this).parent().index();
                if (custTable.column(colIndex).search() !== this.value) {
                    custTable.column(colIndex).search(this.value).draw();
                }
            });
        });
    </script>
</body>
</html>
"""

########################################################################################################################

CUSTOMER_DETAIL_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Customer Details</title>
    <link rel="stylesheet" href="https://cdn.datatables.net/1.13.7/css/jquery.dataTables.min.css">
    <style>
        .form-grid {
            display: flex;
            gap: 40px;
            align-items: flex-start;
        }
        .form-col { flex: 1; }
        label {
            display: inline-block;
            width: 140px;
            margin-bottom: 6px;
        }
        input {
            width: 200px;
        }
        table {
            border-collapse: collapse;
            margin-top: 20px;
            width: 100%;
        }
        table, th, td {
            border: 1px solid #333;
            padding: 5px;
        }
        h2 {
            margin-top: 40px;
        }

        /* Center DataTables search bars */
        div.dataTables_filter {
            text-align: center !important;
            float: none !important;
        }
        div.dataTables_filter label {
            width: 100%;
            display: flex;
            justify-content: center;
            align-items: center;
            gap: 8px;
        }
        div.dataTables_filter input {
            width: 250px;
        }

        /* Sequence column smaller */
        #customer_lineitems_table th:nth-child(1),
        #customer_lineitems_table td:nth-child(1) {
            width: 70px;
            text-align: center;
        }

        /* Popups */
        #popup, #componentsPopup {
            display: none;
            position: fixed;
            top: 10%;
            left: 10%;
            width: 80%;
            height: 70%;
            background: #fff;
            border: 2px solid #333;
            padding: 10px;
            z-index: 1000;
            box-shadow: 0 4px 8px rgba(0,0,0,0.3);
            overflow: auto;
        }
        #popup textarea { width: 100%; height: 80%; }
    </style>
</head>
<body>
    <h1>Customer: {{ customer.customer_name }}</h1>

    <!-- ============================= -->
    <!-- Customer Info Form -->
    <!-- ============================= -->
    <form method="POST">
      <div class="form-grid">
        <!-- Left Column -->
        <div class="form-col">
          <label>Customer Name:</label>
          <input type="text" name="customer_name" value="{{ customer.customer_name }}" required><br><br>

          <label>Customer Email:</label>
          <input type="email" name="customer_email" value="{{ customer.customer_email }}"><br><br>

          <label>Contact First Name:</label>
          <input type="text" name="contact_first_name" value="{{ customer.contact_first_name }}"><br><br>

          <label>Contact Last Name:</label>
          <input type="text" name="contact_last_name" value="{{ customer.contact_last_name }}"><br><br>

          <label>Phone:</label>
          <input type="tel" name="phone" value="{{ customer.phone }}"><br><br>
        </div>

        <!-- Right Column -->
        <div class="form-col">
          <label>Billing Address:</label>
          <input type="text" name="billing_address" value="{{ customer.billing_address }}"><br><br>

          <label>Billing City:</label>
          <input type="text" name="billing_city" value="{{ customer.billing_city }}"><br><br>

          <label>Billing State:</label>
          <input type="text" name="billing_state" maxlength="2" value="{{ customer.billing_state }}"><br><br>

          <label>Billing Zip:</label>
          <input type="text" name="billing_zip" maxlength="10" value="{{ customer.billing_zip }}"><br><br>
        </div>
      </div>

      <button type="submit">Update Customer</button>
    </form>

    <hr>

    <!-- ============================= -->
    <!-- Opportunities Table -->
    <!-- ============================= -->
    <h2>Opportunities for {{ customer.customer_name }}</h2>
    <table id="customer_opps_table">
        <thead>
            <tr>
                <th>ID</th>
                <th>Opportunity Name</th>
                <th>Opportunity Price</th>
                <th>Site Address</th>
                <th>City</th>
                <th>State</th>
                <th>Zip</th>
            </tr>
        </thead>
        <tbody>
            {% for opp in opportunities %}
            <tr>
                <td>{{ opp.opportunity_ID }}</td>
                <td>
                    <a href="{{ url_for('show_opportunity_route', opportunity_id=opp.opportunity_ID) }}">
                        {{ opp.opportunity_name }}
                    </a>
                </td>
                <td>${{ "%.2f"|format(opp.opportunity_price or 0) }}</td>
                <td>{{ opp.site_address }}</td>
                <td>{{ opp.site_city }}</td>
                <td>{{ opp.site_state }}</td>
                <td>{{ opp.site_zip }}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>

    <hr>

    <!-- ============================= -->
    <!-- Add New Saved Line Item -->
    <!-- ============================= -->
    <h2>Add New Saved Line Item</h2>
    <form method="POST" action="{{ url_for('add_customer_line_item', customer_id=customer.customer_ID) }}">
        <label>Description:</label><br>
        <textarea name="description" rows="2" cols="60" required></textarea>
        <br><br>
        <label>Quantity:</label>
        <input type="number" name="quantity" min="1" required>
        <br><br>
        <input type="submit" value="Add Saved Line Item">
    </form>

    <hr>

    <!-- ============================= -->
    <!-- Saved Line Items with Components -->
    <!-- ============================= -->
    <h2>Saved Line Items for {{ customer.customer_name }}</h2>
    <table id="customer_lineitems_table">
        <thead>
            <tr>
                <th>Sequence</th>
                <th>Description</th>
                <th>Quantity</th>
                <th>Unit Price</th>
                <th>Add Component</th>
                <th>Show Components</th>
                <th>Actions</th>
            </tr>
        </thead>
        <tbody>
            {% for li in customer_line_items %}
            <tr>
                <form method="post" action="{{ url_for('update_customer_line_item_and_components', customer_id=customer.customer_ID, line_id=li.line_ID) }}">
                    <td><input type="number" name="sequence_number" value="{{ li.line_item_sequence }}" style="width:60px;"></td>
                    <td>
                        <textarea id="desc_{{ li.line_ID }}" name="description" rows="2" cols="40">{{ li.line_item_description }}</textarea>
                        <button type="button" onclick="openPopup('desc_{{ li.line_ID }}')">Expand</button>
                    </td>
                    <td><input type="number" name="quantity" value="{{ li.quantity }}" min="1"></td>
                    <td>${{ "%.2f"|format(li.unit_price or 0) }}</td>
                    <td>
                        <select name="component_type_id">
                            <option value="">--Add component--</option>
                            {% for ct in component_types %}
                                {% if (ct.component_type_ID|int) not in [3, 8] %}
                                    <option value="{{ ct.component_type_ID }}">{{ ct.component_types_description }}</option>
                                {% endif %}
                            {% endfor %}
                        </select>
                    </td>
                        <td>
    {% if session.employee_type != "Sales" %}
        <button type="button"
            class="show-components-btn"
            data-line-id="{{ li.line_ID }}"
            data-components='{{ li.components|tojson | safe }}'>
            Show Components
        </button>
    {% else %}
        <span style="color:#999;">(Restricted)</span>
    {% endif %}
</td>
                    <td><input type="submit" value="Update / Add Component"></td>
                </form>
            </tr>
            {% endfor %}
        </tbody>
    </table>

    <br>
    <a href="{{ url_for('index') }}">⬅ Back to Opportunities</a>

    <!-- ============================= -->
    <!-- Popups -->
    <!-- ============================= -->
    <div id="popup">
        <h3>Edit Description</h3>
        <textarea id="popupText"></textarea><br>
        <button type="button" onclick="savePopup()">Save</button>
        <button type="button" onclick="closePopup()">Cancel</button>
    </div>

    <div id="componentsPopup">
        <h3>Components</h3>
        <div id="componentsTable"></div>
    </div>

    <!-- ============================= -->
    <!-- Scripts -->
    <!-- ============================= -->
    <script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
    <script src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>
    <script>
        $(document).ready(function() {
    $('#customer_opps_table').DataTable({
        paging: true,
        ordering: false,
        pageLength: 50
    });

    $('#customer_lineitems_table').DataTable({
        paging: true,
        ordering: false,
        pageLength: 50
    });
});

         const COMPONENT_TYPE_MAP = {
        {% for ct in component_types %}
            "{{ ct.component_type_ID }}": "{{ ct.component_types_description|e }}",
        {% endfor %}
    };

    let activeFieldId = null;
    let currentLineId = null;  // which line item’s popup is open

    // ==============================
    // Description Popup
    // ==============================
    function openPopup(fieldId) {
        const field = document.getElementById(fieldId);
        if (field && field.readOnly) return;
        activeFieldId = fieldId;
        document.getElementById('popupText').value = field ? field.value : "";
        document.getElementById('popup').style.display = 'block';
    }

    function savePopup() {
        if (activeFieldId) {
            document.getElementById(activeFieldId).value = document.getElementById('popupText').value;
        }
        closePopup();
    }

    function closePopup() {
        document.getElementById('popup').style.display = 'none';
        activeFieldId = null;
    }

    // ==============================
    // Show Components Popup
    // ==============================
    document.querySelectorAll(".show-components-btn").forEach(button => {
        button.addEventListener("click", () => {
            const raw = button.getAttribute("data-components") || "[]";
            const components = JSON.parse(raw);
            currentLineId = button.getAttribute("data-line-id");
            openComponentsPopup(components);
        });
    });

    function openComponentsPopup(components) {
    let html = "<table border='1' width='100%'>" +
               "<tr><th>Component Type</th><th>Quantity</th><th>Unit Cost</th><th>Unit Price</th><th>Subtotal</th><th>Actions</th></tr>";

    if (components && components.length > 0) {
        components.forEach(c => {
            // ✅ Correct single declaration
            const quoteUrl = `/quote_component/${c.component_ID}/${c.component_type_ID}?customer_id={{ customer.customer_ID }}`;
            const updateUrl = `/component/${c.component_ID}/update_quantity?line_id=${currentLineId}&show_popup=true`;

            const typeName = COMPONENT_TYPE_MAP[c.component_type_ID] ||
                             COMPONENT_TYPE_MAP[String(c.component_type_ID)] ||
                             "(Unknown Type)";
            const qty = parseFloat(c.quantity) || 0;
            const cost = parseFloat(c.unit_cost) || 0;
            const price = parseFloat(c.unit_price) || 0;
            const subtotal = qty * price;

            html += `<tr>
                        <td>${typeName}</td>
                        <td>
                            <form method="POST" action="${updateUrl}" class="component-update-form" style="display:inline;">
                                <input type="number" name="quantity" step="0.01" min="0" value="${qty}" style="width:80px;">
                                <input type="submit" value="Update">
                            </form>
                        </td>
                        <td>$${cost.toFixed(2)}</td>
                        <td>$${price.toFixed(2)}</td>
                        <td>$${subtotal.toFixed(2)}</td>
                        <td><a href="${quoteUrl}">Quote Component</a></td>
                     </tr>`;
        });
    } else {
        html += "<tr><td colspan='6'>No components found</td></tr>";
    }

    html += "</table>";
    html += `<br><button type="button" onclick="closeComponentsPopup()">Close</button>`;
    document.getElementById("componentsTable").innerHTML = html;
    document.getElementById("componentsPopup").style.display = "block";
}

    // ==============================
    // Quantity Update Logic (keeps popup open)
    // ==============================
    document.addEventListener("submit", function(e) {
        const form = e.target.closest(".component-update-form");
        if (!form) return;

        e.preventDefault();
        const fd = new FormData(form);

        fetch(form.action, { method: "POST", body: fd })
          .then(() => {
              // Recompute prices, then reopen popup
              return fetch(`/line_item/${currentLineId}/update_price_from_components`, { method: "POST" })
                     .catch(() => {});
          })
          .finally(() => {
              window.location.href = "{{ url_for('customer_detail_route', customer_id=customer.customer_ID) }}?show_popup=true&line_id=" + currentLineId;
          });
    });

    function closeComponentsPopup() {
        document.getElementById("componentsPopup").style.display = "none";

        if (currentLineId) {
            fetch(`/line_item/${currentLineId}/update_price_from_components`, {
                method: "POST"
            })
            .then(() => {
                window.location.href = "{{ url_for('customer_detail_route', customer_id=customer.customer_ID) }}?show_popup=true&line_id=" + currentLineId;
            })
            .catch(err => console.error("Failed to update line item price:", err));
        }
        currentLineId = null;
    }

    // ==============================
    // Auto-reopen popup on reload
    // ==============================
    window.onload = function() {
        {% if show_popup == "true" and popup_line_id %}
            const btn = document.querySelector(`.show-components-btn[data-line-id='{{ popup_line_id }}']`);
            if (btn) btn.click();
        {% endif %}
    }
</script>
<script>
/* ========= SAVE SCROLL POSITION ========= */
window.addEventListener("beforeunload", function () {
    localStorage.setItem("scrollPos", window.scrollY);
});

/* ========= RESTORE SCROLL POSITION ========= */
window.addEventListener("load", function () {
    const pos = localStorage.getItem("scrollPos");
    if (pos !== null) window.scrollTo(0, parseInt(pos));
});
</script>
</body>
</html>
"""

########################################################################################################################

LINE_ITEMS_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Line Items</title>
    <style>
        /* Popup styling */
        #popup, #componentsPopup, #installersPopup, #customerQuotesPopup {
            display: none;
            position: fixed;
            top: 10%;
            left: 10%;
            width: 80%;
            height: 70%;
            background: #fff;
            border: 2px solid #333;
            padding: 10px;
            z-index: 1000;
            box-shadow: 0px 4px 8px rgba(0,0,0,0.3);
            overflow: auto;
        }

        #popup textarea {
            width: 100%;
            height: 80%;
        }

        /* Grey out inactive rows */
        .inactive {
            background-color: #dcdcdc;
            color: #6b6b6b;
        }
        .inactive textarea, .inactive input[type="number"] {
            background-color: #eaeaea;
            color: #6b6b6b;
            pointer-events: none;
        }
        .inactive textarea[readonly], .inactive input[readonly] {
            pointer-events: auto;
        }
        .inactive input[type="checkbox"] {
            pointer-events: auto;
        }

        /* Form grid layout */
        .form-grid {
            display: flex;
            gap: 40px;
            align-items: flex-start;
            margin-bottom: 20px;
        }
        .form-col {
            flex: 1;
        }
        label {
            display: inline-block;
            width: 140px;
            margin-bottom: 6px;
        }
        input {
            width: 200px;
        }
        button[type="submit"], button[type="button"] {
            margin-top: 10px;
            padding: 6px 14px;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            margin-top: 10px;
        }
        th, td {
            padding: 5px;
            border: 1px solid black;
            text-align: left;
        }
    </style>
</head>

<body>
    <h1>Line Items for Opportunity {{ opportunity_id }}</h1>

    <!-- Edit Opportunity Info -->
    <h2>Opportunity Details</h2>
    <form method="POST" action="{{ url_for('update_opportunity_route', opportunity_id=opportunity_id) }}">
      <div class="form-grid">
        <div class="form-col">
          <label for="opportunity_name">Opportunity Name:</label>
          <input type="text" name="opportunity_name" value="{{ opportunity_name }}" required><br><br>

          <label for="tax_rate">Tax Rate (%):</label>
          <input type="number" name="tax_rate" step="0.01" min="0" value="{{ tax_rate }}"><br><br>
        </div>

        <div class="form-col">
          <label for="site_address">Site Address:</label>
          <input type="text" name="site_address" value="{{ site_address }}"><br><br>

          <label for="site_city">Site City:</label>
          <input type="text" name="site_city" value="{{ site_city }}"><br><br>

          <label for="site_state">Site State:</label>
          <input type="text" name="site_state" maxlength="2" placeholder="TX" value="{{ site_state }}"><br><br>

          <label for="site_zip">Site Zip:</label>
          <input type="text" name="site_zip" maxlength="10" placeholder="75001" value="{{ site_zip }}"><br><br>
        </div>
      </div>
      <button type="submit">Update Opportunity</button>
    </form>

    <hr>  
    <!-- Add New Line Item -->
    <h2>Add New Line Item</h2>
    <form method="post" action="{{ url_for('add_line_item_route', opportunity_id=opportunity_id) }}">
        <label>Description:</label><br>
        <textarea id="new_description" name="description" rows="2" cols="60" required></textarea>
        <button type="button" onclick="openPopup('new_description')">Expand</button>
        <br><br>
        <label>Quantity:</label>
        <input type="number" name="quantity" min="1" required>
        <br><br>
        <input type="submit" value="Add Line Item">
    </form>

    <hr>

    <!-- Standard Line Items -->
    <h2>Add Standard Line Item</h2>
    <form id="standardLineItemForm" method="post">
        <label for="standard_id">Select Standard Item:</label>
        <select id="standard_id" name="standard_id">
            {% for s in standard_line_items %}
            <option value="{{ s.ID }}">{{ s.line_item_description }} - ${{ "%.2f"|format(s.unit_price|float) }}</option>
            {% endfor %}
        </select>
        <button type="submit">Add to Opportunity</button>
    </form>

    <hr>

    <!-- Existing Line Items -->
    <h2>Existing Line Items</h2>
    <button type="button" onclick="openCustomerQuotesPopup()">➕ Add from Customer Quotes</button>

    <table>
        <tr>
            <th>Sequence</th>
            <th>Description</th>
            <th>Quantity</th>
            <th>Unit Price</th>
            <th>Active</th>
            <th>Add Component</th>
            <th>Show Components</th>
            <th>Actions</th>
        </tr>
        {% for item in items %}
        <tr class="{% if item.activation_status != 'ACTIVE' %}inactive{% endif %}">
            <form method="post" action="{{ url_for('update_line_item_and_components', opportunity_id=opportunity_id, line_id=item.line_ID) }}">
                <td>
                    <input type="number" name="sequence_number" value="{{ item.line_item_sequence }}" style="width:60px;"
                           {% if item.activation_status != 'ACTIVE' %}readonly{% endif %}>
                </td>
                <td>
                    <textarea id="desc_{{ item.line_ID }}" name="description" rows="2" cols="40"
                              {% if item.activation_status != 'ACTIVE' %}readonly{% endif %}>{{ item.line_item_description }}</textarea>
                    <button type="button" onclick="openPopup('desc_{{ item.line_ID }}')"
                            {% if item.activation_status != 'ACTIVE' %}disabled{% endif %}>Expand</button>
                </td>
                <td>
                    <input type="number" name="quantity" value="{{ item.quantity }}" min="1"
                           {% if item.activation_status != 'ACTIVE' %}readonly{% endif %}>
                </td>
                <td>${{ "%.2f"|format(item.unit_price) }}</td>
                <td style="text-align:center;">
                    <input type="checkbox" name="activation_status" value="ACTIVE"
                           {% if item.activation_status == 'ACTIVE' %}checked{% endif %}>
                </td>
                <td>
                    <select name="component_type_id">
                        <option value="">--Add component--</option>
                        {% for ct in component_types %}
                            <option value="{{ ct.component_type_ID }}">{{ ct.component_types_description }}</option>
                        {% endfor %}
                    </select>
                </td>

                <!-- HIDE FOR SALES -->
                <td>
                {% if user_type != "Sales" %}
                    <button type="button"
                            class="show-components-btn"
                            data-line-id="{{ item.line_ID }}"
                            data-components='{{ item.components|tojson | safe }}'>
                        Show Components
                    </button>
                {% else %}
                    <span style="color:#888;">(Restricted)</span>
                {% endif %}
                </td>

                <td><input type="submit" value="Update / Add Component"></td>
            </form>
        </tr>
        {% endfor %}
    </table>

    <br>
    <a href="{{ url_for('index') }}">⬅ Back to Opportunities</a>
    <br><br>
    <a href="{{ url_for('contract_route', opportunity_id=opportunity_id) }}" target="_blank">🖨️ Print Contract</a>
    <br><br>
    <button type="button" onclick="findInstallers()">Find Installers</button>

    <!-- Popups -->
    <div id="popup">
        <h3>Edit Description</h3>
        <textarea id="popupText"></textarea>
        <br>
        <button type="button" onclick="savePopup()">Save</button>
        <button type="button" onclick="closePopup()">Cancel</button>
    </div>

    <div id="componentsPopup">
        <h3>Components</h3>
        <div id="componentsTable"></div>
    </div>

    <div id="customerQuotesPopup">
        <h3>Select Customer Quote to Add</h3>
        <div id="customerQuotesTable">Loading...</div>
        <br>
        <button type="button" onclick="closeCustomerQuotesPopup()">Close</button>
    </div>

    <div id="installersPopup">
        <h3>Available Installers</h3>
        <div id="installersTable"></div>
        <br>
        <button type="button" onclick="closeInstallersPopup()">Close</button>
    </div>

<script>
let activeFieldId = null;
let currentLineId = null;

// ===========================
// Add Standard Line Item
// ===========================
document.getElementById("standardLineItemForm").addEventListener("submit", async (e) => {
    e.preventDefault();
    const id = document.getElementById("standard_id").value;
    const res = await fetch(`/add_standard_line_item/{{ opportunity_id }}/${id}`, { method: "POST" });
    const data = await res.json();
    if (data.success) {
        window.location.reload();
    } else {
        alert("Error: " + data.message);
    }
});

// ===========================
// Description popup
// ===========================
function openPopup(fieldId) {
    const field = document.getElementById(fieldId);
    if (field.readOnly) return;
    activeFieldId = fieldId;
    document.getElementById('popupText').value = field.value;
    document.getElementById('popup').style.display = 'block';
}
function savePopup() {
    if (activeFieldId) {
        document.getElementById(activeFieldId).value = document.getElementById('popupText').value;
    }
    closePopup();
}
function closePopup() {
    document.getElementById('popup').style.display = 'none';
    activeFieldId = null;
}

// ===========================
// Components popup logic
// ===========================
document.querySelectorAll(".show-components-btn").forEach(button => {
    button.addEventListener("click", () => {

        const userType = "{{ user_type|e }}";
        if (userType === "Sales") {
            alert("Access restricted. Sales cannot view components.");
            return;
        }

        const raw = button.getAttribute("data-components") || "[]";
        const components = JSON.parse(raw);
        currentLineId = button.getAttribute("data-line-id");
        openComponentsPopup(components);
    });
});

function openComponentsPopup(components) {
    let html = "<table><tr><th>Description</th><th>Quantity</th><th>Unit Cost</th><th>Unit Price</th><th>Subtotal</th><th>Actions</th></tr>";

    if (components && components.length > 0) {
        components.forEach(c => {
            const quoteUrl = `/quote_component/${c.component_ID}/${c.component_type_ID}`;
            const updateUrl = `/component/${c.component_ID}/update_quantity?line_id=${currentLineId}&show_popup=true`;

            const qty = parseFloat(c.quantity) || 0;
            const price = parseFloat(c.unit_price) || 0;
            const subtotal = qty * price;

            html += `<tr>
                        <td>${c.description || ""}</td>
                        <td>
                            <form method="POST" action="${updateUrl}" class="component-update-form" style="display:inline;">
                                <input type="number" name="quantity" step="0.01" min="0" value="${qty}" style="width:80px;">
                                <input type="submit" value="Update">
                            </form>
                        </td>
                        <td>${c.unit_cost ? "$" + parseFloat(c.unit_cost).toFixed(2) : ""}</td>
                        <td>${price ? "$" + price.toFixed(2) : ""}</td>
                        <td>${subtotal ? "$" + subtotal.toFixed(2) : "$0.00"}</td>
                        <td><a href="${quoteUrl}">Quote Component</a></td>
                     </tr>`;
        });
    } else {
        html += "<tr><td colspan='6'>No components found</td></tr>";
    }

    html += "</table><br><button type='button' onclick='closeComponentsPopup()'>Close</button>";
    document.getElementById("componentsTable").innerHTML = html;
    document.getElementById("componentsPopup").style.display = "block";
}

function closeComponentsPopup() {
    document.getElementById("componentsPopup").style.display = "none";
    const cleanUrl = window.location.origin + window.location.pathname;
    window.history.replaceState({}, document.title, cleanUrl);
    window.location.reload();
}

// ===========================
// Auto reopen popup after back
// ===========================
window.onload = function() {
    {% if show_popup == "true" and popup_line_id %}
        const btn = document.querySelector(`.show-components-btn[data-line-id='{{ popup_line_id }}']`);
        if (btn) btn.click();
    {% endif %}
}

// ===========================
// Customer Quotes Popup
// ===========================
function openCustomerQuotesPopup() {
    const popup = document.getElementById("customerQuotesPopup");
    popup.style.display = "block";
    document.getElementById("customerQuotesTable").innerHTML = "Loading...";

    fetch(`/get_customer_quotes/{{ opportunity_id }}`)
      .then(response => response.json())
      .then(data => {
          if (!data.length) {
              document.getElementById("customerQuotesTable").innerHTML = "<p>No customer quotes found.</p>";
              return;
          }
          let html = "<table><tr><th>Description</th><th>Quantity</th><th>Unit Price</th><th>Action</th></tr>";
          data.forEach(q => {
              html += `
                <tr>
                  <td>${q.line_item_description}</td>
                  <td>${q.quantity}</td>
                  <td>$${parseFloat(q.unit_price || 0).toFixed(2)}</td>
                  <td><button type='button' onclick='addCustomerQuoteToOpportunity(${q.line_ID})'>Add to Opportunity</button></td>
                </tr>`;
          });
          html += "</table>";
          document.getElementById("customerQuotesTable").innerHTML = html;
      })
      .catch(err => {
          console.error("Failed to load customer quotes:", err);
          document.getElementById("customerQuotesTable").innerHTML = "<p style='color:red;'>Error loading quotes.</p>";
      });
}

function closeCustomerQuotesPopup() {
    document.getElementById("customerQuotesPopup").style.display = "none";
}

function addCustomerQuoteToOpportunity(customerLineID) {
    fetch(`/add_customer_quote_to_opportunity/{{ opportunity_id }}/${customerLineID}`, { method: "POST" })
      .then(r => r.json())
      .then(resp => {
          if (resp.success) {
              alert("Quote added successfully!");
              window.location.reload();
          } else {
              alert("Failed: " + (resp.message || "Unknown error"));
          }
      })
      .catch(err => {
          console.error("Error adding quote:", err);
          alert("Failed to add quote.");
      });
}

// ===========================
// Installers Popup
// ===========================
function findInstallers() {
  fetch(`/get_install_vendors/{{ opportunity_id }}`)
    .then(response => response.json())
    .then(data => {
      let html = "<table><tr><th>Vendor</th><th>Email</th><th>City</th><th>State</th><th>Zip</th><th>Preferred</th></tr>";
      if (data.length > 0) {
        data.forEach(v => {
          html += `<tr>
                     <td>${v.vendor_name}</td>
                     <td>${v.email}</td>
                     <td>${v.city || ""}</td>
                     <td>${v.state}</td>
                     <td>${v.zip}</td>
                     <td>${v.preferred ? "✅ Yes" : "❌ No"}</td>
                   </tr>`;
        });
      } else {
        html += "<tr><td colspan='6'>No installers found</td></tr>";
      }
      html += "</table>";
      document.getElementById("installersTable").innerHTML = html;
      document.getElementById("installersPopup").style.display = "block";
    })
    .catch(err => {
      console.error("Error fetching installers:", err);
      alert("Failed to load installers.");
    });
}

function closeInstallersPopup() {
  document.getElementById("installersPopup").style.display = "none";
}
</script>

<script>
// ===============================
// Save & Restore Scroll Position
// ===============================
document.addEventListener("DOMContentLoaded", function () {
    const savedPos = localStorage.getItem("scrollPos_" + window.location.pathname);
    if (savedPos) {
        window.scrollTo(0, parseInt(savedPos));
    }
});

window.addEventListener("beforeunload", function () {
    localStorage.setItem("scrollPos_" + window.location.pathname, window.scrollY);
});
</script>

</body>
</html>
"""

########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################

QUOTE_CUSTOM_SIGN = """
<!DOCTYPE html>
<html>
<head>
    <title>Add Materials & Labor</title>
    <link href="https://cdn.jsdelivr.net/npm/tom-select/dist/css/tom-select.css" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/tom-select/dist/js/tom-select.complete.min.js"></script>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        h1 { text-align: center; }
        .top-section { display: flex; justify-content: space-between; align-items: flex-start; gap: 50px; }
        .section { width: 48%; }
        input[name="quantity"], input[name="material_price"] { width: 100px; }
        table { border-collapse: collapse; width: 100%; margin-top: 10px; }
        table, th, td { border: 1px solid black; }
        th, td { padding: 5px; text-align: left; }
        th { background-color: #f3f3f3; }
        .tables-section { display: flex; justify-content: space-between; align-items: flex-start; gap: 40px; margin-top: 30px; }
        .tables-section .table-container { width: 48%; }
        input[type="number"] { text-align: right; }
        a { font-weight: bold; color: #0044cc; text-decoration: none; }
    </style>
</head>
<body>

<h1>Add Materials & Labor to "{{ line_item_description }}"</h1>
<hr>

{% set ns = namespace(material_cost=0.0, labor_cost=0.0) %}

<div class="top-section">
    <!-- Add Material Section -->
    <div class="section">
        <h2>Pick Stock Material</h2>
        <form method="post" action="{{ url_for('add_component_material', component_id=component_id) }}">
            <!-- Preserve IDs -->
            <input type="hidden" name="customer_id" value="{{ customer_id }}">
            <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

            <label>Material:</label>
            <select id="material_id" name="material_id">
                {% for m in materials %}
                <option value="{{ m.material_ID }}">
                    {{ m.material_description }} ({{ m.material_unit }}) - ${{ "%.2f"|format(m.material_price|float) }}
                </option>
                {% endfor %}
            </select>
            <br><br>
            <label>Quantity:</label>
            <input type="number" name="quantity" step="0.01" min="0.01" required>
            <br><br>
            <input type="submit" value="Add Material">
        </form>

        <hr>

        <h2>Add Non-Stock Material</h2>
        <form method="post" action="{{ url_for('add_nonstock_component_material', component_id=component_id) }}">
            <!-- Preserve IDs -->
            <input type="hidden" name="customer_id" value="{{ customer_id }}">
            <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

            <label>Description:</label>
            <input type="text" name="material_description" required>
            <label>Unit:</label>
            <input type="text" name="material_unit" required>
            <br><br>
            <label>Price:</label>
            <input type="number" step="0.01" name="material_price" required>
            <label>Quantity:</label>
            <input type="number" name="quantity" step="0.01" min="0.01" required>
            <br><br>
            <input type="submit" value="Add Non-Stock Material">
        </form>

        <hr>
    </div>

    <!-- Add Labor Section -->
    <div class="section">
        <h2>Add Labor</h2>
        <form method="post" action="{{ url_for('add_component_labor', component_id=component_id) }}">
            <!-- Preserve IDs -->
            <input type="hidden" name="customer_id" value="{{ customer_id }}">
            <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

            <label>Labor Type:</label>
            <select name="labor_id">
                {% for l in labor_types %}
                <option value="{{ l.labor_ID }}">{{ l.labor_type }} - ${{ "%.2f"|format(l.burden_rate|float) }}/unit</option>
                {% endfor %}
            </select>
            <br><br>
            <label>Quantity:</label>
            <input type="number" name="quantity" step="0.01" min="0.01" required>
            <br><br>
            <input type="submit" value="Add Labor">
        </form>

        <hr>
    </div>
</div>

<form method="post" action="{{ url_for('update_component_quantities', component_id=component_id, customer_id=customer_id, component_type_id=component_type_id, hide_back_button=hide_back_button) }}">
  <div class="tables-section">
    <!-- Existing Materials -->
    <div class="table-container">
      <h2>Existing Materials</h2>
      <table>
        <tr>
          <th>Material</th>
          <th>Quantity</th>
          <th>Unit</th>
          <th>Unit Price</th>
          <th>Total</th>
        </tr>
        {% for cm in component_materials %}
          {% set qty = (cm.quantity | float) %}
          {% set price = (cm.material_price | float) %}
          {% set row_total = qty * price %}
          {% set ns.material_cost = ns.material_cost + row_total %}
          <tr>
            <td>{{ cm.material_description }}</td>
            <td>
              <input type="hidden" name="material_row_id[]" value="{{ cm.ID }}">
              <input type="number" step="0.01" name="material_qty[]" value="{{ qty }}" style="width:80px;">
            </td>
            <td>{{ cm.material_unit }}</td>
            <td>${{ "%.2f"|format(price) }}</td>
            <td>${{ "%.2f"|format(row_total) }}</td>
          </tr>
        {% endfor %}
      </table>
    </div>

    <!-- Existing Labor -->
    <div class="table-container">
      <h2>Existing Labor</h2>
      <table>
        <tr>
          <th>Labor Type</th>
          <th>Quantity</th>
          <th>Burden Rate</th>
          <th>Total</th>
        </tr>
        {% for cl in component_labor %}
          {% set qty = (cl.quantity | float) %}
          {% set rate = (cl.burden_rate | float) %}
          {% set row_total = qty * rate %}
          {% set ns.labor_cost = ns.labor_cost + row_total %}
          <tr>
            <td>{{ cl.labor_type }}</td>
            <td>
              <input type="hidden" name="labor_row_id[]" value="{{ cl.line_item_labor_ID }}">
              <input type="number" step="0.01" name="labor_qty[]" value="{{ qty }}" style="width:80px;">
            </td>
            <td>${{ "%.2f"|format(rate) }}</td>
            <td>${{ "%.2f"|format(row_total) }}</td>
          </tr>
        {% endfor %}
      </table>
    </div>
  </div>

  <br>
  <input type="submit" value="Save Quantity Updates">
</form>

<hr>

<!-- Summary Section -->
<div style="margin-top:30px;">
    <h2>Summary</h2>
    <table>
        <tr>
            <th>Total Materials Cost</th>
            <th>Total Labor Cost</th>
            <th>Unit Cost</th>
            <th>Unit Price</th>
        </tr>
        <tr>
            <td>${{ "%.2f"|format(ns.material_cost|float) }}</td>
            <td>${{ "%.2f"|format(ns.labor_cost|float) }}</td>
            <td>${{ "%.2f"|format(component_unit_cost|float) }}</td>
            <td>${{ "%.2f"|format(component_unit_price|float) }}</td>
        </tr>
    </table>
</div>

<!-- Back Navigation (Customer / Opportunity / Home Safe) -->
{% set has_customer = customer_id is defined and customer_id not in (None, "None", "", 0) %}
{% set has_opportunity = opportunity_id is defined and opportunity_id not in (None, "None", "", 0) %}

{% if has_customer %}
    <br>
    <a href="{{ url_for('customer_detail_route', customer_id=customer_id|int) }}">
        ⬅ Back to Customer Details
    </a>

{% elif has_opportunity %}
    <br>
    {% set opp_id_safe = opportunity_id|int %}
    <a id="backToLineItems"
       href="{{ url_for('show_opportunity_route', opportunity_id=opp_id_safe) }}?show_popup=true&line_id={{ line_id }}">
        ⬅ Back to Line Items
    </a>

    <script>
      document.addEventListener("DOMContentLoaded", () => {
          if (document.getElementById("material_id")) {
              new TomSelect("#material_id", {
                  create: false,
                  sortField: { field: "text", direction: "asc" }
              });
          }

          const btn = document.getElementById("backToLineItems");
          if (btn) {
              btn.addEventListener("click", async (e) => {
                  e.preventDefault();
                  try {
                      await fetch(`/line_item/{{ line_id }}/update_price_from_components`, { method: "POST" });
                  } catch (err) {
                      console.warn("Update failed:", err);
                  }
                  window.location.href = "{{ url_for('show_opportunity_route', opportunity_id=opp_id_safe) }}?show_popup=true&line_id={{ line_id }}";
              });
          }
      });
    </script>

{% else %}
    <br>
    <a href="{{ url_for('index') }}">
        ⬅ Back to Home
    </a>
{% endif %}

</body>
</html>
"""

########################################################################################################################

QUOTE_INSTALLATION = """
<!DOCTYPE html>
<html>
<head>
    <title>Installation Quote</title>
    <style>
        body { font-family: Arial, sans-serif; }
        .top-section { display: flex; justify-content: space-between; align-items: flex-start; gap: 50px; }
        .section { width: 48%; }
        input[name="quantity"], input[name="unit_price"] { width: 100px; }
        table { border-collapse: collapse; width: 100%; margin-top: 10px; }
        table, th, td { border: 1px solid black; }
        th, td { padding: 5px; text-align: left; }
        .tables-section { display: flex; justify-content: space-between; align-items: flex-start; gap: 40px; margin-top: 30px; }
        .tables-section .table-container { width: 48%; }
        input[type="number"] { text-align: right; width: 80px; }
        h2 { margin-bottom: 10px; }
    </style>
</head>
<body>
<h1>Add Installation Materials & Labor to "{{ line_item_description }}"</h1>
<hr>

{% set ns = namespace(install_material_cost=0.0, install_labor_cost=0.0, subcontract_cost=0.0) %}

<div class="top-section">

    <!-- Add Install Material -->
    <div class="section">
        <h2>Add Installation Material</h2>
        <form method="post" action="{{ url_for('add_install_material', component_id=component_id) }}">
            <input type="hidden" name="customer_id" value="{{ customer_id }}">
            <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

            <label>Description:</label>
            <input type="text" name="material_description" required>
            <label>Unit:</label>
            <input type="text" name="material_unit" required><br><br>

            <label>Unit Cost:</label>
            <input type="number" step="0.01" name="unit_cost" required>
            <label>Quantity:</label>
            <input type="number" name="quantity" step="0.01" min="0.01" required><br><br>

            <input type="submit" value="Add Installation Material">
        </form>

        <hr>

        <!-- Add Subcontract Install Cost -->
        <h2>Add Subcontractor Install Cost</h2>
        <form method="post" action="{{ url_for('add_sub_install_cost', component_id=component_id) }}">
            <input type="hidden" name="customer_id" value="{{ customer_id }}">
            <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

            <label>Subcontractor Cost:</label>
            <input type="number" name="subcontractor_cost" step="0.01" min="0.00" required>

            <br><br>
            <input type="submit" value="Add Subcontract Cost">
        </form>
    </div>

    <!-- Add Install Labor -->
    <div class="section">
        <h2>Add Installation Labor</h2>
        <form method="post" action="{{ url_for('add_install_labor', component_id=component_id) }}">
            <input type="hidden" name="customer_id" value="{{ customer_id }}">
            <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

            <label>Labor Type:</label>
            <select name="install_labor_id">
                {% for l in install_labor_types %}
                <option value="{{ l.install_labor_ID }}">{{ l.install_labor_type }} - ${{ "%.2f"|format(l.burden_rate|float) }}/unit</option>
                {% endfor %}
            </select><br><br>

            <label>Quantity:</label>
            <input type="number" name="quantity" step="0.01" min="0.01" required><br><br>

            <input type="submit" value="Add Installation Labor">
        </form>
    </div>
</div>

<hr>

<form method="post" action="{{ url_for('update_install_quantities', component_id=component_id) }}">
    <input type="hidden" name="customer_id" value="{{ customer_id }}">
    <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

    <div class="tables-section">

        <!-- Existing Installation Materials -->
        <div class="table-container">
            <h2>Existing Installation Materials</h2>
            <table>
                <tr>
                    <th>Description</th>
                    <th>Unit</th>
                    <th>Quantity</th>
                    <th>Unit Cost</th>
                    <th>Total Cost</th>
                </tr>
                {% for m in component_install_materials %}
                    {% set qty = m.quantity | float %}
                    {% set price = m.unit_cost | float %}
                    {% set row_total = qty * price %}
                    {% set ns.install_material_cost = ns.install_material_cost + row_total %}
                <tr>
                    <td>{{ m.material_description }}</td>
                    <td>{{ m.material_unit }}</td>
                    <td>
                        <input type="hidden" name="install_material_row_id[]" value="{{ m.ID }}">
                        <input type="number" step="0.01" name="install_material_qty[]" value="{{ qty }}">
                    </td>
                    <td>${{ "%.2f"|format(price) }}</td>
                    <td>${{ "%.2f"|format(row_total) }}</td>
                </tr>
                {% endfor %}
            </table>
        </div>

        <!-- Existing Installation Labor -->
        <div class="table-container">
            <h2>Existing Installation Labor</h2>
            <table>
                <tr>
                    <th>Labor Type</th>
                    <th>Quantity</th>
                    <th>Burden Rate</th>
                    <th>Total Cost</th>
                </tr>
                {% for l in component_install_labor %}
                    {% set qty = l.quantity | float %}
                    {% set rate = l.burden_rate | float %}
                    {% set row_total = qty * rate %}
                    {% set ns.install_labor_cost = ns.install_labor_cost + row_total %}
                <tr>
                    <td>{{ l.install_labor_type }}</td>
                    <td>
                        <input type="hidden" name="install_labor_row_id[]" value="{{ l.ID }}">
                        <input type="number" step="0.01" name="install_labor_qty[]" value="{{ qty }}">
                    </td>
                    <td>${{ "%.2f"|format(rate) }}</td>
                    <td>${{ "%.2f"|format(row_total) }}</td>
                </tr>
                {% endfor %}
            </table>
        </div>

    </div>

    <br>
    <input type="submit" value="Save Quantity Updates">
</form>

<hr>

<div class="table-container">
    <h2>Subcontractor Install Costs</h2>
    <table>
        <tr>
            <th>Subcontractor Cost</th>
        </tr>
        {% for s in subcontract_costs %}
            {% set ns.subcontract_cost = ns.subcontract_cost + (s.subcontractor_cost | float) %}
        <tr>
            <td>${{ "%.2f"|format(s.subcontractor_cost) }}</td>
        </tr>
        {% endfor %}
    </table>
</div>

<hr>

<div style="margin-top:30px;">
    <h2>Summary</h2>
    <table>
        <tr>
            <th>Total Install Materials</th>
            <th>Total Install Labor</th>
            <th>Total Subcontract</th>
            <th>Unit Cost</th>
            <th>Unit Price</th>
        </tr>
        <tr>
            <td>${{ "%.2f"|format(ns.install_material_cost) }}</td>
            <td>${{ "%.2f"|format(ns.install_labor_cost) }}</td>
            <td>${{ "%.2f"|format(ns.subcontract_cost) }}</td>
            <td>${{ "%.2f"|format(component_unit_cost) }}</td>
            <td>${{ "%.2f"|format(component_unit_price) }}</td>
        </tr>
    </table>
</div>

<br>
{% if opportunity_id and opportunity_id != 0 %}
    <a href="{{ url_for('show_opportunity_route', opportunity_id=opportunity_id) }}?show_popup=true&line_id={{ line_id }}">
        ⬅ Back to Line Items
    </a>
{% elif customer_id and customer_id != 0 %}
    <a href="{{ url_for('customer_detail_route', customer_id=customer_id) }}">
        ⬅ Back to Customer Page
    </a>
{% endif %}

</body>
</html>
"""

########################################################################################################################

QUOTE_EMC = """
<!DOCTYPE html>
<html>
<head>
    <title>EMC Quote</title>
    <style>
        body { font-family: Arial, sans-serif; }
        table { border-collapse: collapse; width: 100%; margin-top: 10px; }
        table, th, td { border: 1px solid black; }
        th, td { padding: 5px; text-align: left; }
        a { text-decoration: none; color: #007BFF; font-weight: bold; }
        a:hover { text-decoration: underline; }
    </style>
</head>
<body>
<h1>Add EMC (Electronic Message Center) to "{{ line_item_description }}"</h1>

<hr>

<form method="post" action="{{ url_for('add_emc_unit', component_id=component_id) }}">
    <input type="hidden" name="customer_id" value="{{ customer_id }}">
    <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

    <label>Description:</label>
    <input type="text" name="EMC_description" required>
    <label>Unit Cost:</label>
    <input type="number" step="0.01" name="unit_cost" required>
    <label>Quantity:</label>
    <input type="number" step="0.01" name="quantity" required>
    <br><br>
    <input type="submit" value="Add EMC Unit">
</form>

<hr>

<h2>Existing EMC Units</h2>
<table>
    <tr>
        <th>Description</th>
        <th>Quantity</th>
        <th>Unit Cost</th>
        <th>Total Cost</th>
    </tr>
    {% set ns = namespace(emc_total=0.0) %}
    {% for e in component_emc %}
        {% set qty = e.quantity | float %}
        {% set cost = e.unit_cost | float %}
        {% set row_total = qty * cost %}
        {% set ns.emc_total = ns.emc_total + row_total %}
    <tr>
        <td>{{ e.EMC_description }}</td>
        <td>{{ qty }}</td>
        <td>${{ "%.2f"|format(cost) }}</td>
        <td>${{ "%.2f"|format(row_total) }}</td>
    </tr>
    {% endfor %}
</table>

<hr>

<h2>Summary</h2>
<table>
    <tr>
        <th>Total EMC Cost</th>
        <th>Unit Cost</th>
        <th>Unit Price</th>
    </tr>
    <tr>
        <td>${{ "%.2f"|format(ns.emc_total) }}</td>
        <td>${{ "%.2f"|format(component_unit_cost) }}</td>
        <td>${{ "%.2f"|format(component_unit_price) }}</td>
    </tr>
</table>

<br>
{% if opportunity_id and opportunity_id != 0 %}
    <a href="{{ url_for('show_opportunity_route', opportunity_id=opportunity_id) }}?show_popup=true&line_id={{ line_id }}">
        ⬅ Back to Line Items
    </a>
{% elif customer_id and customer_id != 0 %}
    <a href="{{ url_for('customer_detail_route', customer_id=customer_id) }}">
        ⬅ Back to Customer Page
    </a>
{% endif %}
</body>
</html>
"""

########################################################################################################################

QUOTE_PIPE_FOUNDATIONS = """
<!DOCTYPE html>
<html>
<head>
    <title>Pipe and Foundations Quote</title>
    <style>
        body { font-family: Arial, sans-serif; }
        form label { display: inline-block; width: 250px; margin-top: 6px; }
        form input { width: 150px; margin-bottom: 6px; }
        table { border-collapse: collapse; width: 100%; margin-top: 20px; }
        table, th, td { border: 1px solid black; }
        th, td { padding: 5px; text-align: left; }
        .form-section { margin-bottom: 30px; padding: 10px; border: 1px solid #ccc; }
        h2 { margin-top: 30px; }
        .cabinet-row { margin-bottom: 8px; }
        .cabinet-row label { margin-right: 6px; }
        .small-input { width: 70px; margin-right: 12px; }
    </style>
</head>
<body>
<h1>Pipe and Foundations for "{{ line_item_description }}"</h1>
<hr>

<form method="post" action="{{ url_for('save_pipe_foundation_factors', component_id=component_id) }}">

    <div class="form-section">
        <h2>General</h2>
        <label>Overall Sign Height:</label>
        <input type="number" step="0.01" name="overall_height" value="{{ factor1 or 0 }}"><br>
        <label>Head Cabinet Height:</label>
        <input type="number" step="0.01" name="head_cabinet_height" value="{{ factor2 or 0 }}">
        <label>Head Cabinet Width:</label>
        <input type="number" step="0.01" name="head_cabinet_width" value="{{ factor3 or 0 }}"><br>
        <label>Wind Speed Requirement:</label>
        <input type="number" step="1" name="wind_speed" value="{{ factor4 or 0 }}">
        <label>Exposure Type:</label>
        <select name="exposure_type">
            <option value="1" {% if factor5 == 1 %}selected{% endif %}>A</option>
            <option value="2" {% if factor5 == 2 %}selected{% endif %}>B</option>
            <option value="3" {% if factor5 == 3 %}selected{% endif %}>C</option>
            <option value="4" {% if factor5 == 4 %}selected{% endif %}>D</option>
        </select><br>
        <label>Number of Pipes:</label>
        <input type="number" name="num_pipes" value="{{ factor6 or 0 }}">
        <label>Pipe Yield Strength (ksi):</label>
        <input type="number" step="0.01" name="pipe_yield_strength" value="{{ factor7 or 0 }}">
    </div>

    <div class="form-section">
        <h2>Cabinets</h2>
        <div class="cabinet-row">
            <label>Cabinet 2 Max Height:</label>
            <input type="number" step="0.01" name="cab2_max_height" value="{{ factor8 or 0 }}" class="small-input">
            <label>Cabinet 2 Height:</label>
            <input type="number" step="0.01" name="cab2_height" value="{{ factor9 or 0 }}" class="small-input">
            <label>Cabinet 2 Width:</label>
            <input type="number" step="0.01" name="cab2_width" value="{{ factor10 or 0 }}" class="small-input">
        </div>

        <div class="cabinet-row">
            <label>Cabinet 3 Max Height:</label>
            <input type="number" step="0.01" name="cab3_max_height" value="{{ factor11 or 0 }}" class="small-input">
            <label>Cabinet 3 Height:</label>
            <input type="number" step="0.01" name="cab3_height" value="{{ factor12 or 0 }}" class="small-input">
            <label>Cabinet 3 Width:</label>
            <input type="number" step="0.01" name="cab3_width" value="{{ factor13 or 0 }}" class="small-input">
        </div>

        <div class="cabinet-row">
            <label>Cabinet 4 Max Height:</label>
            <input type="number" step="0.01" name="cab4_max_height" value="{{ factor14 or 0 }}" class="small-input">
            <label>Cabinet 4 Height:</label>
            <input type="number" step="0.01" name="cab4_height" value="{{ factor15 or 0 }}" class="small-input">
            <label>Cabinet 4 Width:</label>
            <input type="number" step="0.01" name="cab4_width" value="{{ factor16 or 0 }}" class="small-input">
        </div>
    </div>

    <div class="form-section">
        <h2>Pipe Transitions</h2>
        <label>Pipe 1 Transition Height:</label>
        <input type="number" step="0.01" name="pipe1_transition_height" value="{{ factor17 or 0 }}">
        <label>Pipe 2 Transition Height:</label>
        <input type="number" step="0.01" name="pipe2_transition_height" value="{{ factor18 or 0 }}"><br>
        <label>Pipe 3 Transition Height:</label>
        <input type="number" step="0.01" name="pipe3_transition_height" value="{{ factor19 or 0 }}">
        <label>Pipe 4 Transition Height:</label>
        <input type="number" step="0.01" name="pipe4_transition_height" value="{{ factor20 or 0 }}">
    </div>

    <div class="form-section">
        <h2>Foundation</h2>
        <label>Foundation Type:</label>
        <select name="foundation_type">
            <option value="1" {% if factor21 == 1 %}selected{% endif %}>Pier Footer</option>
            <option value="2" {% if factor21 == 2 %}selected{% endif %}>Rectangular Pier Footer</option>
            <option value="3" {% if factor21 == 3 %}selected{% endif %}>Spread Footer</option>
        </select>
        <label>Rectangular Footer Length:</label>
        <input type="number" step="0.01" name="rect_footer_length" value="{{ factor22 or 0 }}"><br>
        <label>Rectangular Footer Width:</label>
        <input type="number" step="0.01" name="rect_footer_width" value="{{ factor23 or 0 }}">
    </div>

    <br>
    <input type="submit" value="Save Parameters">
</form>

<br>
<a href="{{ url_for('show_opportunity_route', opportunity_id=opportunity_id) }}?show_popup=true&line_id={{ line_id }}">
    ⬅ Back to Line Items
</a>
</body>
</html>
"""

########################################################################################################################

QUOTE_PIPE_FOUNDATION_OPTIONS = """
<!DOCTYPE html>
<html>
<head>
    <title>Select Pipe & Foundation</title>
    <style>
        body { font-family: Arial, sans-serif; }
        table { border-collapse: collapse; width: 100%; margin-top: 20px; }
        table, th, td { border: 1px solid black; }
        th, td { padding: 5px; text-align: left; }
    </style>
</head>
<body>
<h1>Select Foundation Option</h1>
<form method="post" action="{{ url_for('save_pipe_foundation_choice', component_id=component_id) }}">
    <table>
        <tr>
            <th>Select</th>
            <th>Pier Depth</th>
            <th>Pier Diameter</th>
        </tr>
        {% for opt in options %}
        <tr>
            <td><input type="radio" name="choice" value="{{ opt.depth }}|{{ opt.diameter }}" required></td>
            <td>{{ opt.depth }}</td>
            <td>{{ opt.diameter }}</td>
        </tr>
        {% endfor %}
    </table>
    <br>
    <input type="submit" value="Save Selection">
</form>
</body>
</html>
"""

########################################################################################################################

QUOTE_PIPE_FOUNDATION_COSTS = """
<!DOCTYPE html>
<html>
<head>
    <title>Pipe & Foundation Costs</title>
    <style>
        body { font-family: Arial, sans-serif; }
        form label { display: inline-block; width: 250px; margin-top: 6px; }
        form input { width: 150px; margin-bottom: 6px; }
        table { border-collapse: collapse; margin-top: 20px; }
        table, th, td { border: 1px solid black; padding: 6px; }
        th { background-color: #f0f0f0; }
    </style>
</head>
<body>
<h1>Pipe & Foundation Costs</h1>

{% if pier_diameter != 0 %}
<p><strong>Pier Diameter:</strong> {{ pier_diameter }} ft</p>
{% endif %}
{% if pier_depth != 0 %}
<p><strong>Pier Depth:</strong> {{ pier_depth }} ft</p>
{% endif %}
{% if pier_quantity != 0 %}
<p><strong>Number of Piers:</strong> {{ pier_quantity }}</p>
{% endif %}

{% if rectangular_footer_length != 0 %}
<p><strong>Rectangular Footer Length:</strong> {{ rectangular_footer_length }} ft</p>
{% endif %}
{% if rectangular_footer_width != 0 %}
<p><strong>Rectangular Footer Width:</strong> {{ rectangular_footer_width }} ft</p>
{% endif %}
{% if rectangular_footer_depth != 0 %}
<p><strong>Rectangular Footer Depth:</strong> {{ rectangular_footer_depth }} ft</p>
{% endif %}

<p><strong>Base Pipe Diameter:</strong> {{ base_pipe_diameter }}</p>
<p><strong>Base Pipe Required Footage:</strong> {{ base_pipe_footage }} ft</p>

{% if stack_pipe1_diameter != "0" %}
<p><strong>Stack Pipe 1 Diameter:</strong> {{ stack_pipe1_diameter }} ft</p>
{% endif %}
{% if stack_pipe1_footage != 0 %}
<p><strong>Stack Pipe 1 Required Footage:</strong> {{ stack_pipe1_footage }} ft</p>
{% endif %}

{% if stack_pipe2_diameter != "0" %}
<p><strong>Stack Pipe 2 Diameter:</strong> {{ stack_pipe2_diameter }} ft</p>
{% endif %}
{% if stack_pipe2_footage != 0 %}
<p><strong>Stack Pipe 2 Required Footage:</strong> {{ stack_pipe2_footage }} ft</p>
{% endif %}

{% if stack_pipe3_diameter != "0" %}
<p><strong>Stack Pipe 3 Diameter:</strong> {{ stack_pipe3_diameter }} ft</p>
{% endif %}
{% if stack_pipe3_footage != 0 %}
<p><strong>Stack Pipe 3 Required Footage:</strong> {{ stack_pipe3_footage }} ft</p>
{% endif %}

{% if stack_pipe4_diameter != "0" %}
<p><strong>Stack Pipe 4 Diameter:</strong> {{ stack_pipe4_diameter }} ft</p>
{% endif %}
{% if stack_pipe4_footage != 0 %}
<p><strong>Stack Pipe 4 Required Footage:</strong> {{ stack_pipe4_footage }} ft</p>
{% endif %}

<form method="post">
    <label>Digging Cost:</label>
    <input type="number" step="0.01" name="digging_cost" value="{{ digging_cost }}"><br>

    <label>Concrete Cost:</label>
    <input type="number" step="0.01" name="concrete_cost" value="{{ concrete_cost }}"><br>

    <label>Additional Footer Cost:</label>
    <input type="number" step="0.01" name="additional_footer_cost" value="{{ additional_footer_cost }}"><br>

    <label>Pipe Cost:</label>
    <input type="number" step="0.01" name="pipe_cost" value="{{ pipe_cost }}"><br>

    <br>
    <input type="submit" value="Save Costs">
</form>

{% if unit_cost is defined and unit_price is defined %}
<h2>Calculated Totals</h2>
<table>
    <tr>
        <th>Unit Cost</th>
        <th>Unit Price</th>
    </tr>
    <tr>
        <td>${{ "%.2f"|format(unit_cost) }}</td>
        <td>${{ "%.2f"|format(unit_price) }}</td>
    </tr>
</table>
{% endif %}

</body>
</html>
"""

########################################################################################################################

QUOTE_MASONRY = """
<!DOCTYPE html>
<html>
<head>
    <title>Masonry Quote</title>
    <style>
        body { font-family: Arial, sans-serif; }
        table { border-collapse: collapse; width: 100%; margin-top: 10px; }
        table, th, td { border: 1px solid black; }
        th, td { padding: 5px; text-align: left; }
    </style>
</head>
<body>
<h1>Add Masonry to "{{ line_item_description }}"</h1>

<hr>

<form method="post" action="{{ url_for('add_masonry', component_id=component_id) }}">
    <label>Description:</label>
    <input type="text" name="masonry_description" required>
    <label>Unit Cost:</label>
    <input type="number" step="0.01" name="unit_cost" required>
    <label>Quantity:</label>
    <input type="number" step="0.01" name="quantity" required>
    <br><br>
    <input type="submit" value="Add Masonry">
</form>

<hr>

<h2>Existing Masonry</h2>
<table>
    <tr>
        <th>Description</th>
        <th>Quantity</th>
        <th>Unit Cost</th>
        <th>Total Cost</th>
    </tr>
    {% set ns = namespace(masonry_total=0.0) %}
    {% for m in component_masonry %}
        {% set qty = m.quantity | float %}
        {% set cost = m.unit_cost | float %}
        {% set row_total = qty * cost %}
        {% set ns.masonry_total = ns.masonry_total + row_total %}
    <tr>
        <td>{{ m.masonry_description }}</td>
        <td>{{ qty }}</td>
        <td>${{ "%.2f"|format(cost) }}</td>
        <td>${{ "%.2f"|format(row_total) }}</td>
    </tr>
    {% endfor %}
</table>

<hr>

<h2>Summary</h2>
<table>
    <tr>
        <th>Total Masonry Cost</th>
        <th>Unit Cost</th>
        <th>Unit Price</th>
    </tr>
    <tr>
        <td>${{ "%.2f"|format(ns.masonry_total) }}</td>
        <td>${{ "%.2f"|format(component_unit_cost) }}</td>
        <td>${{ "%.2f"|format(component_unit_price) }}</td>
    </tr>
</table>

<br>
<a href="{{ url_for('show_opportunity_route', opportunity_id=opportunity_id) }}?show_popup=true&line_id={{ line_id }}">
    ⬅ Back to Line Items
</a>
</body>
</html>
"""

########################################################################################################################

QUOTE_RENTAL_EQUIPMENT = """
<!DOCTYPE html>
<html>
<head>
    <title>Rental Equipment Quote</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            margin: 20px;
        }
        h1, h2 {
            text-align: center;
        }
        form {
            display: flex;
            flex-wrap: wrap;
            gap: 15px;
            align-items: center;
            justify-content: center;
            margin-bottom: 15px;
        }
        label {
            font-weight: bold;
            margin-right: 5px;
        }
        input[type="text"], input[type="number"] {
            padding: 5px;
            width: 150px;
        }
        table {
            border-collapse: collapse;
            width: 100%;
            margin-top: 10px;
        }
        th, td {
            border: 1px solid black;
            padding: 8px;
            text-align: center;
        }
        th {
            background-color: #f2f2f2;
        }
        hr {
            margin: 25px 0;
        }
        .summary-table {
            width: 50%;
            margin: 0 auto;
        }
        .back-link {
            display: block;
            text-align: center;
            margin-top: 20px;
        }
    </style>
</head>
<body>
<h1>Add Rental Equipment to "{{ line_item_description }}"</h1>

<hr>

<form method="post" action="{{ url_for('add_rental_equipment', component_id=component_id) }}">
    <input type="hidden" name="customer_id" value="{{ customer_id }}">
    <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

    <label>Description:</label>
    <input type="text" name="equipment_description" required>

    <label>Unit Cost ($):</label>
    <input type="number" step="0.01" name="unit_cost" required>

    <label>Quantity:</label>
    <input type="number" step="0.01" name="quantity" required>

    <input type="submit" value="Add Rental Equipment">
</form>

<hr>

<h2>Existing Rental Equipment</h2>
<table>
    <tr>
        <th>Description</th>
        <th>Quantity</th>
        <th>Unit Cost ($)</th>
        <th>Total Cost ($)</th>
    </tr>
    {% set ns = namespace(equipment_total=0.0) %}
    {% for re in component_rental_equipment %}
        {% set qty = re.quantity | float %}
        {% set cost = re.unit_cost | float %}
        {% set row_total = qty * cost %}
        {% set ns.equipment_total = ns.equipment_total + row_total %}
        <tr>
            <td>{{ re.equipment_description }}</td>
            <td>{{ "%.2f"|format(qty) }}</td>
            <td>{{ "%.2f"|format(cost) }}</td>
            <td>{{ "%.2f"|format(row_total) }}</td>
        </tr>
    {% endfor %}
</table>

<hr>

<h2>Summary</h2>
<table class="summary-table">
    <tr>
        <th>Total Equipment Cost</th>
        <th>Component Unit Cost</th>
        <th>Component Unit Price</th>
    </tr>
    <tr>
        <td>${{ "%.2f"|format(ns.equipment_total) }}</td>
        <td>${{ "%.2f"|format(component_unit_cost) }}</td>
        <td>${{ "%.2f"|format(component_unit_price) }}</td>
    </tr>
</table>

{% if customer_id and customer_id|int > 0 %}
    <a class="back-link" href="{{ url_for('customer_detail_route', customer_id=customer_id) }}?show_popup=true&line_id={{ line_id }}">
        ⬅ Back to Customer
    </a>
{% else %}
    <a class="back-link" href="{{ url_for('show_opportunity_route', opportunity_id=opportunity_id) }}?show_popup=true&line_id={{ line_id }}">
        ⬅ Back to Line Items
    </a>
{% endif %}

</body>
</html>
"""

########################################################################################################################

QUOTE_FACE_LIT_CHANNEL_LETTERS_COMPONENT = """
<!DOCTYPE html>
<html>
<head>
  <title>Face Lit Channel Letters Component</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 20px; }
    table { border-collapse: collapse; width: 80%; margin-top: 16px; }
    th, td { border: 1px solid #333; padding: 8px; text-align: left; }
    .form-section { margin-top: 20px; padding: 12px; border: 1px solid #ccc; border-radius: 6px; }
    label { display: inline-block; width: 180px; margin-top: 6px; }
    select, input[type="number"], input[type="file"] { padding: 4px; }
    button { margin-top: 16px; padding: 8px 14px; }
    .back-link { display: block; text-align: center; margin-top: 25px; }
  </style>
</head>
<body>
  <h1>Face Lit Channel Letters Component for "{{ line_item_description }}"</h1>

  <h2>Excel Data</h2>
  <table>
    <tr><th>Modules</th><th>Areas</th><th>Perimeters</th></tr>
    {% for mod, area, peri in excel_rows %}
      {% if (mod|string != "0" and mod not in (None, "", 0)) 
            or (area|string != "0" and area not in (None, "", 0)) 
            or (peri|string != "0" and peri not in (None, "", 0)) %}
        <tr>
          <td>{{ mod }}</td>
          <td>{{ area }}</td>
          <td>{{ peri }}</td>
        </tr>
      {% endif %}
    {% endfor %}
  </table>

  <hr>

  <h2>Upload Excel & Design Selections</h2>
  <form action="{{ url_for('quote_component', component_id=component_id, component_type_id=5) }}" 
        method="post" enctype="multipart/form-data">

    <!-- Hidden IDs -->
    <input type="hidden" name="customer_id" value="{{ customer_id }}">
    <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

    <!-- Excel upload -->
    <div class="form-section">
      <label>Excel File:</label>
      <input type="file" name="excel_file" accept=".xlsx,.xls">
    </div>

    <div class="form-section">
      <!-- Letter Backs -->
      <label>Letter Backs:</label>
      <select name="factor1">
        <option value="1" {% if factor1 == 1 %}selected{% endif %}>ACM</option>
        <option value="2" {% if factor1 == 2 %}selected{% endif %}>Aluminum</option>
      </select><br>

      <!-- Returns -->
      <label>Returns:</label>
      <select name="factor2">
        <option value="1" {% if factor2 == 1 %}selected{% endif %}>5"</option>
        <option value="2" {% if factor2 == 2 %}selected{% endif %}>3"</option>
      </select><br>

      <!-- Face Material -->
      <label>Face Material:</label>
      <select name="factor3">
        <option value="1" {% if factor3 == 1 %}selected{% endif %}>White Acrylic</option>
        <option value="2" {% if factor3 == 2 %}selected{% endif %}>Clear Acrylic</option>
        <option value="3" {% if factor3 == 3 %}selected{% endif %}>White Poly</option>
        <option value="4" {% if factor3 == 4 %}selected{% endif %}>Clear Poly</option>
      </select><br>

      <!-- Paint -->
      <label>Paint:</label>
      <select name="factor4">
        <option value="1" {% if factor4 == 1 %}selected{% endif %}>Returns</option>
        <option value="2" {% if factor4 == 2 %}selected{% endif %}>Trimcap</option>
        <option value="3" {% if factor4 == 3 %}selected{% endif %}>Returns and Trimcap</option>
        <option value="4" {% if factor4 == 4 %}selected{% endif %}>None</option>
      </select><br>

      <!-- Raceway -->
      <label>Raceway:</label>
      <select name="factor5">
        <option value="1" {% if factor5 == 1 %}selected{% endif %}>None</option>
        <option value="2" {% if factor5 == 2 %}selected{% endif %}>5"</option>
        <option value="3" {% if factor5 == 3 %}selected{% endif %}>7"</option>
      </select><br>

      <!-- Trimcap -->
      <label>Trimcap:</label>
      <select name="factor6">
        <option value="1" {% if factor6 == 1 %}selected{% endif %}>Trimcap</option>
        <option value="2" {% if factor6 == 2 %}selected{% endif %}>Retainers</option>
      </select><br>

      <!-- Graphics -->
      <label>Graphics:</label>
      <select name="factor7">
        <option value="1" {% if factor7 == 1 %}selected{% endif %}>Direct Print</option>
        <option value="2" {% if factor7 == 2 %}selected{% endif %}>Standard Color Vinyl</option>
        <option value="3" {% if factor7 == 3 %}selected{% endif %}>DayNight Vinyl</option>
        <option value="4" {% if factor7 == 4 %}selected{% endif %}>Printed Vinyl</option>
        <option value="5" {% if factor7 == 5 %}selected{% endif %}>Printed DayNight Vinyl</option>
        <option value="6" {% if factor7 == 6 %}selected{% endif %}>None</option>
      </select><br>

      <!-- Number of colors -->
      <label>Number of Printed Colors:</label>
      <input type="number" name="factor8" min="0" value="{{ factor8 or '' }}">
    </div>

    <button type="submit">Upload & Save</button>
  </form>

  <form action="{{ url_for('quote_component', component_id=component_id, component_type_id=99) }}" method="get">
    <input type="hidden" name="customer_id" value="{{ customer_id }}">
    <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">
    <button type="submit">Go to Quoting Screen</button>
  </form>

{% set has_customer = customer_id is defined and customer_id not in (None, "None", "") %}
{% set has_opportunity = opportunity_id is defined and opportunity_id not in (None, "None", "") %}

{% if has_customer %}
  <a href="{{ url_for('customer_detail_route', customer_id=customer_id|int) }}?show_popup=true&line_id={{ line_id }}">
    ⬅ Back to Customer
  </a>

{% elif has_opportunity %}
  {% set opp_id_safe = opportunity_id|int %}
  <a href="{{ url_for('show_opportunity_route', opportunity_id=opp_id_safe) }}?show_popup=true&line_id={{ line_id }}">
    ⬅ Back to Line Items
  </a>

{% else %}
  <a href="{{ url_for('index') }}">⬅ Back to Home</a>
{% endif %}
</body>
</html>
"""

########################################################################################################################

QUOTE_REVERSE_LIT_CHANNEL_LETTERS_COMPONENT = """
<!DOCTYPE html>
<html>
<head>
  <title>Reverse Lit Channel Letters Component</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 20px; }
    table { border-collapse: collapse; width: 80%; margin-top: 16px; }
    th, td { border: 1px solid #333; padding: 8px; text-align: left; }
    .form-section { margin-top: 20px; padding: 12px; border: 1px solid #ccc; border-radius: 6px; }
    label { display: inline-block; width: 180px; margin-top: 6px; }
    select, input[type="number"], input[type="file"] { padding: 4px; }
    button { margin-top: 16px; padding: 8px 14px; }
  </style>
</head>
<body>
  <h1>Reverse Lit Channel Letters Component for "{{ line_item_description }}"</h1>

  <h2>Excel Data</h2>
  <table>
    <tr><th>Modules</th><th>Areas</th><th>Perimeters</th></tr>
    {% for mod, area, peri in excel_rows %}
      {% if mod|float != 0 or area|float != 0 or peri|float != 0 %}
        <tr>
          <td>{{ mod }}</td>
          <td>{{ area }}</td>
          <td>{{ peri }}</td>
        </tr>
      {% endif %}
    {% endfor %}
  </table>

  <hr>

  <h2>Upload Excel & Design Selections</h2>
  <form action="{{ url_for('quote_component', component_id=component_id, component_type_id=6) }}" 
        method="post" enctype="multipart/form-data">

    <!-- Hidden IDs -->
    <input type="hidden" name="customer_id" value="{{ customer_id }}">
    <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">

    <!-- Excel upload -->
    <div class="form-section">
      <label>Excel File:</label>
      <input type="file" name="excel_file" accept=".xlsx,.xls">
    </div>

    <div class="form-section">
      <!-- Letter Backs -->
      <label>Letter Backs:</label>
      <select name="factor1">
        <option value="1" {% if factor1 == 1 %}selected{% endif %}>Clear Poly</option>
        <option value="2" {% if factor1 == 2 %}selected{% endif %}>White Poly</option>
        <option value="3" {% if factor1 == 3 %}selected{% endif %}>Clear Acrylic</option>
        <option value="4" {% if factor1 == 4 %}selected{% endif %}>White Acrylic</option>
      </select><br>

      <!-- Face Graphics -->
      <label>Face Graphics:</label>
      <select name="factor2">
        <option value="1" {% if factor2 == 1 %}selected{% endif %}>None</option>
        <option value="2" {% if factor2 == 2 %}selected{% endif %}>Opaque Vinyl</option>
        <option value="3" {% if factor2 == 3 %}selected{% endif %}>Printed Opaque Vinyl</option>
        <option value="4" {% if factor2 == 4 %}selected{% endif %}>Direct Print</option>
      </select><br>

      <!-- Back Graphics -->
      <label>Back Graphics:</label>
      <select name="factor3">
        <option value="1" {% if factor3 == 1 %}selected{% endif %}>None</option>
        <option value="2" {% if factor3 == 2 %}selected{% endif %}>Translucent Vinyl</option>
        <option value="3" {% if factor3 == 3 %}selected{% endif %}>Printed Translucent Vinyl</option>
        <option value="4" {% if factor3 == 4 %}selected{% endif %}>Direct Print</option>
      </select><br>

      <!-- Raceway -->
      <label>Raceway:</label>
      <select name="factor4">
        <option value="1" {% if factor4 == 1 %}selected{% endif %}>None</option>
        <option value="2" {% if factor4 == 2 %}selected{% endif %}>5"</option>
        <option value="3" {% if factor4 == 3 %}selected{% endif %}>7"</option>
      </select><br>

      <!-- Number of colors -->
      <label>Number of Printed Colors:</label>
      <input type="number" name="factor8" min="0" value="{{ factor8 or '' }}">
    </div>

    <button type="submit">Upload & Save</button>
  </form>

  <form action="{{ url_for('quote_component', component_id=component_id, component_type_id=99) }}" method="get">
    <input type="hidden" name="customer_id" value="{{ customer_id }}">
    <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">
    <button type="submit">Go to Quoting Screen</button>
  </form>

{% set has_customer = customer_id is defined and customer_id not in (None, "None", "") %}
{% set has_opportunity = opportunity_id is defined and opportunity_id not in (None, "None", "") %}

{% if has_customer %}
  <a href="{{ url_for('customer_detail_route', customer_id=customer_id|int) }}?show_popup=true&line_id={{ line_id }}">
    ⬅ Back to Customer
  </a>

{% elif has_opportunity %}
  {% set opp_id_safe = opportunity_id|int %}
  <a href="{{ url_for('show_opportunity_route', opportunity_id=opp_id_safe) }}?show_popup=true&line_id={{ line_id }}">
    ⬅ Back to Line Items
  </a>

{% else %}
  <a href="{{ url_for('index') }}">⬅ Back to Home</a>
{% endif %}
</body>
</html>
"""

########################################################################################################################

QUOTE_MANUAL_PRICE_ENTRY = """
<!DOCTYPE html>
<html>
<head>
  <title>Manual Price Entry</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 20px; }
    h1 { text-align: center; }
    form { width: 400px; margin: 0 auto; text-align: center; }
    label { display: inline-block; width: 150px; text-align: right; margin-right: 10px; }
    input[type="number"] { width: 120px; text-align: right; }
    button { margin-top: 20px; padding: 8px 16px; }
    a { display: block; margin-top: 20px; text-align: center; font-weight: bold; color: #0044cc; text-decoration: none; }
  </style>
</head>
<body>
  <h1>Manual Price Entry for "{{ line_item_description }}"</h1>

  <form method="post" action="{{ url_for('quote_component', component_id=component_id, component_type_id=10) }}">
    <label for="unit_price">Unit Price ($):</label>
    <input type="number" id="unit_price" name="unit_price" step="0.01" min="0" value="{{ component_unit_price or '' }}" required>
    <br><br>
    <button type="submit">Save Price</button>

    <input type="hidden" name="customer_id" value="{{ customer_id }}">
    <input type="hidden" name="opportunity_id" value="{{ opportunity_id }}">
  </form>

  {% set has_customer = customer_id is defined and customer_id not in (None, "None", "", 0) %}
  {% set has_opportunity = opportunity_id is defined and opportunity_id not in (None, "None", "", 0) %}

  {% if has_customer %}
    <a href="{{ url_for('customer_detail_route', customer_id=customer_id|int) }}?show_popup=true&line_id={{ line_id }}">
      ⬅ Back to Customer
    </a>
  {% elif has_opportunity %}
    <a href="{{ url_for('show_opportunity_route', opportunity_id=opportunity_id|int) }}?show_popup=true&line_id={{ line_id }}">
      ⬅ Back to Line Items
    </a>
  {% else %}
    <a href="{{ url_for('index') }}">⬅ Back to Home</a>
  {% endif %}
</body>
</html>
"""

########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################

CONTRACT_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Contract - Opportunity {{ opportunity.opportunity_name }}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        h1, h2 { text-align: center; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid black; padding: 8px; text-align: left; }
        th { background-color: #f2f2f2; }

        .totals-table {
            width: 40%;
            float: right;
            margin-top: 30px;
            font-size: 1.2em;
            border: 2px solid black;
        }
        .totals-table th, .totals-table td {
            border: 1px solid black;
            padding: 10px;
        }
        .totals-table th {
            background-color: #e6e6e6;
            text-align: left;
        }
        .totals-table td {
            text-align: right;
        }
        .grand-total {
            font-size: 1.4em;
            font-weight: bold;
            background-color: #d9edf7;
        }

        /* ✅ Print-safe dark headers */
        .black-header {
            background-color: black !important;
            color: white !important;
            text-align: center;
            -webkit-print-color-adjust: exact !important;
            print-color-adjust: exact !important;
        }

        @media print {
            body {
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
            }
        }

        /* Smaller text for contract terms */
        .contract-terms h4 {
            font-size: 0.7em;
            margin-bottom: 4px;
        }
        .contract-terms p, .contract-terms li {
            font-size: 0.5em;
            line-height: 1.4;
        }

    </style>
</head>
<body>
    <!-- Header with larger logo and perfectly centered titles -->
    <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 20px;">
        <!-- Left: Logo -->
        {% if company_logo %}
            <div style="flex: 1; display: flex; justify-content: flex-start;">
                <img src="data:image/jpeg;base64,{{ company_logo }}" 
                     alt="Company Logo"
                     style="height: 120px; width: auto;">
            </div>
        {% else %}
            <div style="flex: 1;"></div>
        {% endif %}

        <!-- Center: Titles (stays centered regardless of logo) -->
        <div style="flex: 1; text-align: center;">
            <h1 style="margin: 0;">Contract</h1>
            <h2 style="margin-top: 5px;">Opportunity: {{ opportunity.opportunity_name }}</h2>
        </div>

        <!-- Right: Empty spacer for centering symmetry -->
        <div style="flex: 1;"></div>
    </div>

    <hr>

    <!-- Project & Customer Info side-by-side -->
    <div style="display: flex; justify-content: space-between; align-items: flex-start; gap: 40px; margin-top: 20px;">
        <!-- Left: Project Information -->
        <div style="flex: 1; line-height: 1.6; width: 50%;">
            <h3 style="background-color: #f2f2f2; padding: 6px; border: 1px solid #ccc;">Project Information</h3>
            <p><strong>Site Address:</strong> {{ opportunity.site_address }}</p>
            <p><strong>City/State/ZIP:</strong> {{ opportunity.site_city }}, {{ opportunity.site_state }} {{ opportunity.site_zip }}</p>
            <p><strong>Contact Name:</strong> {{ opportunity.contact_first_name }} {{ opportunity.contact_last_name }}</p>
            <p><strong>Contact Phone:</strong> {{ opportunity.contact_phone }}</p>
        </div>

        <!-- Right: Customer Information -->
        <div style="flex: 1; line-height: 1.6; width: 50%;">
            <h3 style="background-color: #f2f2f2; padding: 6px; border: 1px solid #ccc;">Customer Information</h3>
            <p><strong>Customer:</strong> {{ opportunity.customer_name }}</p>
            <p><strong>Email:</strong> {{ opportunity.customer_email }}</p>
            <p><strong>Billing Address:</strong> {{ opportunity.billing_address }}</p>
            <p><strong>City/State/ZIP:</strong> {{ opportunity.billing_city }}, {{ opportunity.billing_state }} {{ opportunity.billing_zip }}</p>
        </div>
    </div>

    <hr>

    {% set ns = namespace(total=0.0) %}

{% set ns_row = namespace(num=1) %}

<h3>Line Items</h3>
<table style="width: 100%; border-collapse: collapse; table-layout: fixed;">
    <tr>
        <th style="width: 5%; text-align: center;">#</th>
        <th style="width: 55%; text-align: left;">Description</th>
        <th style="width: 10%; text-align: center;">Quantity</th>
        <th style="width: 17.5%; text-align: right;">Unit Price</th>
        <th style="width: 12.5%; text-align: right;">Subtotal</th>
    </tr>

    {% for item in line_items %}
        {% if item["activation_status"] == "ACTIVE" %}
            {% set qty = item["quantity"] | float(0) %}
            {% set price = item["unit_price"] | float(0) %}
            {% set row_subtotal = qty * price %}
            {% set ns.total = ns.total + row_subtotal %}

            <tr>
                <td style="text-align: center;">{{ ns_row.num }}</td>
                <td style="white-space: pre-wrap; text-align: left;">{{ item["line_item_description"] }}</td>
                <td style="text-align: center;">{{ qty }}</td>
                <td style="text-align: right;">${{ "{:,.2f}".format(price) }}</td>
                <td style="text-align: right;">${{ "{:,.2f}".format(row_subtotal) }}</td>
            </tr>

            {% set ns_row.num = ns_row.num + 1 %}
        {% endif %}
    {% endfor %}
</table>


<hr>

{% set tax_rate = opportunity.tax_rate | float(0) %}
{% set tax_amount = ns.total * (tax_rate / 100) %}
{% set grand_total = ns.total + tax_amount %}

<!-- ✅ Totals Table with wrapper and enforced spacing -->
<div class="totals-wrapper" style="width: 100%; overflow: hidden; margin-top: 20px;">
    <div class="totals" style="float: right; width: 40%;">
        <table class="totals-table" style="margin-bottom: 40px; white-space: nowrap;">
            <tr><td><strong>Subtotal:</strong></td><td>${{ "{:,.2f}".format(ns.total) }}</td>
            <tr><td><strong>Tax Rate:</strong></td><td>{{ '%.2f'|format(opportunity.tax_rate) }}%</td></tr>
            <tr><td><strong>Tax:</strong></td><td>${{ "{:,.2f}".format(tax_amount) }}</td></tr>
            <tr class="grand-total"><td><strong>Grand Total:</strong></td><td><strong>${{ "{:,.2f}".format(grand_total) }}</strong></td></tr>

            {% set deposit = grand_total / 2 %}
            {% set final_balance = grand_total / 2 %}
            <tr><td><strong>Deposit (50%):</strong></td><td>${{ "{:,.2f}".format(deposit) }}</td></tr>
            <tr><td><strong>Final Balance (50%):</strong></td><td>${{ "{:,.2f}".format(final_balance) }}</td></tr>
        </table>
    </div>
</div>

<!-- ✅ This ensures space below the floated table in all views -->
<div style="clear: both; height: 1px;"></div>

<!-- Signature Section -->
<div class="signature-section" style="margin-top: 40px;">
    <h3 class="black-header" style="padding: 8px;">
        THIS AGREEMENT IS ACCEPTED AND APPROVED BY
    </h3>
    <p style="text-align: center; margin: 10px 0;">
        By signing below, Customer accepts Company's proposal for the Project and agrees to 
        all of the terms and conditions stated on the Contract.
    </p>
    <br>

    <div style="display: flex; justify-content: space-between; margin-top: 40px;">
        <div style="width: 45%; border-top: 1px solid black; padding-top: 5px;">
            <br>
            <strong>By:</strong> ___________________________<br>
            <br>
            <strong>Date:</strong> _________________________
        </div>
        <div style="width: 45%; border-top: 1px solid black; padding-top: 5px; text-align: right;">
            <br>
            <strong>By:</strong> Michael Taylor / For: FSG - Signs<br>
            <br>
            <strong>Date:</strong> _________________________
        </div>
    </div>
</div>

<!-- Contract Terms -->
<div style="margin-top: 60px;">
    <h3 class="black-header" style="padding: 10px; font-size: 1.3em;">
        CONTRACT TERMS
    </h3>

    <div class="contract-terms" style="font-size: 0.75em; line-height: 1.4;">
        <ul style="list-style-type: disc; padding-left: 25px;">
            <li>Please make checks payable to FSG-Signs. Terms are Due Upon Completion, unless otherwise agreed.</li>
            <li>Sales tax calculated at 8.25% unless otherwise stated. Subject to change based on jurisdiction and will reflect on invoice.</li>
            <li>This proposal may be withdrawn if not accepted within 14 days due to fluctuating steel, material and fuel costs.</li>
            <li>Customer is to furnish all primary electrical service (120V UNLESS OTHERWISE AGREED) and connection to the sign BASE including: timers, photocells, switches, and/or other controls required by local city ordinances at Customers own expense.</li>
            <li>Installation portion of this estimate is based on adequate access to front and backside of the install area. Unforeseen obstacles may require additional charges.</li>
            <li>All private lines must be clearly marked by the customer (such as sprinkler systems and ground lighting). Any damage to private lines not clearly marked is the responsibility of the customer.</li>
            <li>Projects that are "NEW CONSTRUCTION" are taxed on the cost of materials only. Taxes are charged and itemized as a pass-through item to the customer. The final invoice is the controlling element of this contract (labor and materials separated on invoice).</li>
            <li>All shipping quotes expire after 60 days. Any price differences billed on final invoice.</li>
            <li>FSG imposes a 3% surcharge on credit cards, barring state laws, that is not greater than our cost of acceptance.</li>
        </ul>
        <hr>

            <p style="margin-top: 20px;">
                Company and Customer enter into the following customer contract (“Contract”) regarding services provided for the Job Number identified above and more specifically described on the reverse side of this Contract (the “Project”) and agree to the following terms and conditions regarding such Project:
            </p>

            <hr>

            <h4>STANDARD SPECIFICATIONS</h4>
            <p>The Project shall be completed in accordance with the sign drawing and elevation specifications corresponding to the Job Number listed above which are approved by Customer (“Standard Specifications”), unless changes to the Standard Specifications are approved by the parties in writing in accordance with the Change Order process described below.</p>

            <hr>

            <h4>CONTRACT AMOUNT</h4>
            <p>Customer shall pay Company for the Project as invoiced by Company in the amount and in increments listed on the reverse side of this Contract (“Contract Amount”). Time is of the essence with regard to Customer’s payment obligation.</p>

            <hr>

            <h4>ADDITIONAL WORK</h4>
            <p>Unless stated as part of the Contract Amount on the reverse side of this Contract, Customer shall pay an additional amount for the Project in the event that: (i) abnormal soil conditions or underground obstructions exist, including, without limitation, existence of solid rock, pipes, underground wires, etc.; (ii) Company must perform services related to obtaining a variance; (iii) Company is required to provide documentation to obtain permits and approvals for the Project other than the Standard Specifications described above, including, without limitation, shop drawings, samples, design layouts and modifications to architectural site plans; (iv) Company is required to remove free-standing signs or prior signs on a structure located on or near the installation site; (v) Company must obtain permits or approvals; or (vi) Company is requested or required to do any other additional work related to the Project that is not described in the Services section on the reverse side of this Contract.</p>

            <hr>

            <h4>TAXES</h4>
            <p>Customer agrees to pay all taxes that are due or may become due by Customer or that may be levied upon Company in connection with the Project, including without limitation, all sales, use, and rental taxes levied by any federal, state, county or municipal authority or political subdivision thereof.</p>

            <hr>

            <h4>LATE FEES</h4>
            <p>Customer agrees that all amounts not paid by due date stated on invoice sent by Company are subject to a late fee of 18% per annum or the maximum rate allowable by law, and Customer agrees to pay such late fee.</p>

            <hr>

            <h4>OWNERSHIP OF SIGNAGE PROPERTY</h4>
            <p>Company shall contribute parts and materials to manufacture the signage related to the Project (“Signage Property”). Customer acknowledges and agrees that all Signage Property is owned by the Company until receipt of final payment for the Project. Customer expressly agrees that title to the Signage Property is retained by Company and in Company's name until Customer’s full payment for the Project is received. Customer further agrees that if Customer fails to make payment in full for the Project within 90 days of completion of the Project, then Company, or Company's representative, in its sole discretion, shall have the right, and is hereby authorized and empowered to take and remove the Signage Property from the installation site, and resume possession of the Signage Property, wherever found, without any liability for damages or other claim whatsoever, with or without process of law, and without prejudice to further enforcement of any balance of such obligation or expenses remaining due.</p>

            <hr>

            <h4>OWNERSHIP OF COMPANY DESIGNS</h4>
            <p>Company may provide Customer with designs and artwork created by the Company in connection with the Project (“Company Designs”). All right, title and interest in and to the Company Designs is owned exclusively, throughout the world, and in perpetuity by the Company (including all copyrights and patents, derivatives, renewals and extensions thereof). Any and all use of the Company Designs by Customer, its employees or agents is expressly prohibited without the written consent of the Company; and such written consent is subject to payment in full for the Project and the Company’s design service. Until payment in full is received, the Company shall have the sole and exclusive right to use the Company Designs, in whole or in part, in whatever manner the Company may desire, including without limitation, the right to cut, edit, revise, alter and/or otherwise modify the Company Designs and to freely use, perform, distribute, exhibit and exploit such materials and license others to do so in any and all media now known or hereafter devised and shall have the sole and exclusive right to copyright or patent the Company Designs in the Company’s name, as the owner and author thereof.</p>

            <hr>

            <h4>PERMITS AND LICENSES</h4>
            <p>Unless otherwise stated on the reverse side of this Contract, the Company shall obtain all necessary installation permits related to the Project. Customer shall be responsible for maintaining all necessary permits or variances from public authorities.</p>

            <hr>

            <h4>CHANGE ORDERS</h4>
            <p>Any changes to the Standard Specifications that are requested by Customer shall be agreed to by the parties in a Client Change Order Contract, which, upon signature by all parties shall be made part of this Contract. Company may, in its sole discretion, stop all work in connection with the Project until the Client Change Order Contract is signed by Customer.</p>

            <hr>

            <h4>CUSTOMER DELAY</h4>
            <p>Company shall not be liable for any delay in the performance of this Contract caused by or resulting from Customer’s acts, omissions, or delays in its obligations under this Contract.</p>

            <hr>

            <h4>TERMINATION</h4>
            <p>In the event that this Contract is terminated, Customer shall pay Company for all work in progress related to the Project up to date of termination plus a cancellation fee of 25% of the total Contract Amount.</p>

            <hr>

            <h4>LIMITED WARRANTY</h4>
            <p>ALL SIGNAGE PROPERTY AND SERVICES PROVIDED BY COMPANY IS GUARANTEED, PARTS & LABOR, FOR A PERIOD OF FIVE YEARS AGAINST MECHANICAL DEFECTS WITH AN ADDITIONAL LIFETIME MANUFACTURERS PARTS WARRANTY FOR PRINCIPAL LEDS AND POWER SUPPLIES (ref Principal Warranty and Terms). Expressly excluded from this warranty are acts of God, vandalism, customer modification or defects due to Customer negligence and any other causes beyond the control of the Company. All warranties are void with respect to portions of Project not manufactured, performed or serviced by Company, its employees or agents.</p>

            <hr>

            <h4>DISCLAIMER OF WARRANTIES</h4>
            <p>This contract is made with the understanding that there are no expressed or implied warranties other than those contained in this contract and that there are no warranties of any kind, expressed or implied, that the goods shall be merchantable or fit for any particular use or purpose other than those specifically mentioned herein.</p>

            <hr>

            <h4>FORCE MAJEURE</h4>
            <p>Company shall not be liable for failure of or delays in the performance of the terms of this Contract resulting from strikes, breakage, fire, labor disputes, unforeseen commercial delays, war, acts of God, or other causes beyond the control of the Company. In addition, the Customer shall not hold the Company responsible, and Company shall not be liable for any damage to landscaping that occurs during installation. Company shall not be liable for roof warranty work if roof membrane is penetrated as part of the normal installation process. The Customer agrees to pay original roof contractor for any needed repairs or patching in order to keep original roof warranty in tact.</p>

            <hr>

            <h4>MECHANICS LIEN</h4>
            <p>Customer acknowledges and agrees that the Company shall provide services and furnish materials and labor to manufacture signage related to the Project and further agrees that such materials or labor is for improvement of real property. Therefore, unless otherwise prohibited by state law, Customer authorizes Company to file a Mechanic's Lien for any amounts due under this Contract after (30) days. The Customer further agrees to pay Company an administration fee for all costs incurred in filing of a Mechanic's Lien, and further agrees to pay all legal fees and court costs in connection with the enforcement of a Mechanic’s Lien.</p>

            <hr>

            <h4>INDEMNIFICATION</h4>
            <p>The Customer shall indemnify and hold Company, and its employees and agents harmless from and against any and all claims, damages, losses and expenses, including, without limitation, attorneys’ fees and court costs arising out of or resulting from the performance of the services, if any such claims, damage, loss or expense is caused in whole or in any part by any act or omission of the Customer, or Customer’s employees or agents.</p>

            <hr>

            <h4>INDEPENDENT CONTRACTOR STATUS</h4>
            <p>In this Contract, Company shall be deemed an independent contractor. It is the intention of the parties that: (i) the Company shall specifically not occupy the status of an agent, servant, or employee of the Customer; and (ii) the relationship between the Company and the Customer shall specifically not be that of a partnership, joint venture, or other similar association. During the progress of performance of the Project, Customer will not, without Company’s prior written consent, direct or attempt to direct the employees, agents, or subcontractors involved in performance of services related to the Project or the installation of the Signage Property.</p>

            <hr>

            <h4>MISCELLANEOUS</h4>
            <p>The parties hereby agree that: (i) This Contract shall be governed by and construed in accordance with the laws of the State of the Company’s principal office without regard to choice of law principles, and Customer hereby irrevocably submits to the jurisdiction of the state and federal courts in such state for all disputes or legal claims arising from this Contract; (ii) In any legal action brought by or against the Company in relation to this Contract, the prevailing party shall be entitled to recover its costs and reasonable attorney fees in addition to any other relief that may be awarded; (iii) If a court should find one or more of the terms of this Contract unenforceable, the remaining terms will nonetheless remain binding on the parties; (iv) This Contract is the complete agreement between the parties regarding the subject matter set forth herein, and this Contract supercedes all previous oral or written agreements regarding this subject matter; (v) This Contract cannot be voided or amended without the written agreement by an officer of the Company; and (vi) Customer shall not make any assignment of this Contract, but the Company may assign this Contract in its sole discretion.</p>

    </div>
</div>

</body>
</html>
"""

########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
""""""
""""""
"""FLASK ROUTES"""
""""""
""""""
########################################################################################################################

"""ROUTE TO CALL THE INITIAL OPPORTUNITIES TEMPLATE"""

@app.route("/")
@login_required
def index():
    opportunities = get_opportunities()
    customers = get_customers()
    return render_template_string(OPPORTUNITIES_TEMPLATE, opportunities=opportunities, customers=customers)

########################################################################################################################

"""ROUTE TO ADD A NEW CUSTOMER"""

@app.route("/add_customer", methods=["POST"])
def add_customer_route():
    customer_name = request.form["customer_name"]
    customer_email = request.form["customer_email"]
    billing_address = request.form["billing_address"]
    billing_city = request.form["billing_city"]
    billing_state = request.form["billing_state"]
    billing_zip = request.form["billing_zip"]
    contact_first_name = request.form["contact_first_name"]
    contact_last_name = request.form["contact_last_name"]
    phone = request.form["phone"]

    # Call your backend helper with all fields
    add_customer(
        customer_name,
        customer_email,
        billing_address,
        billing_city,
        billing_state,
        billing_zip,
        contact_first_name,
        contact_last_name,
        phone
    )

    return redirect(url_for("index"))

########################################################################################################################

"""ROUTE TO UPDATE CUSTOMER INFO"""

@app.route("/customer/<int:customer_id>", methods=["GET", "POST"])
def customer_detail_route(customer_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # ==============================================
    # 1. Handle Customer Info Update
    # ==============================================
    if request.method == "POST" and not request.form.get("description"):
        customer_name = request.form.get("customer_name")
        customer_email = request.form.get("customer_email")
        billing_address = request.form.get("billing_address")
        billing_city = request.form.get("billing_city")
        billing_state = request.form.get("billing_state")
        billing_zip = request.form.get("billing_zip")
        contact_first_name = request.form.get("contact_first_name")
        contact_last_name = request.form.get("contact_last_name")
        contact_phone = request.form.get("phone")

        cursor.execute("""
            UPDATE Customers
            SET customer_name=?, customer_email=?, billing_address=?, billing_city=?, billing_state=?, billing_zip=?,
                contact_first_name=?, contact_last_name=?, contact_phone=?
            WHERE customer_ID=?
        """, (
            customer_name, customer_email, billing_address, billing_city, billing_state, billing_zip,
            contact_first_name, contact_last_name, contact_phone, customer_id
        ))
        conn.commit()

    # ==============================================
    # 2. Auto-update all line item unit prices
    # ==============================================
    cursor.execute("SELECT line_ID FROM Customer_Line_Items WHERE customer_ID=?", (customer_id,))
    line_ids = [row[0] for row in cursor.fetchall()]

    for line_id in line_ids:
        # Compute total unit price from linked components
        cursor.execute("""
            SELECT SUM(unit_price * quantity)
            FROM Components
            WHERE line_ID = ?
        """, (line_id,))
        total_price = cursor.fetchone()[0] or 0

        # (Optional) Compute unit cost the same way
        cursor.execute("""
            SELECT SUM(unit_cost * quantity)
            FROM Components
            WHERE line_ID = ?
        """, (line_id,))
        total_cost = cursor.fetchone()[0] or 0

        cursor.execute("""
            UPDATE Customer_Line_Items
            SET unit_price = ?, unit_cost = ?
            WHERE line_ID = ?
        """, (total_price, total_cost, line_id))

    conn.commit()

    # ==============================================
    # 3. Get Customer Record
    # ==============================================
    cursor.execute("SELECT * FROM Customers WHERE customer_ID=?", (customer_id,))
    customer = cursor.fetchone()

    # ==============================================
    # 4. Get Opportunities
    # ==============================================
    cursor.execute("""
        SELECT opportunity_ID, opportunity_name, opportunity_price, tax_rate,
               site_address, site_city, site_state, site_zip
        FROM Opportunities
        WHERE customer_ID=?
    """, (customer_id,))
    opportunities = cursor.fetchall()

    # ==============================================
    # 5. Get Saved Customer Line Items + Components
    # ==============================================
    cursor.execute("""
        SELECT line_ID, customer_ID, line_item_description, quantity, unit_cost, unit_price, line_item_sequence
        FROM Customer_Line_Items
        WHERE customer_ID=?
        ORDER BY line_item_sequence
    """, (customer_id,))
    line_items = cursor.fetchall()

    # Build line item dicts including their components
    items = []
    for li in line_items:
        cursor2 = conn.cursor()
        cursor2.execute("SELECT * FROM Components WHERE line_ID=?", (li.line_ID,))
        components = [dict(zip([d[0] for d in cursor2.description], row)) for row in cursor2.fetchall()]

        li_dict = dict(zip([d[0] for d in cursor.description], li))
        li_dict["components"] = components
        items.append(li_dict)

    # ==============================================
    # 6. Get Component Types (for Add Component dropdown)
    # ==============================================
    cursor.execute("SELECT * FROM Component_Types ORDER BY component_types_description")
    component_types = cursor.fetchall()

    conn.close()

    # ==============================================
    # 7. Render Full Template
    # ==============================================
    return render_template_string(
        CUSTOMER_DETAIL_TEMPLATE,
        customer=customer,
        opportunities=opportunities,
        customer_line_items=items,
        component_types=component_types
    )

########################################################################################################################

"""ROUTE TO ADD A NEW OPPORTUNITY"""

@app.route("/add_opportunity", methods=["POST"])
def add_opportunity_route():
    customer_id = request.form["customer_id"]
    opportunity_name = request.form["opportunity_name"]
    tax_rate = float(request.form.get("tax_rate", 0))
    site_address = request.form["site_address"]
    site_city = request.form["site_city"]
    site_state = request.form["site_state"]
    site_zip = request.form["site_zip"]

    add_opportunity(customer_id, opportunity_name, tax_rate, site_address, site_city, site_state, site_zip)
    return redirect(url_for("index"))

########################################################################################################################

"""ROUTE TO UPDATE AN OPPORTUNITY"""

@app.route("/update_opportunity/<int:opportunity_id>", methods=["POST"])
def update_opportunity_route(opportunity_id):
    opportunity_name = request.form.get("opportunity_name")
    tax_rate = request.form.get("tax_rate")
    site_address = request.form.get("site_address")
    site_city = request.form.get("site_city")
    site_state = request.form.get("site_state")
    site_zip = request.form.get("site_zip")

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE Opportunities
        SET opportunity_name = ?, tax_rate = ?, site_address = ?, site_city = ?, site_state = ?, site_zip = ?
        WHERE opportunity_ID = ?
    """, (opportunity_name, tax_rate, site_address, site_city, site_state, site_zip, opportunity_id))
    conn.commit()
    conn.close()

    return redirect(url_for("show_opportunity_route", opportunity_id=opportunity_id))

########################################################################################################################

"""ROUTE TO SHOW THE DETAILS OF AN OPPORTUNITY (THE LINE ITEMS)"""

@app.route("/opportunity/<int:opportunity_id>")
def show_opportunity_route(opportunity_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # ✅ STEP 1: Create or replace a temp table for component totals
    try:
        cursor.execute("""
            SELECT C.line_ID, SUM(C.quantity * C.unit_price) AS new_price
            INTO TempPriceUpdate
            FROM Components AS C
            INNER JOIN Line_Items AS L ON C.line_ID = L.line_ID
            WHERE L.opportunity_ID = ?
            GROUP BY C.line_ID
        """, (opportunity_id,))

        cursor.execute("""
            UPDATE Line_Items AS L
            INNER JOIN TempPriceUpdate AS T
            ON L.line_ID = T.line_ID
            SET L.unit_price = T.new_price
        """)

        cursor.execute("DROP TABLE TempPriceUpdate")
        conn.commit()
    except Exception as e:
        print("⚠️ TempPriceUpdate skipped (likely no components):", e)
        conn.rollback()

    # ✅ STEP 2: Load Opportunity data
    rows = get_line_items(opportunity_id)
    component_types = get_component_types()
    component_type_lookup = {ct[0]: ct[1] for ct in component_types}

    items = []
    for row in rows:
        item = dict(zip([column[0] for column in row.cursor_description], row))

        comps = get_components(item["line_ID"])
        components_list = []

        for c in comps:
            comp_dict = {
                "component_ID": int(c[0]),
                "component_type_ID": int(c[2]) if c[2] else None,
                "description": component_type_lookup.get(c[2], "Unknown"),
                "quantity": float(c[3] or 0),
                "unit_cost": float(c[4] or 0),
                "unit_price": float(c[5] or 0),
            }
            components_list.append(comp_dict)

        item["components"] = components_list
        items.append(item)

    # ✅ STEP 3: Load Standard Line Items
    try:
        cursor.execute("SELECT * FROM Standard_Line_Items")
        standard_line_items = cursor.fetchall()
    except Exception as e:
        print("⚠️ Could not load Standard_Line_Items:", e)
        standard_line_items = []

    conn.close()

    # ====== ⛔ CRITICAL ADDITION BELOW: PASS USER TYPE TO TEMPLATE ======
    user_type = session.get("employee_type", "Unknown")
    # ===================================================================

    # Popup state
    show_popup = request.args.get("show_popup", "false")
    popup_line_id = request.args.get("line_id")

    return render_template_string(
        LINE_ITEMS_TEMPLATE,
        items=items,
        opportunity_id=opportunity_id,
        component_types=component_types,
        show_popup=show_popup,
        popup_line_id=popup_line_id,
        standard_line_items=standard_line_items,
        user_type=user_type     # ✅ MUST BE PASSED IN
    )

########################################################################################################################

"""ROUTE TO ADD A LINE ITEM TO AN OPPORTUNITY"""

@app.route("/opportunity/<int:opportunity_id>/add_line_item", methods=["POST"])
def add_line_item_route(opportunity_id):
    description = request.form["description"]
    quantity = int(request.form["quantity"])

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # --- Get unified next available line_ID (check both tables) ---
    cursor.execute("SELECT MAX(line_ID) FROM Line_Items")
    max_line_items = cursor.fetchone()[0]
    max_line_items = int(max_line_items) if max_line_items is not None else 0

    cursor.execute("SELECT MAX(line_ID) FROM Customer_Line_Items")
    max_customer_items = cursor.fetchone()[0]
    max_customer_items = int(max_customer_items) if max_customer_items is not None else 0

    next_line_id = max(int(max_line_items), int(max_customer_items)) + 1

    # --- Get next sequence number within this opportunity ---
    cursor.execute("""
        SELECT MAX(line_item_sequence) FROM Line_Items WHERE opportunity_ID = ?
    """, (opportunity_id,))
    max_seq = cursor.fetchone()[0] or 0
    new_sequence = max_seq + 100

    # --- Insert new line item with unified line_ID and ACTIVE status ---
    cursor.execute("""
        INSERT INTO Line_Items (
            line_ID, opportunity_ID, line_item_description, quantity, unit_cost, unit_price, line_item_sequence, activation_status
        )
        VALUES (?, ?, ?, ?, 0, 0, ?, 'ACTIVE')
    """, (next_line_id, opportunity_id, description, quantity, new_sequence))

    # --- Update opportunity total ---
    update_opportunity_price(opportunity_id)

    conn.commit()
    conn.close()

    return redirect(url_for("show_opportunity_route", opportunity_id=opportunity_id))

########################################################################################################################

"""ROUTE TO UPDATE A LINE ITEMS ACTIVATION STATUS, SEQUENCE NUMBER AND QUANTITY"""

@app.route("/opportunity/<int:opportunity_id>/update_line_item/<int:line_id>", methods=["POST"])
def update_line_item_route(opportunity_id, line_id):
    description = request.form["description"]
    quantity = int(request.form["quantity"])
    activation_status = request.form.get("activation_status", "INACTIVE")  # or however you handle it
    sequence_number = int(request.form["sequence_number"])  # <-- new

    # Now update the line item in DB
    update_line_item(line_id, description, quantity, activation_status, sequence_number)

    # Update opportunity total
    update_opportunity_price(opportunity_id)

    return redirect(url_for("show_opportunity_route", opportunity_id=opportunity_id))

########################################################################################################################

"""ROUTE TO ADD STOCK MATERIALS TO A CUSTOM COMPONENT"""

@app.route("/component/<int:component_id>/add_material", methods=["POST"])
def add_component_material(component_id):
    customer_id = request.args.get("customer_id")
    opportunity_id = request.args.get("opportunity_id")
    component_type_id = request.args.get("component_type_id", 1)

    material_id = request.form["material_id"]
    quantity = float(request.form["quantity"] or 0)

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
        VALUES (?, ?, ?)
    """, (component_id, material_id, quantity))
    conn.commit()
    conn.close()

    update_component_totals(component_id)

    return redirect(url_for(
        'quote_component',
        component_id=component_id,
        component_type_id=1,
        customer_id=request.args.get('customer_id') or request.form.get('customer_id') or request.args.get('cust'),
        opportunity_id=request.args.get('opportunity_id') or request.form.get('opportunity_id') or request.args.get(
            'opp')
    ))

########################################################################################################################

"""ROUTE TO ADD NONSTOCK MATERIALS TO A CUSTOM COMPONENT"""

@app.route("/component/<int:component_id>/add_nonstock_material", methods=["POST"])
def add_nonstock_component_material(component_id):
    customer_id = request.args.get("customer_id")
    opportunity_id = request.args.get("opportunity_id")
    component_type_id = request.args.get("component_type_id", 1)

    material_description = request.form["material_description"]
    material_unit = request.form["material_unit"]
    material_price = float(request.form["material_price"])
    quantity = float(request.form["quantity"] or 0)

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO Materials (material_description, material_unit, material_price, stock)
        VALUES (?, ?, ?, 'NO')
    """, (material_description, material_unit, material_price))
    conn.commit()

    cursor.execute("SELECT @@IDENTITY AS new_id")
    new_material_id = cursor.fetchone().new_id

    cursor.execute("""
        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
        VALUES (?, ?, ?)
    """, (component_id, new_material_id, quantity))
    conn.commit()
    conn.close()

    update_component_totals(component_id)

    return redirect(url_for(
        'quote_component',
        component_id=component_id,
        component_type_id=1,
        customer_id=request.args.get('customer_id') or request.form.get('customer_id') or request.args.get('cust'),
        opportunity_id=request.args.get('opportunity_id') or request.form.get('opportunity_id') or request.args.get(
            'opp')
    ))

########################################################################################################################

"""ROUTE TO ADD LABOR TO A CUSTOM COMPONENT"""

@app.route("/component/<int:component_id>/add_labor", methods=["POST"])
def add_component_labor(component_id):
    customer_id = request.args.get("customer_id")
    opportunity_id = request.args.get("opportunity_id")
    component_type_id = request.args.get("component_type_id", 1)

    labor_id = int(request.form["labor_id"])
    quantity = float(request.form["quantity"] or 0)

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
        VALUES (?, ?, ?)
    """, (component_id, labor_id, quantity))
    conn.commit()
    conn.close()

    update_component_totals(component_id)

    return redirect(url_for(
        'quote_component',
        component_id=component_id,
        component_type_id=1,
        customer_id=request.args.get('customer_id') or request.form.get('customer_id') or request.args.get('cust'),
        opportunity_id=request.args.get('opportunity_id') or request.form.get('opportunity_id') or request.args.get(
            'opp')
    ))

########################################################################################################################

"""ROUTE TO MAKE CONTRACT"""

@app.route("/opportunity/<int:opportunity_id>/contract")
def contract_route(opportunity_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Fetch opportunity info
    cursor.execute("""
        SELECT o.opportunity_ID, o.opportunity_name, o.opportunity_price, o.tax_rate, o.site_address, o.site_city,
               o.site_state, o.site_zip, 
               c.customer_name, c.customer_email, c.contact_first_name, c.contact_last_name, c.contact_phone,
               c.billing_address, c.billing_city, c.billing_state, c.billing_zip
        FROM Opportunities o
        INNER JOIN Customers c ON o.customer_ID = c.customer_ID
        WHERE o.opportunity_ID = ?
    """, opportunity_id)
    opportunity = cursor.fetchone()
    if not opportunity:
        conn.close()
        return "Opportunity not found", 404

    # Convert opportunity to dict for template
    opportunity_dict = {
        "opportunity_ID": opportunity.opportunity_ID,
        "opportunity_name": opportunity.opportunity_name,
        "opportunity_price": float(opportunity.opportunity_price or 0),
        "tax_rate": float(opportunity.tax_rate or 0),
        "site_address": opportunity.site_address,
        "site_city": opportunity.site_city,
        "site_state": opportunity.site_state,
        "site_zip": opportunity.site_zip,
        "customer_name": opportunity.customer_name,
        "customer_email": opportunity.customer_email,
        "contact_first_name": opportunity.contact_first_name,
        "contact_last_name": opportunity.contact_last_name,
        "contact_phone": opportunity.contact_phone,
        "billing_address": opportunity.billing_address,
        "billing_city": opportunity.billing_city,
        "billing_state": opportunity.billing_state,
        "billing_zip": opportunity.billing_zip
    }

    # Fetch line items for this opportunity
    cursor.execute("""
        SELECT line_ID, line_item_description, quantity, unit_price, activation_status
        FROM Line_Items
        WHERE opportunity_ID = ?
        ORDER BY line_item_sequence
    """, opportunity_id)

    columns = [column[0] for column in cursor.description]
    line_items = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Filter only ACTIVE line items
    active_items = []
    subtotal = 0.0
    for item in line_items:
        if item["activation_status"] == "ACTIVE":
            item["quantity"] = float(item["quantity"] or 0)
            item["unit_price"] = float(item["unit_price"] or 0)
            item["subtotal"] = item["quantity"] * item["unit_price"]
            subtotal += item["subtotal"]
            active_items.append(item)

    tax_rate = opportunity_dict["tax_rate"]
    tax_amount = subtotal * (tax_rate / 100)
    grand_total = subtotal + tax_amount

    # ✅ Embed your company logo directly as Base64
    try:
        with open(r"C:\Users\Brooks\OneDrive\Desktop\Picture1.jpg", "rb") as img_file:
            logo_b64 = base64.b64encode(img_file.read()).decode("utf-8")
    except Exception as e:
        logo_b64 = None
        print("⚠️ Logo load failed:", e)

    conn.close()

    # Render the contract template with embedded logo
    return render_template_string(
        CONTRACT_TEMPLATE,
        opportunity=opportunity_dict,
        line_items=active_items,
        subtotal=subtotal,
        tax_amount=tax_amount,
        grand_total=grand_total,
        company_logo=logo_b64  # ✅ pass Base64 image to template
    )

########################################################################################################################

"""ROUTE TO ADD A COMPONENT TO A LINE ITEM"""

@app.route("/line_item/<int:line_id>/add_component", methods=["POST"])
def add_component_route(line_id):
    """
    Add a new component to a given line item.
    """
    try:
        # Get the selected component type from the form
        component_type_id = request.form.get("component_type_id")
        print("Form component_type_id:", component_type_id)  # DEBUG

        if not component_type_id:
            flash("Please select a component type.")
            print("No component_type_id provided, redirecting.")  # DEBUG
            return redirect(url_for("show_opportunity_route",
                                    opportunity_id=get_opportunity_id_by_line(line_id)))

        # Convert to integer just in case
        component_type_id = int(component_type_id)

        # Default values for a new component
        quantity = 1
        unit_cost = 0
        unit_price = 0
        factor1 = factor2 = factor3 = factor4 = factor5 = 0

        # Insert new component into Components table
        conn = pyodbc.connect(CONN_STR)
        cursor = conn.cursor()
        sql = """
            INSERT INTO Components
            (line_ID, component_type_ID, quantity, unit_cost, unit_price,
             factor1, factor2, factor3, factor4, factor5)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        params = (line_id, component_type_id, quantity, unit_cost, unit_price,
                  factor1, factor2, factor3, factor4, factor5)

        print("Executing SQL:", sql)  # DEBUG
        print("With params:", params)  # DEBUG
        cursor.execute(sql, params)
        print("Inserted rows:", cursor.rowcount)  # DEBUG

        conn.commit()
        conn.close()
        print("Insert committed and connection closed.")  # DEBUG

        # Redirect back to the line items page for this opportunity
        opportunity_id = get_opportunity_id_by_line(line_id)
        print("Redirecting to opportunity:", opportunity_id)  # DEBUG
        return redirect(url_for("show_opportunity_route", opportunity_id=opportunity_id))

    except Exception as e:
        print("Error inserting component:", str(e))  # DEBUG
        flash("Error adding component. Check logs for details.")
        return redirect(url_for("show_opportunity_route",
                                opportunity_id=get_opportunity_id_by_line(line_id)))

########################################################################################################################

"""ROUTE TO UPDATE LINE ITEMS AND COMPONENTS"""

@app.route("/opportunity/<int:opportunity_id>/update_line_item_and_components/<int:line_id>", methods=["POST"])
def update_line_item_and_components(opportunity_id, line_id):
    description = request.form["description"]
    quantity = int(request.form["quantity"])
    activation_status = request.form.get("activation_status", "INACTIVE")
    sequence_number = int(request.form["sequence_number"])
    component_type_id = request.form.get("component_type_id")

    # Update line item
    update_line_item(line_id, description, quantity, activation_status, sequence_number)

    # Add component if selected
    if component_type_id:
        # Insert new component with default values
        conn = pyodbc.connect(CONN_STR)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO Components
            (line_ID, component_type_ID, quantity, unit_cost, unit_price,
             factor1, factor2, factor3, factor4, factor5, factor6)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (line_id, component_type_id, 1, 0, 0, 0, 0, 0, 0, 0, 0))
        conn.commit()
        conn.close()

    # Update line totals
    update_line_item_totals(line_id)
    update_opportunity_price(opportunity_id)

    return redirect(url_for("show_opportunity_route", opportunity_id=opportunity_id))

########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################

"""ROUTE FOR QUOTING COMPONENTS.  THIS IS WHERE ALL THE QUOTING BACKEND HAPPENS"""

@app.route("/quote_component/<int:component_id>/<int:component_type_id>", methods=["GET", "POST"])
def quote_component(component_id, component_type_id):

    # --- 🔒 SAFELY get both IDs from args or form ---
    raw_customer_id = request.args.get("customer_id") or request.form.get("customer_id")
    raw_opportunity_id = request.args.get("opportunity_id") or request.form.get("opportunity_id")

    try:
        customer_id = int(raw_customer_id) if raw_customer_id not in (None, "", "None") else None
    except ValueError:
        customer_id = None

    try:
        opportunity_id = int(raw_opportunity_id) if raw_opportunity_id not in (None, "", "None") else None
    except ValueError:
        opportunity_id = None

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Try normal Line_Items first (opportunity-level)
    cursor.execute("""
        SELECT c.component_ID, c.unit_cost, c.unit_price,
               c.factor1, c.factor2, c.factor3, c.factor4, c.factor5,
               c.factor6, c.factor7, c.factor8, c.factor9, c.factor10,
               c.factor11, c.factor12, c.factor13, c.factor14, c.factor15,
               c.factor16, c.factor17, c.factor18, c.factor19, c.factor20,
               c.factor21, c.factor22, c.factor23, c.factor24, 
               c.factor25, c.factor26, c.factor27,
               l.line_item_description, l.opportunity_id, l.line_ID
        FROM Components c
        INNER JOIN Line_Items l ON c.line_ID = l.line_ID
        WHERE c.component_ID = ?
    """, (component_id,))
    row = cursor.fetchone()

    # If not found, try the Customer_Line_Items table instead
    if not row:
        cursor.execute("""
            SELECT c.component_ID, c.unit_cost, c.unit_price,
                   c.factor1, c.factor2, c.factor3, c.factor4, c.factor5,
                   c.factor6, c.factor7, c.factor8, c.factor9, c.factor10,
                   c.factor11, c.factor12, c.factor13, c.factor14, c.factor15,
                   c.factor16, c.factor17, c.factor18, c.factor19, c.factor20,
                   c.factor21, c.factor22, c.factor23, c.factor24, 
                   c.factor25, c.factor26, c.factor27,
                   l.line_item_description, NULL AS opportunity_id, l.line_ID
            FROM Components c
            INNER JOIN Customer_Line_Items l ON c.line_ID = l.line_ID
            WHERE c.component_ID = ?
        """, (component_id,))
        row = cursor.fetchone()

    conn.close()

    if not row:
        return "Component not found", 404

    # Helper to parse stored comma-separated factor strings
    def parse_factor_string(val):
        if not val:
            return []
        return [x for x in str(val).split(",") if x and x != "0"]

    component_unit_cost = float(row.unit_cost or 0)
    component_unit_price = float(row.unit_price or 0)
    line_item_description = row.line_item_description
    # ⚙️ If DB returns opportunity_id as None, fall back to our parsed one
    opportunity_id = row.opportunity_id or opportunity_id
    line_id = row.line_ID

    # --- Base context for templates ---
    context = dict(
        component_id=component_id,
        line_item_description=line_item_description,
        component_unit_cost=component_unit_cost,
        component_unit_price=component_unit_price,
        line_id=line_id,
    )

    # FORCE these into the context so CUSTOM SIGN never loses them
    context["opportunity_id"] = opportunity_id or 0
    context["customer_id"] = customer_id or 0

    # Add factor1..factor24 normally
    for i in range(1, 25):
        context[f"factor{i}"] = getattr(row, f"factor{i}", None)

    # Parse factor25–27 into lists
    factor25_list = parse_factor_string(getattr(row, "factor25", ""))
    factor26_list = parse_factor_string(getattr(row, "factor26", ""))
    factor27_list = parse_factor_string(getattr(row, "factor27", ""))

    # Combine into rows for display
    excel_rows = list(zip(factor25_list, factor26_list, factor27_list))

    # Add to context
    context["factor25_list"] = factor25_list
    context["factor26_list"] = factor26_list
    context["factor27_list"] = factor27_list
    context["excel_rows"] = excel_rows


    if component_type_id == 2:  # INSTALLATION

        install_labor_types = get_install_labor_types()
        component_install_labor = get_component_install_labor(component_id)
        component_install_materials = get_component_install_materials(component_id)
        subcontract_costs = get_subcontract_install_costs(component_id)

        context.update(
            install_labor_types=install_labor_types,
            component_install_labor=component_install_labor,
            component_install_materials=component_install_materials,
            subcontract_costs=subcontract_costs,
            customer_id=customer_id,
            opportunity_id=opportunity_id
        )
    # ------------------------------------------------------------
    # Handle opportunity/customer context
    # ------------------------------------------------------------
        if not opportunity_id:
            opportunity_id = 0
            context["opportunity_id"] = opportunity_id
            context["hide_back_button"] = True
        else:
            context["hide_back_button"] = False

    # Try to fetch or infer the customer_id (same pattern as Custom Sign)
        customer_id = (
            locals().get("customer_id")
            or context.get("customer_id")
            or getattr(row, "customer_id", None)
            or 0
        )
        context["customer_id"] = customer_id or 0
        context["hide_back_button"] = bool(customer_id)

        print("customer_id:", context.get("customer_id"), "hide_back_button:", context.get("hide_back_button"))

        # ------------------------------------------------------------
    # Render the Install Quote Template
    # ------------------------------------------------------------
        return render_template_string(QUOTE_INSTALLATION, **context)

    elif component_type_id == 3:  # PIPE & FOUNDATIONS
        params = {}
        options = []

        if request.method == "POST":
            params = {
                "overall_height": request.form.get("overall_height"),
                "wind_speed": request.form.get("wind_speed"),
                "exposure": request.form.get("exposure"),
                "foundation_type": request.form.get("foundation_type"),
            }

        context.update(params=params, options=options)
        return render_template_string(QUOTE_PIPE_FOUNDATIONS, **context)

    elif component_type_id == 4:  # EMC
        # Fetch EMC rows
        component_emc = get_component_emc(component_id)

        # Inject into template context
        context.update(
            component_emc=component_emc,
        )

        # Mirror the same context handling you used for Custom Sign / Install
        # (Assumes `row` above already has component + line info like before)
        if not opportunity_id:
            opportunity_id = 0
            context["opportunity_id"] = opportunity_id

        # Try to infer customer_id the same way
        customer_id = (
                locals().get("customer_id")
                or context.get("customer_id")
                or getattr(row, "customer_id", None)
                or 0
        )
        context["customer_id"] = customer_id or 0

        # (Optional) If you still use hide_back_button elsewhere, you can omit it here
        # and just let the template show the correct link based on which ID is nonzero.

        return render_template_string(QUOTE_EMC, **context)

    elif component_type_id == 5:  # Face-Lit Channel Letters
        # ✅ Always capture IDs safely
        customer_id = request.form.get("customer_id") or request.args.get("customer_id") or None
        opportunity_id = request.form.get("opportunity_id") or request.args.get("opportunity_id") or None

        # ✅ Clean type coercion
        try:
            opportunity_id = int(opportunity_id) if opportunity_id not in (None, "", "None") else None
        except ValueError:
            opportunity_id = None

        try:
            customer_id = int(customer_id) if customer_id not in (None, "", "None") else None
        except ValueError:
            customer_id = None

        # ✅ Add to template context
        context.update(customer_id=customer_id, opportunity_id=opportunity_id)

        if request.method == "POST":
            file = request.files.get("excel_file")

            import tempfile, os
            from openpyxl import load_workbook

            modules_str, areas_str, perimeters_str = None, None, None
            modules, areas, perimeters = [], [], []

            # Only process Excel if a file was actually uploaded
            if file and file.filename:
                temp_path = os.path.join(tempfile.gettempdir(), file.filename)
                file.save(temp_path)

                wb = load_workbook(temp_path, data_only=True)
                if "Sheet2" not in wb.sheetnames:
                    return "Excel file missing Sheet2", 400
                sheet = wb["Sheet2"]

                for vals in sheet.iter_rows(min_row=2, values_only=True):
                    mod, area, peri = vals[0], vals[1], vals[2]
                    if mod is None and area is None and peri is None:
                        continue
                    modules.append(mod)
                    areas.append(area)
                    perimeters.append(peri)

                wb.close()
                os.remove(temp_path)

                modules_str = ",".join(str(x) for x in modules)
                areas_str = ",".join(str(x) for x in areas)
                perimeters_str = ",".join(str(x) for x in perimeters)

            # Get current factor1–8 from DB row
            existing_factors = {f"factor{i}": getattr(row, f"factor{i}", None) for i in range(1, 9)}

            # Merge user input with existing values
            factors = {}
            for i in range(1, 9):
                val = request.form.get(f"factor{i}")
                factors[f"factor{i}"] = int(val) if val not in (None, "") else existing_factors[f"factor{i}"]

            # Preserve Excel data if not uploaded
            if not modules_str:
                modules_str = getattr(row, "factor25", None)
            if not areas_str:
                areas_str = getattr(row, "factor26", None)
            if not perimeters_str:
                perimeters_str = getattr(row, "factor27", None)

            # Update factors in Components table
            conn = pyodbc.connect(CONN_STR)
            cursor = conn.cursor()
            cursor.execute("""
                        UPDATE Components
                        SET factor1 = ?, factor2 = ?, factor3 = ?, factor4 = ?,
                            factor5 = ?, factor6 = ?, factor7 = ?, factor8 = ?,
                            factor25 = ?, factor26 = ?, factor27 = ?
                        WHERE component_ID = ?
                    """, (
                factors["factor1"], factors["factor2"], factors["factor3"], factors["factor4"],
                factors["factor5"], factors["factor6"], factors["factor7"], factors["factor8"],
                modules_str, areas_str, perimeters_str, component_id
            ))

            conn.commit()

            # Clear old material/labor records
            cursor.execute("DELETE FROM component_MFG_Materials WHERE component_ID = ?", (component_id,))
            cursor.execute("DELETE FROM component_MFG_Labor WHERE component_ID = ?", (component_id,))

            # ------------------------------
            # 🔥 backend calculations
            # ------------------------------

            factor25_list = modules_str.split(",") if modules_str else []
            factor26_list = areas_str.split(",") if areas_str else []
            factor27_list = perimeters_str.split(",") if perimeters_str else []

            list_of_modules = [float(x) for x in factor25_list if x not in (None, "", "0")]
            list_of_areas = [float(x) for x in factor26_list if x not in (None, "", "0")]
            list_of_perimeters = [float(x) for x in factor27_list if x not in (None, "", "0")]

            ############################################################################################################
            ############################################################################################################
            """CODE FOR MATERIALS"""
            ############################################################################################################
            ############################################################################################################

            ############################################################################################################
            """CODE FOR LETTER BACKS"""
            ############################################################################################################

            if factors["factor1"] == 1:
                ACM_SHEETS = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                """, (component_id, 10, ACM_SHEETS))

            else:
                _063_SHEETS = math.ceil(sum(list_of_areas) / .45 / 40)

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 4, _063_SHEETS))

            ############################################################################################################
            """CODE FOR RETAINERS/TRIMCAP"""
            ############################################################################################################

            if factors["factor2"] == 1:

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 60, math.ceil(sum(list_of_perimeters))))

            elif factors["factor6"] == 2:

                _090_sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 6, _090_sheets))

                _063_sheets = math.ceil(sum(list_of_perimeters) / 240 * 1.1)

                cursor.execute("""
                                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                                    VALUES (?, ?, ?)
                                                    """, (component_id, 4, _063_sheets))

            ############################################################################################################
            """CODE FOR COIL"""
            ############################################################################################################

            if factors["factor2"] == 1:

                _5in_coil = math.ceil(sum(list_of_perimeters))

                cursor.execute("""
                                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                                    VALUES (?, ?, ?)
                                                    """, (component_id, 58, _5in_coil))

            else:

                _3in_coil = math.ceil(sum(list_of_perimeters))

                cursor.execute("""
                                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                                    VALUES (?, ?, ?)
                                                    """, (component_id, 59, _3in_coil))

            ############################################################################################################
            """CODE FOR RACEWAYS"""
            ############################################################################################################

            if factors["factor5"] == 2:
                inserts = [
                    (component_id, 71, 1),
                    (component_id, 72, 1),
                    (component_id, 73, 2),
                    (component_id, 77, 7)
                ]

                cursor.executemany("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, inserts)

            elif factors["factor5"] == 3:
                inserts = [
                    (component_id, 74, 1),
                    (component_id, 75, 1),
                    (component_id, 76, 2),
                    (component_id, 77, 7)
                ]

                cursor.executemany("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, inserts)

            ############################################################################################################
            """CODE FOR FACES"""
            ############################################################################################################

            if factors["factor3"] == 1:

                white_acrylic_sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                """, (component_id, 39, white_acrylic_sheets))

            elif factors["factor3"] == 2:

                clear_acrylic_sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                """, (component_id, 40, clear_acrylic_sheets))

            elif factors["factor3"] == 3:

                white_poly_sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                """, (component_id, 43, white_poly_sheets))

            elif factors["factor3"] == 4:
                clear_poly_sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                """, (component_id, 44, clear_poly_sheets))

            ############################################################################################################
            """CODE FOR LEDS/PS"""
            ############################################################################################################

            inserts = [
                (component_id, 49, math.ceil(sum(list_of_modules) / 200)),
                (component_id, 52, math.ceil(sum(list_of_modules) * .461 / 60)),
                (component_id, 54, 1),
                (component_id, 55, 1)
            ]

            if factors["factor5"] == 1:
                inserts.append(
                    (component_id, 56, math.ceil(math.ceil(sum(list_of_modules) * .461 / 60)) / 2))
                inserts.append((component_id, 57, len(list_of_perimeters)))

            else:
                inserts.append(
                    (component_id, 57, 1)
                )

            cursor.executemany("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, inserts)

            ############################################################################################################
            """CODE FOR INK"""
            ############################################################################################################

            if factors["factor7"] == 1 or factors["factor7"] == 4 or factors["factor7"] == 5:
                ml_ink = math.ceil(sum(list_of_areas))

                cursor.execute("""
                                        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 45, ml_ink))

            if factors["factor7"] == 2 or factors["factor7"] == 4:

                sqft_vinyl = math.ceil(sum(list_of_areas) / .45 / 32) * 32

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 46, sqft_vinyl))

            elif factors["factor7"] == 3 or factors["factor7"] == 5:

                sqft_vinyl = math.ceil(sum(list_of_areas) / .45 / 32) * 32

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 47, sqft_vinyl))

            ############################################################################################################
            """CODE FOR PAINT"""
            ############################################################################################################

            paint_area = 0

            if factors["factor2"] == 1 and (factors["factor4"] == 1 or factors["factor4"] == 3):
                paint_area += math.ceil(sum(list_of_perimeters) * .75)

            if factors["factor2"] == 2 and (factors["factor4"] == 1 or factors["factor4"] == 3):
                paint_area += math.ceil(sum(list_of_perimeters) * .5)

            if factors["factor4"] == 2 or factors["factor4"] == 3:
                paint_area += math.ceil(sum(list_of_perimeters) * .25)

            if factors["factor6"] == 2:
                paint_area += math.ceil(sum(list_of_perimeters) * .75)

            if factors["factor5"] == 2:
                paint_area += 24.5 * 2

            if factors["factor5"] == 3:
                paint_area += 24.5 * 2.5

            if paint_area != 0:
                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 48, paint_area))

            ############################################################################################################
            ############################################################################################################
            """CODE FOR LABOR"""
            ############################################################################################################
            ############################################################################################################

            ############################################################################################################
            """DRAWING LABOR"""
            ############################################################################################################

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 1, 1))

            ############################################################################################################
            """FILE SETUP LABOR"""
            ############################################################################################################

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 2, 1.5))

            ############################################################################################################
            """ROUTING LABOR - BACKS"""
            ############################################################################################################

            if factors["factor1"] == 1:

                sheets = math.ceil(sum(list_of_areas) / .45 / 32)

            else:
                sheets = math.ceil(sum(list_of_areas) / .45 / 40)

            route_time = sheets * (.0081 * (sum(list_of_perimeters) / sheets) + 30.736) / 60

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                """, (component_id, 3, route_time))

            ############################################################################################################
            """COLOR MATCHING LABOR"""
            ############################################################################################################

            if factors["factor8"] != 0:
                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (
                component_id, 14, (2.2392 * factors["factor8"] ** 2 + 15.442 * factors["factor8"] + 32.608) / 60))

            ############################################################################################################
            """PRINTING LABOR"""
            ############################################################################################################

            if factors["factor7"] == 1 or factors["factor7"] == 4 or factors["factor7"] == 5:
                sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                print_time = sheets * (.1533 * (sum(list_of_areas) / sheets) + 36.519) / 60

                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 14, print_time))

            ############################################################################################################
            """ROUTING LABOR - FACES"""
            ############################################################################################################

            sheets = math.ceil(sum(list_of_areas) / .45 / 32)

            route_time = sheets * (.0081 * (sum(list_of_perimeters) / sheets) + 30.736) / 60

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 3, route_time))

            ############################################################################################################
            """VINYL CUTTING LABOR"""
            ############################################################################################################

            if factors['factor7'] == 2 or factors['factor7'] == 3 or factors['factor7'] == 4 or factors['factor7'] == 5:
                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 13, (10 + .0081 * sum(list_of_perimeters)) / 60))

            ############################################################################################################
            """VINYL APPLICATION LABOR"""
            ############################################################################################################

            if factors['factor7'] == 2 or factors['factor7'] == 3 or factors['factor7'] == 4 or factors['factor7'] == 5:
                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 13, (10 * len(list_of_perimeters)) / 60))

            ############################################################################################################
            """TRIMCAP LABOR"""
            ############################################################################################################

            if factors['factor6'] == 1:
                values = [i ** 2 / j for i, j in zip(list_of_perimeters, list_of_areas)]

                trimcap_labor = sum([max(10, (17.135 * math.log(i) - 35.446)) for i in values]) / 60

                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 8, trimcap_labor))

            ############################################################################################################
            """AUTOBENDER LABOR"""
            ############################################################################################################

            autobender_labor = sum([.0224 * i + 2.4177 for i in list_of_perimeters]) / 60

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 4, autobender_labor))

            ############################################################################################################
            """REMOVE COIL PLASTIC LABOR"""
            ############################################################################################################

            coil_plastic_labor = sum([max(1.7329 * math.log(i) - 4.054, 1) for i in list_of_perimeters]) / 60

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 4, coil_plastic_labor))

            ############################################################################################################
            """ATTACH COIL TO BACKS LABOR"""
            ############################################################################################################

            attach_coil_labor = sum([.0003 * i ** - 0.037 * i + 3.1456 for i in list_of_perimeters]) / 60

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 10, attach_coil_labor))

            ############################################################################################################
            """INSTALL LEDS LABOR"""
            ############################################################################################################

            install_LEDs_labor = sum([.0024 * i ** 2 + 1.1666 * i + 3.2765 for i in list_of_modules]) / 60

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 10, install_LEDs_labor))

            ############################################################################################################
            """SILICONE LEDS LABOR"""
            ############################################################################################################

            silicone_LEDs_labor = sum([max(.5, 3.2445 * math.log(i) - 3.5799) for i in list_of_modules]) / 60

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 10, silicone_LEDs_labor))

            ############################################################################################################
            """CAULK SIDEWALLS LABOR"""
            ############################################################################################################

            caulking_labor = sum([.0003 * i ** 2 - .0485 * i + 3.6158 for i in list_of_perimeters]) / 60

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 10, caulking_labor))

            ############################################################################################################
            """ASSEMBLE FACES LABOR"""
            ############################################################################################################

            assemble_faces_labor = 1.5 * len(list_of_perimeters) / 60

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 10, assemble_faces_labor))

            ############################################################################################################
            """ROUTING LABOR - RETAINER LANDINGS"""
            ############################################################################################################

            if factors["factor6"] == 2:
                sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                route_time = (sheets * (.0081 * (sum(list_of_perimeters) * 2.1) / sheets) + 30.736) / 60

                cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 3, route_time))

            ############################################################################################################
            """SHEARING LABOR - RETAINER RETURNS"""
            ############################################################################################################

            if factors["factor6"] == 2:
                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 7, sum(list_of_perimeters) / 10 * 5.5 / 60))

            ############################################################################################################
            """FAB RETAINERS LABOR"""
            ############################################################################################################

            if factors["factor6"] == 2:
                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 7, sum(list_of_perimeters) * 2 / 60))

            ############################################################################################################
            """PAINT AND PAINT PREP LABOR"""
            ############################################################################################################

            if paint_area != 0:
                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 11, max(paint_area / 60, 1)))

                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 12, max(paint_area / 60, 1)))

            ############################################################################################################
            """PRODUCE PATTERN LABOR"""
            ############################################################################################################

            cursor.execute("""
                                INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 5, .5))

            ############################################################################################################
            """RACEWAY LABOR"""
            ############################################################################################################

            if factors["factor5"] != 1:
                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 7, 20 / 60))

                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 7, 10 / 60))

                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 10, len(list_of_perimeters) * 15 / 60))

            ################################################################################################################

            conn.commit()
            cursor.close()
            conn.close()

            # ✅ Update totals
            update_component_totals(component_id)

            # Fetch new totals for display
            component_row_conn = pyodbc.connect(CONN_STR)
            component_cursor = component_row_conn.cursor()
            component_cursor.execute("""
                            SELECT unit_cost, unit_price 
                            FROM Components 
                            WHERE component_ID = ?
                        """, (component_id,))
            uc, up = component_cursor.fetchone()
            component_cursor.close()
            component_row_conn.close()

            context["component_unit_cost"] = uc
            context["component_unit_price"] = up

            # Load linked materials/labor data
            context.update(
                materials=get_materials(),
                labor_types=get_labor_types(),
                component_materials=get_component_materials(component_id),
                component_labor=get_component_labor(component_id),
            )

            # ✅ Render same quoting template for both customer or opp
            return render_template_string(QUOTE_CUSTOM_SIGN, **context)

            # ---------- GET request ----------
            # ✅ Make sure Excel table + navigation IDs show correctly
        context.update(
            excel_rows=[
                (x, y, z)
                for x, y, z in zip(
                    (getattr(row, "factor25", "") or "").split(","),
                    (getattr(row, "factor26", "") or "").split(","),
                    (getattr(row, "factor27", "") or "").split(",")
                )
                if x or y or z
            ],
        )
        return render_template_string(QUOTE_FACE_LIT_CHANNEL_LETTERS_COMPONENT, **context)

    elif component_type_id == 6:  # Reverse Lit Channel Letters
        if request.method == "POST":
            file = request.files.get("excel_file")

            import tempfile, os
            from openpyxl import load_workbook

            modules_str, areas_str, perimeters_str = None, None, None
            modules, areas, perimeters = [], [], []

            # ------------------------------
            # 🔹 Process Excel if uploaded
            # ------------------------------
            if file and file.filename:
                temp_path = os.path.join(tempfile.gettempdir(), file.filename)
                file.save(temp_path)

                wb = load_workbook(temp_path, data_only=True)
                if "Sheet2" not in wb.sheetnames:
                    return "Excel file missing Sheet2", 400
                sheet = wb["Sheet2"]

                for vals in sheet.iter_rows(min_row=2, values_only=True):
                    mod, area, peri = vals[0], vals[1], vals[2]
                    if not any([mod, area, peri]):
                        continue
                    modules.append(mod)
                    areas.append(area)
                    perimeters.append(peri)

                wb.close()
                os.remove(temp_path)

                # Convert to strings for DB
                modules_str = ",".join(str(x) for x in modules)
                areas_str = ",".join(str(x) for x in areas)
                perimeters_str = ",".join(str(x) for x in perimeters)

            # ------------------------------
            # 🔹 Handle factors (1–8)
            # ------------------------------
            existing_factors = {f"factor{i}": getattr(row, f"factor{i}", None) for i in range(1, 9)}
            factors = {}
            for i in range(1, 9):
                val = request.form.get(f"factor{i}")
                if val not in (None, ""):
                    factors[f"factor{i}"] = int(val)
                else:
                    factors[f"factor{i}"] = existing_factors[f"factor{i}"]

            # Keep previous Excel data if no new upload
            if not modules_str:
                modules_str = getattr(row, "factor25", None)
            if not areas_str:
                areas_str = getattr(row, "factor26", None)
            if not perimeters_str:
                perimeters_str = getattr(row, "factor27", None)

            # ------------------------------
            # 🔹 Update DB factors
            # ------------------------------
            conn = pyodbc.connect(CONN_STR)
            cursor = conn.cursor()
            cursor.execute("""
                       UPDATE Components
                       SET factor1 = ?, factor2 = ?, factor3 = ?, factor4 = ?,
                           factor5 = ?, factor6 = ?, factor7 = ?, factor8 = ?,
                           factor25 = ?, factor26 = ?, factor27 = ?
                       WHERE component_ID = ?
                   """, (
                factors["factor1"], factors["factor2"], factors["factor3"], factors["factor4"],
                factors["factor5"], factors["factor6"], factors["factor7"], factors["factor8"],
                modules_str, areas_str, perimeters_str, component_id
            ))

            # Clear previous materials/labor
            cursor.execute("DELETE FROM component_MFG_Materials WHERE component_ID = ?", (component_id,))
            cursor.execute("DELETE FROM component_MFG_Labor WHERE component_ID = ?", (component_id,))

            # ------------------------------
            # 🔹 Core Calculations (unchanged)
            # ------------------------------
            list_of_modules = [float(x) for x in (modules_str or "").split(",") if x not in ("", "0", None)]
            list_of_areas = [float(x) for x in (areas_str or "").split(",") if x not in ("", "0", None)]
            list_of_perimeters = [float(x) for x in (perimeters_str or "").split(",") if x not in ("", "0", None)]

            ############################################################################################################
            ############################################################################################################
            """CODE FOR MATERIALS"""
            ############################################################################################################
            ############################################################################################################

            ############################################################################################################
            """CODE FOR LETTER FACES"""
            ############################################################################################################

            _090_SHEETS = math.ceil(sum(list_of_areas) / .45 / 32)

            cursor.execute("""
                                INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                """, (component_id, 6, _090_SHEETS))

            ############################################################################################################
            """CODE FOR COIL"""
            ############################################################################################################

            _3in_coil = math.ceil(sum(list_of_perimeters))

            cursor.execute("""
                                INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                VALUES (?, ?, ?)
                                """, (component_id, 59, _3in_coil))

            ############################################################################################################
            """CODE FOR RACEWAYS"""
            ############################################################################################################

            if factors["factor4"] == 2:
                inserts = [
                    (component_id, 71, 1),
                    (component_id, 72, 1),
                    (component_id, 73, 2),
                    (component_id, 77, 7)
                ]

                cursor.executemany("""
                                        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, inserts)

            elif factors["factor4"] == 3:
                inserts = [
                    (component_id, 74, 1),
                    (component_id, 75, 1),
                    (component_id, 76, 2),
                    (component_id, 77, 7)
                ]

                cursor.executemany("""
                                        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, inserts)

            ############################################################################################################
            """CODE FOR BACKS"""
            ############################################################################################################

            if factors["factor1"] == 4:

                white_acrylic_sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                        VALUES (?, ?, ?)
                                    """, (component_id, 39, white_acrylic_sheets))

            elif factors["factor1"] == 3:

                clear_acrylic_sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                        VALUES (?, ?, ?)
                                    """, (component_id, 40, clear_acrylic_sheets))

            elif factors["factor1"] == 2:

                white_poly_sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                        VALUES (?, ?, ?)
                                    """, (component_id, 43, white_poly_sheets))

            elif factors["factor1"] == 1:
                clear_poly_sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                cursor.execute("""
                                        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                        VALUES (?, ?, ?)
                                    """, (component_id, 44, clear_poly_sheets))

            ############################################################################################################
            """CODE FOR LEDS/PS"""
            ############################################################################################################

            inserts = [
                (component_id, 49, math.ceil(sum(list_of_modules) / 200)),
                (component_id, 52, math.ceil(sum(list_of_modules) * .461 / 60)),
                (component_id, 54, 1),
                (component_id, 55, 1)
            ]

            if factors["factor4"] == 1:
                inserts.append(
                    (component_id, 56, math.ceil(math.ceil(sum(list_of_modules) * .461 / 60)) / 2))
                inserts.append((component_id, 57, len(list_of_perimeters)))

            else:
                inserts.append(
                    (component_id, 57, 1)
                )

            cursor.executemany("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, inserts)

            ############################################################################################################
            """CODE FOR INK/VINYL"""
            ############################################################################################################

            if factors["factor2"] == 3 or factors["factor2"] == 4:
                ml_ink = math.ceil(sum(list_of_areas))

                cursor.execute("""
                                            INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                            VALUES (?, ?, ?)
                                            """, (component_id, 45, ml_ink))

            if factors["factor3"] == 3 or factors["factor2"] == 4:
                ml_ink = math.ceil(sum(list_of_areas))

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 45, ml_ink))

            if factors["factor2"] == 2 or factors["factor7"] == 3:
                sqft_vinyl = math.ceil(sum(list_of_areas) / .45 / 32) * 32

                cursor.execute("""
                                        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 46, sqft_vinyl))

            if factors["factor3"] == 2 or factors["factor3"] == 3:
                sqft_vinyl = math.ceil(sum(list_of_areas) / .45 / 32) * 32

                cursor.execute("""
                                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 46, sqft_vinyl))

            ############################################################################################################
            """CODE FOR PAINT"""
            ############################################################################################################

            paint_area = 0

            paint_area += sum(list_of_areas) + sum(list_of_perimeters) * .5

            if factors["factor4"] == 2:
                paint_area += 24.5 * 2

            if factors["factor4"] == 3:
                paint_area += 24.5 * 2.5

            if paint_area != 0:
                cursor.execute("""
                                        INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 48, paint_area))

            ############################################################################################################
            ############################################################################################################
            """CODE FOR LABOR"""
            ############################################################################################################
            ############################################################################################################

            ############################################################################################################
            """DRAWING LABOR"""
            ############################################################################################################

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 1, 1))

            ############################################################################################################
            """FILE SETUP LABOR"""
            ############################################################################################################

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 2, 1.5))

            ############################################################################################################
            """ROUTING LABOR - BACKS"""
            ############################################################################################################

            sheets = math.ceil(sum(list_of_areas) / .45 / 32)

            route_time = sheets * (.0081 * (sum(list_of_perimeters) / sheets) + 30.736) / 60

            cursor.execute("""
                                        INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                        VALUES (?, ?, ?)
                                    """, (component_id, 3, route_time))

            ############################################################################################################
            """COLOR MATCHING LABOR"""
            ############################################################################################################

            if factors["factor2"] != 3 or factors["factor2"] != 4 or factors["factor3"] != 3 or factors["factor3"] != 4:
                cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                        """, (
                    component_id, 14, (2.2392 * factors["factor8"] ** 2 + 15.442 * factors["factor8"] + 32.608) / 60))

            ############################################################################################################
            """PRINTING LABOR"""
            ############################################################################################################

            if factors["factor2"] != 3 or factors["factor2"] != 4 or factors["factor3"] != 3 or factors["factor3"] != 4:
                sheets = math.ceil(sum(list_of_areas) / .45 / 32)

                print_time = sheets * (.1533 * (sum(list_of_areas) / sheets) + 36.519) / 60

                cursor.execute("""
                                        INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 14, print_time))

            ############################################################################################################
            """ROUTING LABOR - FACES"""
            ############################################################################################################

            sheets = math.ceil(sum(list_of_areas) / .45 / 32)

            route_time = sheets * (.0081 * (sum(list_of_perimeters) / sheets) + 30.736) / 60

            cursor.execute("""
                                        INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 3, route_time))

            ############################################################################################################
            """VINYL CUTTING LABOR"""
            ############################################################################################################

            if factors['factor7'] == 2 or factors['factor7'] == 3 or factors['factor7'] == 4 or factors['factor7'] == 5:
                cursor.execute("""
                                        INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 13, (10 + .0081 * sum(list_of_perimeters)) / 60))

            ############################################################################################################
            """VINYL APPLICATION LABOR"""
            ############################################################################################################

            if factors['factor7'] == 2 or factors['factor7'] == 3 or factors['factor7'] == 4 or factors['factor7'] == 5:
                cursor.execute("""
                                        INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 13, (10 * len(list_of_perimeters)) / 60))

            ############################################################################################################
            """AUTOBENDER LABOR"""
            ############################################################################################################

            autobender_labor = sum([.0224 * i + 2.4177 for i in list_of_perimeters]) / 60

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 4, autobender_labor))

            ############################################################################################################
            """REMOVE COIL PLASTIC LABOR"""
            ############################################################################################################

            coil_plastic_labor = sum([max(1.7329 * math.log(i) - 4.054, 1) for i in list_of_perimeters]) / 60

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 4, coil_plastic_labor))

            ############################################################################################################
            """WELD RETURNS TO FACES LABOR"""
            ############################################################################################################

            attach_coil_labor = sum(list_of_perimeters) * 12 / 60 / 2

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 7, attach_coil_labor))

            ############################################################################################################
            """PAINT AND PAINT PREP LABOR"""
            ############################################################################################################

            if paint_area != 0:

                cursor.execute("""
                                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                                    VALUES (?, ?, ?)
                                                    """, (component_id, 11, max(paint_area / 60, 1)))

                cursor.execute("""
                                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                                    VALUES (?, ?, ?)
                                                    """, (component_id, 12, max(paint_area / 60, 1)))

            ############################################################################################################
            """PAINT AND PAINT PREP LABOR"""
            ############################################################################################################

            cursor.execute("""
                                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                                    VALUES (?, ?, ?)
                                                    """, (component_id, 11, sum(list_of_areas) / 60))

            ############################################################################################################
            """INSTALL LEDS LABOR"""
            ############################################################################################################

            install_LEDs_labor = sum([.0024 * i ** 2 + 1.1666 * i + 3.2765 for i in list_of_modules]) / 60

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 10, install_LEDs_labor))

            ############################################################################################################
            """SILICONE LEDS LABOR"""
            ############################################################################################################

            silicone_LEDs_labor = sum([max(.5, 3.2445 * math.log(i) - 3.5799) for i in list_of_modules]) / 60

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 10, silicone_LEDs_labor))

            ############################################################################################################
            """ASSEMBLE LETTERS"""
            ############################################################################################################

            assemble_letters_labor = math.ceil(sum(list_of_perimeters) * 12 / 8) / 60 * 2

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 10, assemble_letters_labor))

            ############################################################################################################
            """PRODUCE PATTERN LABOR"""
            ############################################################################################################

            cursor.execute("""
                                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                    VALUES (?, ?, ?)
                                    """, (component_id, 5, .5))

            ############################################################################################################
            """RACEWAY LABOR"""
            ############################################################################################################

            if factors["factor4"] != 1:
                cursor.execute("""
                                        INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 7, 20 / 60))

                cursor.execute("""
                                        INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 7, 10 / 60))

                cursor.execute("""
                                        INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                                        VALUES (?, ?, ?)
                                        """, (component_id, 10, len(list_of_perimeters) * 15 / 60))

            ################################################################################################################

            conn.commit()
            update_component_totals(component_id)
            conn.close()

            # ------------------------------
            # 🔹 Refresh context
            # ------------------------------
            comp_conn = pyodbc.connect(CONN_STR)
            c = comp_conn.cursor()
            c.execute("SELECT unit_cost, unit_price FROM Components WHERE component_ID = ?", (component_id,))
            uc, up = c.fetchone()
            comp_conn.close()

            context["component_unit_cost"] = uc
            context["component_unit_price"] = up

            # Reload material/labor data for quoting screen
            context.update(
                materials=get_materials(),
                labor_types=get_labor_types(),
                component_materials=get_component_materials(component_id),
                component_labor=get_component_labor(component_id)
            )

            # ✅ Determine navigation target
            customer_id = request.form.get("customer_id", type=int)
            opportunity_id = request.form.get("opportunity_id", type=int)
            context.update(customer_id=customer_id, opportunity_id=opportunity_id)

            return render_template_string(QUOTE_CUSTOM_SIGN, **context)

            # ------------------------------
            # 🔹 GET request: show UI
            # ------------------------------
        context.update(
            excel_rows=[
                (x, y, z)
                for x, y, z in zip(
                    (getattr(row, "factor25", "") or "").split(","),
                    (getattr(row, "factor26", "") or "").split(","),
                    (getattr(row, "factor27", "") or "").split(",")
                )
                if any([x, y, z]) and not (x == "0" and y == "0" and z == "0")
            ]
        )

        return render_template_string(QUOTE_REVERSE_LIT_CHANNEL_LETTERS_COMPONENT, **context)

    ########################################################################################################################
    ########################################################################################################################
    ########################################################################################################################

    elif component_type_id == 8:
        component_masonry = get_component_masonry(component_id)
        context.update(component_masonry=component_masonry)
        return render_template_string(QUOTE_MASONRY, **context)

    elif component_type_id == 9:  # RENTAL EQUIPMENT
        component_rental_equipment = get_component_rental_equipment(component_id)

        # Try to fetch correct IDs from request or context
        opportunity_id = (
                request.args.get("opportunity_id", type=int)
                or context.get("opportunity_id")
                or getattr(context.get("line_item", {}), "opportunity_ID", 0)
                or 0
        )
        customer_id = (
                request.args.get("customer_id", type=int)
                or context.get("customer_id")
                or 0
        )

        # ✅ Update context so template knows where to go back to
        context.update(
            component_rental_equipment=component_rental_equipment,
            opportunity_id=opportunity_id,
            customer_id=customer_id,
        )

        return render_template_string(QUOTE_RENTAL_EQUIPMENT, **context)

    elif component_type_id == 10:  # Manual Price Entry
        if request.method == "POST":
            from decimal import Decimal, InvalidOperation

            # Get inputs
            price_raw = (request.form.get("unit_price") or "").strip().replace(",", "")
            customer_id = request.form.get("customer_id", type=int)
            opportunity_id = request.form.get("opportunity_id", type=int)

            try:
                new_price = float(Decimal(price_raw))
            except (InvalidOperation, ValueError):
                new_price = None

            if new_price is not None:
                conn = pyodbc.connect(CONN_STR)
                cur = conn.cursor()

                # Update component price directly
                cur.execute("""
                    UPDATE Components
                    SET unit_price = ?, unit_cost = ?
                    WHERE component_ID = ?
                """, (new_price, new_price, component_id))
                conn.commit()

                # Get line_ID for rollup
                cur.execute("SELECT line_ID FROM Components WHERE component_ID = ?", (component_id,))
                line_row = cur.fetchone()
                line_id = int(line_row.line_ID) if line_row else None

                conn.close()

                # ✅ Use your existing route helper instead of adding a new function
                if line_id:
                    update_line_item_totals_from_components(line_id)

            # ✅ Redirect correctly
            if customer_id:
                return redirect(
                    url_for("customer_detail_route", customer_id=customer_id, show_popup=True, line_id=line_id))
            elif opportunity_id:
                return redirect(
                    url_for("show_opportunity_route", opportunity_id=opportunity_id, show_popup=True, line_id=line_id))
            else:
                return redirect(url_for("index"))

        # ---------- GET request ----------
        return render_template_string(QUOTE_MANUAL_PRICE_ENTRY, **context)


    else:  # CUSTOM SIGN
        materials = get_materials()
        labor_types = get_labor_types()
        component_materials = get_component_materials(component_id)
        component_labor = get_component_labor(component_id)

        context.update(
            materials=materials,
            component_materials=component_materials,
            labor_types=labor_types,
            component_labor=component_labor,
        )

        if not opportunity_id:
            opportunity_id = 0
            context["opportunity_id"] = opportunity_id
            context["hide_back_button"] = True
        else:
            context["hide_back_button"] = False

        # Try to fetch a customer_id if possible
        customer_id = (
                locals().get("customer_id")
                or context.get("customer_id")
                or getattr(row, "customer_id", None)
                or 0  # fallback
        )
        context["customer_id"] = customer_id or 0
        context["hide_back_button"] = bool(customer_id)

        return render_template_string(QUOTE_CUSTOM_SIGN, **context)

########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################

"""ROUTE TO UPDATE THE COMPONENT QUANTITIES"""

@app.route("/component/<int:component_id>/update_quantity", methods=["POST"])
def update_component_quantity(component_id):
    new_qty = float(request.form.get("quantity", 0))

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE Components
        SET quantity = ?
        WHERE component_ID = ?
    """, (new_qty, component_id))
    conn.commit()

    # get parent line item
    cursor.execute("SELECT line_ID FROM Components WHERE component_ID = ?", (component_id,))
    line_id = cursor.fetchone().line_ID
    conn.close()

    # update totals
    update_line_item_totals(line_id)

    opportunity_id = get_opportunity_id_by_line(line_id)
    return redirect(url_for("show_opportunity_route", opportunity_id=opportunity_id))

########################################################################################################################

"""ROUTE TO UPDATE THE LINE ITEM UNIT PRICE FROM THE COMPONENT TOTALS"""

@app.route("/line_item/<int:line_id>/update_price_from_components", methods=["POST"])
def update_price_from_components(line_id):
    update_line_item_totals_from_components(line_id)
    return {"status": "ok"}

########################################################################################################################

"""ROUTE TO ADD INSTALL MATERIALS"""

@app.route("/add_install_material/<int:component_id>", methods=["POST"])
def add_install_material(component_id):
    description = request.form["material_description"].strip()
    unit = request.form["material_unit"].strip()
    unit_cost = float(request.form["unit_cost"])
    qty = float(request.form["quantity"])

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO Component_Install_Materials (component_ID, material_description, material_unit, unit_cost, quantity)
        VALUES (?, ?, ?, ?, ?)
    """, (component_id, description, unit, unit_cost, qty))
    conn.commit()
    conn.close()

    # ✅ Recalculate totals for the component
    update_install_component_totals(component_id)

    # 🔁 Preserve context
    customer_id = request.form.get("customer_id", type=int)
    opportunity_id = request.form.get("opportunity_id", type=int)

    if customer_id:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=2,
            customer_id=customer_id
        ))
    elif opportunity_id:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=2,
            opportunity_id=opportunity_id
        ))
    else:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=2
        ))

########################################################################################################################

"""ROUTE TO ADD INSTALL LABOR"""

@app.route("/add_install_labor/<int:component_id>", methods=["POST"])
def add_install_labor(component_id):
    labor_id = int(request.form["install_labor_id"])
    qty = float(request.form["quantity"])

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO Component_Install_Labor (component_ID, install_labor_ID, quantity)
        VALUES (?, ?, ?)
    """, (component_id, labor_id, qty))
    conn.commit()
    conn.close()

    # ✅ Recalculate totals for the component
    update_install_component_totals(component_id)

    # 🔁 Preserve context
    customer_id = request.form.get("customer_id", type=int)
    opportunity_id = request.form.get("opportunity_id", type=int)

    if customer_id:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=2,
            customer_id=customer_id
        ))
    elif opportunity_id:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=2,
            opportunity_id=opportunity_id
        ))
    else:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=2
        ))

########################################################################################################################

@app.route("/component/<int:component_id>/add_sub_install_cost", methods=["POST"])
def add_sub_install_cost(component_id):
    subcontractor_cost = float(request.form["subcontractor_cost"] or 0)

    # Preserve context for proper back navigation
    customer_id = request.form.get("customer_id")
    opportunity_id = request.form.get("opportunity_id")

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Insert subcontractor cost
    cursor.execute("""
        INSERT INTO subcontractor_install_cost (component_ID, subcontractor_cost)
        VALUES (?, ?)
    """, (component_id, subcontractor_cost))

    conn.commit()
    conn.close()

    # 🔥 REQUIRED: Recalculate installation totals
    update_install_component_totals(component_id)

    # Go back to the correct installation component screen
    return redirect(url_for(
        "quote_component",
        component_id=component_id,
        component_type_id=2,   # ← Correct install component type
        customer_id=customer_id,
        opportunity_id=opportunity_id
    ))

########################################################################################################################

"""ROUTE TO SHOW INSTALL QUOTE"""

@app.route("/installation/<int:component_id>")
def show_installation(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Materials
    cursor.execute("""
        SELECT material_description, material_unit, quantity, unit_cost
        FROM compoent_install_Materials
        WHERE component_id = ?
    """, (component_id,))
    component_install_materials = cursor.fetchall()

    cursor.execute("SELECT COALESCE(SUM(quantity * unit_cost), 0) FROM Install_Materials WHERE component_id = ?",
                   (component_id,))
    install_material_cost = cursor.fetchone()[0]

    # Labor
    cursor.execute("""
        SELECT install_labor_type, quantity, burden_rate
        FROM Install_Labor
        WHERE component_id = ?
    """, (component_id,))
    component_install_labor = cursor.fetchall()

    cursor.execute("SELECT COALESCE(SUM(quantity * unit_cost), 0) FROM Install_Labor WHERE component_id = ?",
                   (component_id,))
    install_labor_cost = cursor.fetchone()[0]

    # Totals already updated by update_component_totals_install
    cursor.execute("SELECT unit_cost, unit_price FROM Components WHERE component_ID = ?", (component_id,))
    component_unit_cost, component_unit_price = cursor.fetchone()

    conn.close()

    return render_template_string(
        QUOTE_INSTALLATION,
        component_install_materials=component_install_materials,
        component_install_labor=component_install_labor,
        install_material_cost=install_material_cost,
        install_labor_cost=install_labor_cost,
        component_unit_cost=component_unit_cost,
        component_unit_price=component_unit_price,
        component_id=component_id,
        # plus opportunity_id, line_id, etc
    )

########################################################################################################################

"""ROUTE TO ADD EMC UNIT"""

@app.route("/add_emc_unit/<int:component_id>", methods=["POST"])
def add_emc_unit(component_id):
    # 🔹 Read form inputs
    description = request.form["EMC_description"].strip()
    unit_cost = float(request.form["unit_cost"])
    qty = float(request.form["quantity"])

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO Component_EMC (component_ID, EMC_description, unit_cost, quantity)
        VALUES (?, ?, ?, ?)
    """, (component_id, description, unit_cost, qty))
    conn.commit()
    conn.close()

    # ✅ Recalculate component totals
    update_emc_component_totals(component_id)

    # 🔁 Preserve context (customer / opportunity)
    customer_id = request.form.get("customer_id", type=int)
    opportunity_id = request.form.get("opportunity_id", type=int)

    if customer_id:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=4,
            customer_id=customer_id
        ))
    elif opportunity_id:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=4,
            opportunity_id=opportunity_id
        ))
    else:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=4
        ))

########################################################################################################################

"""ROUTE TO SAVE PIPE AND FOUNDATION PRICING"""

@app.route("/save_pipe_foundation/<int:component_id>", methods=["POST"])
def save_pipe_foundation(component_id):
    form = request.form

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # check if already exists
    cursor.execute("""
        SELECT component_pipe_and_foundation_ID
        FROM component_pipe_and_foundation
        WHERE component_ID = ?
    """, (component_id,))
    row = cursor.fetchone()

    if row:  # update
        cursor.execute("""
            UPDATE component_pipe_and_foundation
            SET base_pipe_diameter=?, base_pipe_footage=?,
                stack_pipe1_diameter=?, stack_pipe1_footage=?,
                stack_pipe2_diameter=?, stack_pipe2_footage=?,
                stack_pipe3_diameter=?, stack_pipe3_footage=?,
                stack_pipe4_diameter=?, stack_pipe4_footage=?,
                pier_diameter=?, pier_depth=?, pier_quanitity=?,
                rectangular_footer_length=?, rectangular_footer_width=?, rectangular_footer_depth=?,
                digging_cost=?, concrete_cost=?, additional_footer_cost=?, pipe_cost=?
            WHERE component_ID=?
        """, (
            form["base_pipe_diameter"], form["base_pipe_footage"],
            form["stack_pipe1_diameter"], form["stack_pipe1_footage"],
            form["stack_pipe2_diameter"], form["stack_pipe2_footage"],
            form["stack_pipe3_diameter"], form["stack_pipe3_footage"],
            form["stack_pipe4_diameter"], form["stack_pipe4_footage"],
            form["pier_diameter"], form["pier_depth"], form["pier_quanitity"],
            form["rectangular_footer_length"], form["rectangular_footer_width"], form["rectangular_footer_depth"],
            form["digging_cost"], form["concrete_cost"], form["additional_footer_cost"], form["pipe_cost"],
            component_id
        ))
    else:  # insert new
        cursor.execute("""
            INSERT INTO component_pipe_and_foundation
            (component_ID, base_pipe_diameter, base_pipe_footage,
             stack_pipe1_diameter, stack_pipe1_footage,
             stack_pipe2_diameter, stack_pipe2_footage,
             stack_pipe3_diameter, stack_pipe3_footage,
             stack_pipe4_diameter, stack_pipe4_footage,
             pier_diameter, pier_depth, pier_quantity,
             rectangular_footer_length, rectangular_footer_width, rectangular_footer_depth,
             digging_cost, concrete_cost, additional_footer_cost, pipe_cost)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            component_id,
            form["base_pipe_diameter"], form["base_pipe_footage"],
            form["stack_pipe1_diameter"], form["stack_pipe1_footage"],
            form["stack_pipe2_diameter"], form["stack_pipe2_footage"],
            form["stack_pipe3_diameter"], form["stack_pipe3_footage"],
            form["stack_pipe4_diameter"], form["stack_pipe4_footage"],
            form["pier_diameter"], form["pier_depth"], form["pier_quantity"],
            form["rectangular_footer_length"], form["rectangular_footer_width"], form["rectangular_footer_depth"],
            form["digging_cost"], form["concrete_cost"], form["additional_footer_cost"], form["pipe_cost"]
        ))

    conn.commit()
    conn.close()

    # Update totals
    update_pipe_foundation_totals(component_id)

    return redirect(url_for("quote_component", component_id=component_id, component_type_id=3))

########################################################################################################################

"""ROUTE TO SAVE PIPE AND FOUNDATION FACTORS AND CALCULATE ENGINEERING"""

@app.route("/save_pipe_foundation_factors/<int:component_id>", methods=["POST"])
def save_pipe_foundation_factors(component_id):
    form = request.form

    # convert to floats safely (default 0 if blank)
    factors = []
    for f in [
        form.get("overall_height", 0),
        form.get("head_cabinet_height", 0),
        form.get("head_cabinet_width", 0),
        form.get("wind_speed", 0),
        form.get("exposure_type", 0),
        form.get("num_pipes", 0),
        form.get("pipe_yield_strength", 0),
        form.get("cab2_max_height", 0),
        form.get("cab2_height", 0),
        form.get("cab2_width", 0),
        form.get("cab3_max_height", 0),
        form.get("cab3_height", 0),
        form.get("cab3_width", 0),
        form.get("cab4_max_height", 0),
        form.get("cab4_height", 0),
        form.get("cab4_width", 0),
        form.get("pipe1_transition_height", 0),
        form.get("pipe2_transition_height", 0),
        form.get("pipe3_transition_height", 0),
        form.get("pipe4_transition_height", 0),
        form.get("foundation_type", 0),
        form.get("rect_footer_length", 0),
        form.get("rect_footer_width", 0),
    ]:
        try:
            factors.append(float(f))
        except ValueError:
            factors.append(0.0)
    # save to DB
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE Components
        SET factor1=?, factor2=?, factor3=?, factor4=?, factor5=?,
            factor6=?, factor7=?, factor8=?, factor9=?, factor10=?,
            factor11=?, factor12=?, factor13=?, factor14=?, factor15=?,
            factor16=?, factor17=?, factor18=?, factor19=?, factor20=?,
            factor21=?, factor22=?, factor23=?
        WHERE component_ID=?
    """, (*factors, component_id))
    conn.commit()
    conn.close()

    overall_height = factors[0]
    head_cabinet_height = factors[1]
    head_cabinet_width = factors[2]
    wind_speed = factors[3]
    exposure_type = factors[4]
    num_pipes = factors[5]
    pipe_yield_strength = factors[6]
    cab2_max_height = factors[7]
    cab2_height = factors[8]
    cab2_width = factors[9]
    cab3_max_height = factors[10]
    cab3_height = factors[11]
    cab3_width = factors[12]
    cab4_max_height = factors[13]
    cab4_height = factors[14]
    cab4_width = factors[15]
    pipe1_transition_height = factors[16]
    pipe2_transition_height = factors[17]
    pipe3_transition_height = factors[18]
    pipe4_transition_height = factors[19]
    if factors[20] == 1:

        foundation_type = "pier footer"

    elif factors[20] == 2:

        foundation_type = "rectangular pier"

    else:

        foundation_type = "other"

    rect_footer_length = factors[21]
    rect_footer_width = factors[22]

    ########################################################################################################################

    pier_depths = []
    allowable_diameters = []

    def pylon_sign_engineering_calculator(overall_height, head_cabinet_height, head_cabinet_width, wind_speed,
                                          exposure_type, number_of_pipes, pipe_yield_strength_PSI,
                                          cabinet_2_max_height=0,
                                          cabinet_2_height=0, cabinet_2_width=0, cabinet_3_max_height=0,
                                          cabinet_3_height=0, cabinet_3_width=0, cabinet_4_max_height=0,
                                          cabinet_4_height=0, cabinet_4_width=0, pipe_1_transition_height=0,
                                          pipe_2_transition_height=0, pipe_3_transition_height=0,
                                          pipe_4_transition_height=0, foundation_type='pier footer',
                                          length_of_rectangular_footer=0, width_of_rectangular_footer=0):

        base_pipe_diameter = 0
        base_pipe_footage = 0
        stack_pipe1_diameter = 0
        stack_pipe1_footage = 0
        stack_pipe2_diameter = 0
        stack_pipe2_footage = 0
        stack_pipe3_diameter = 0
        stack_pipe3_footage = 0
        stack_pipe4_diameter = 0
        stack_pipe4_footage = 0
        depth = 0

        """This code determine the required pipe thickness and diameter and foundation depth for a pylon sign.  The code
        takes an input of the overall height of the sign [ft], the height and width of up to four cabinets [ft], the
        required wind speed as determined by the municipality [mph], the exposure type [A, B or C], the yield strength of
        the pipe to be used [psi], and the heights of up to four transitions [ft]."""

        """Notes on Engineering:
         - STRUCTURE IS DESIGNED IN ACCORDANCE WITH ASCE 7-16: MINIMUM DESIGN LOADS FOR BUILDINGS AND OTHER STRUCTURES.
         - CONCRETE MINIMUM 28-DAY COMPRESSIVE STRENGTH SHALL BE 2,500 PSI.
         - FOUNDATIONS DESIGN BEARING PRESSURES ARE PER IBC CLASS 4 PRESUMPTIVE VALUES (NO SPECIAL INSPECTION REQUIRED):
         LATERAL BEARING = 150 PSF/FT.  [PER IBC 2018 LATERAL BEARING IS DOUBLED FOR SIGNAGE - TO MATCH ENGINEERING FROM 
         REVERENCE ENGINEERING.]"""

        ########################################################################################################################

        """Dictionary of commonly available steel pipe diameters and their thickness.  The dictionary key is the name that 
        the pipes are commonly called (diameter x thickness).  The first value in the dictionary is the section modulus of 
        the pipe in in^3 [pipes that would be subject to failure due to localized buckling have had their section modulus
        reduced to account for this].  The second value of the dictionary is the diameter of the pipe in inches.  The third 
        value of the dictionary is the weight of the pipe per foot in lbs-f/foot.  The last value of the dictionary is the 
        thickness of the pipe in inches."""

        available_pipes = {"3.5\"x.216\"": [1.72, 3.5, 7.58, .216], "4.5\"x.237\"": [3.21, 4.5, 10.79, .237],
                           "6.625\"x.280\"": [8.5, 6.625, 18.97, .280], "8.625\"x.322\"": [16.81, 8.625, 28.55, .322],
                           "10.75\"x.365\"": [29.2, 10.75, 40.48, .365], "12.75\"x.375\"": [43.8, 12.75, 49.56, .375],
                           "14.375\"x.375\"": [53.2, 14.375, 54.57, .375], "16\"x.375\"": [70.3, 16, 62.58, .375],
                           "18\"x.375\"": [89.6, 18, 70.59, .375], "20\"x.375\"": [111.3, 20, 78.6, .375],
                           "24\"x.375\"": [161.9, 24, 94.62, .375], "26\"x.375\"": [190.6, 26, 102.63, .375],
                           "30\"x.375\"": [254.8, 30, 118.65, .375], "32\"x.375\"": [291, 32, 126.66, .375],
                           "34\"x.375\"": [329.2, 34, 134.67, .375], "36\"x.375\"": [370.2, 36, 142.68, .375],
                           "42\"x.375\"": [506.1, 42, 166.71, .375], "48\"x.375\"": [663, 48, 190.74, .375],
                           "54\"x.375\"": [841, 54, 214.77, .375], "60\"x.375\"": [1040, 60, 238.8, .375],
                           "72\"x.375\"": [1503, 72, 286.66, .375], "84\"x.375\"": [2051, 84, 334.92, .375],
                           "96\"x.375": [2683, 96, 382.98, .375]}

        """List of  pier diameters that be commonly drilled with an auger.  Note that 10' augers aren't actually common.  
        'I cain't find a 10' diameter auger in all of Arkansas' - Some Subcontractor"""

        common_pier_diameters = [1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5, 5.5, 6, 6.5, 7, 8, 10]

        ########################################################################################################################

        """Determine the aspect ratio and clearance ratio of the sign.  Note that its common practice to only calculate the 
        shape coefficient for the top cabinet and use that value for all cabinets."""

        aspect_ratio = head_cabinet_width / head_cabinet_height  # Aspect ratio = head cabinet width / head cabinet height.
        clearance_ratio = head_cabinet_height / overall_height  # Clearance ratio = head cabinet height / overall height.

        # print(exposure_type)

        def shape_coefficient_double_linear_interpolator(aspect_ratio, clearance_ratio):
            """shape_coefficient_double_linear_interpolator takes the aspect ratio and clearance ratio of a sign and returns
            the shape coefficient per ASCE 7-16 Figure 29.3-1."""

            """Note On Why Shape Coefficient is Necessary: The forces on signs due to wind is an example of Conservation of 
            Momentum.  That is, the individual air molecules have both mass and velocity and some of that momentum is transferred
            to the sign upon impact.  However, the sign also creates a low pressure zone behind its surface, creating a negative
            pressure that also creates a force on the sign.  The shape of the sign determines the force that is applied.  Tables
            for this are created experimentally by putting shapes into fluid tunnels, hence the need for interpolation."""

            ########################################################################################################################

            """Lists of the aspect ratios and clearance ratios used for interpolation.  Values and table provided in ASCE
            7-16 Figure 29.3-1."""
            aspect_ratio_values = [.05, .1, .2, .5, 1, 2, 4, 5, 10, 20, 30, 45]
            clearance_ratio_values = [1, .9, .7, .5, .3, .2, .16]

            """List of lists of the shape coefficients sorted by aspect ratio and clearance ratio."""
            array = [[1.80, 1.70, 1.65, 1.55, 1.45, 1.40, 1.35, 1.35, 1.30, 1.30, 1.30, 1.30],
                     [1.85, 1.75, 1.70, 1.60, 1.55, 1.50, 1.45, 1.45, 1.40, 1.40, 1.40, 1.40],
                     [1.90, 1.85, 1.75, 1.70, 1.65, 1.60, 1.60, 1.55, 1.55, 1.55, 1.55, 1.55],
                     [1.95, 1.85, 1.80, 1.75, 1.75, 1.70, 1.70, 1.70, 1.70, 1.70, 1.70, 1.75],
                     [1.95, 1.90, 1.85, 1.80, 1.80, 1.80, 1.80, 1.80, 1.80, 1.85, 1.85, 1.85],
                     [1.95, 1.90, 1.85, 1.80, 1.80, 1.80, 1.80, 1.80, 1.85, 1.90, 1.90, 1.95],
                     [1.95, 1.90, 1.85, 1.85, 1.80, 1.80, 1.85, 1.85, 1.85, 1.90, 1.90, 1.95]]

            ########################################################################################################################

            """If the aspect ratio is a value that is in the list of aspect ratios set both the upper index of the aspect ratio
            list and the lower index of the aspect ratio list to that index.  If the aspect ratio is a value that is not in the 
            list of aspect ratios, loop through the list of aspect ratios and determine the first index which holds a value 
            greater than the aspect ratio of the sign.  Set the lower index equal to the upper index -1."""

            if aspect_ratio in aspect_ratio_values:  # Check if the aspect ratio is in the list of aspect ratios.

                aspect_ratio_upper_index = aspect_ratio_values.index(
                    aspect_ratio)  # Set the aspect ratio upper index to the index that holds the value of the signs aspect ratio.

                aspect_ratio_lower_index = aspect_ratio_upper_index  # Set the aspect ratio lower index to the index of the aspect ratio upper index.

                # print(aspect_ratio_upper_index)  # Print the aspect ratio upper index for testing.

                # print(aspect_ratio_lower_index)  # Print the aspect ratio lower index for testing.

            else:  # Loop if the aspect ratio is not in the list of aspect ratios.

                aspect_ratio_upper_index = 0  # Create a variable to store the aspect ratio upper index value.

                for i in aspect_ratio_values:  # Loop through the list of aspect ratio values.

                    if aspect_ratio >= i:  # If the value in the list is less than or equal to the aspect ratio go through the if statement.

                        aspect_ratio_upper_index += 1

                        aspect_ratio_lower_index = aspect_ratio_upper_index - 1

                        # print(aspect_ratio_upper_index)  # Print the aspect ratio upper index for testing.

                        # print(aspect_ratio_lower_index)  # Print the aspect ratio lower index for testing.

            ########################################################################################################################

            """If the clearance ratio is a value that is in the list of clearance ratios set both the upper index of the 
            clearance ratio list and the lower index of the clearance ratio list to that index.  If the clearance ratio is a 
            value that is not in the list of clearance ratios, loop through the list of clearance ratios and determine the first
            index which holds a value greater than the clearance ratio of the sign.  Set the lower index equal to the upper 
            index -1.  If the clearance ratio of the sign in less than .16 set the both the upper and lower indices of the
            clearance ratio to 6."""

            if clearance_ratio in clearance_ratio_values:  # Check if the clearance ratio is in the list of clearance ratios.

                clearance_ratio_upper_index = clearance_ratio_values.index(
                    clearance_ratio)  # Set the aspect clearance upper index to the index that holds the value of the signs clearance ratio.

                clearance_ratio_lower_index = clearance_ratio_upper_index  # Set the clearance ratio lower index to the index of the clearance ratio upper index.

                # print(clearance_ratio_upper_index)  # Print the clearance ratio upper index for testing.

                # print(clearance_ratio_lower_index)  # Print the clearance ratio lower index for testing.

            elif clearance_ratio < .16:  # Check if the clearance ratio is less than .16 (occurred in testing.)

                clearance_ratio_upper_index = 6  # Set the clearance ratio upper index to 6.

                clearance_ratio_lower_index = 6  # Set the clearance ratio lower index to 6.

                # print(clearance_ratio_upper_index)  # Print the clearance ratio upper index for testing.

                # print(clearance_ratio_lower_index)  # Print the clearance ratio lower index for testing.

            else:  # Loop if the clearance ratio is not in the list of clearance ratios.

                clearance_ratio_upper_index = 0  # Create a variable to store the clearance ratio upper index value.

                for i in clearance_ratio_values:  # Loop through the list of clearance ratio values.

                    if clearance_ratio <= i:  # If the value in the list is greater than or equal to the clearance ratio fo through the if statement.

                        clearance_ratio_upper_index += 1  # Add 1 to the value that stores the u[[er index of the clearance ratio.

                        clearance_ratio_lower_index = clearance_ratio_upper_index - 1  # Set the lower index of the aspect ratio eqaul the upper index minus 1.

                        # print(clearance_ratio_upper_index)  # Print the clearance ratio upper index for testing.

                        # print(clearance_ratio_lower_index)  # Print the clearance ratio lower index for testing.

            ########################################################################################################################

            """Use the upper and lower indices to to determine the values that will be used for the double linear 
            interpolation."""

            first_row = array[clearance_ratio_lower_index]
            value_one = first_row[aspect_ratio_lower_index]
            # print(value_one)
            value_two = first_row[aspect_ratio_upper_index]
            # print(value_two)
            second_row = array[clearance_ratio_upper_index]
            value_three = second_row[aspect_ratio_lower_index]
            # print(value_three)
            value_four = second_row[aspect_ratio_upper_index]
            # print(value_four)

            x = aspect_ratio
            y = clearance_ratio
            x1 = aspect_ratio_values[aspect_ratio_lower_index]
            x2 = aspect_ratio_values[aspect_ratio_upper_index]
            y1 = clearance_ratio_values[clearance_ratio_lower_index]
            y2 = clearance_ratio_values[clearance_ratio_upper_index]

            ########################################################################################################################

            """If both the clearance ratio and aspect ratio are not in the list of clearance ratios and list of aspect ratios
            respectively, perform a double linear interpolation.  If either the clearance ratio or the aspect ratio are in the 
            list of clearance ratios and list of aspect ratios respectively, perform a linear interpolation.  If both the 
            clearance ratio and the aspect ratio are in the list of clearance ratios and list of aspect ratios respectively,
            simply return any of the four values."""

            if x1 != x2 and y1 != y2:
                shape_coefficient = (value_one * (x2 - x) * (y2 - y) +
                                     value_two * (x - x1) * (y2 - y) +
                                     value_three * (x2 - x) * (y - y1) +
                                     value_four * (x - x1) * (y - y1)) / ((x2 - x1) * (y2 - y1))

                # print("The shape coefficient of the top cabinet is:", shape_coefficient)

            elif x1 == x2 and y1 != y2:
                shape_coefficient = value_one + (value_three - value_one) / (y2 - y1) * (clearance_ratio - y1)
                # print("The shape coefficient of the top cabinet is:", shape_coefficient)

            elif y1 == y2 and x1 != x2:
                shape_coefficient = value_one + (value_two - value_one) / (x2 - x1) * (aspect_ratio - x1)
                # print("The shape coefficient of the top cabinet is:", shape_coefficient)

            elif y1 == y2 and x1 == x2:
                shape_coefficient = value_one
                # print("The shape coefficient of the top cabinet is:", shape_coefficient)

            # print(shape_coefficient)

            ########################################################################################################################

            return shape_coefficient

        shape_coefficient = shape_coefficient_double_linear_interpolator(aspect_ratio, clearance_ratio)

        ########################################################################################################################

        def pipe_rings_and_plate_calculator(upper_pipe_diameter, lower_pipe_diameter, lower_pipe_thickness):
            """This function takes the diameter of the upper stage pipe (so where two pipes meet, you want to use the upper
            pipe, and the length of the pipe outside of the lower pipe and returns the required dimensions of the steel plates
            and the total required length of pipe."""

            """Calculations not required, suggested plate thicknesses provided by Texas Sign Association as determined by 
            Reverence Engineering."""

            ########################################################################################################################

            if upper_pipe_diameter < 12:

                required_cap_plate_thickness = .75

                required_ring_plate_thickness = .75

                additional_pipe_length_required = (
                                                              24 + required_cap_plate_thickness + 2 * required_ring_plate_thickness) / 12

            elif 12 <= upper_pipe_diameter < 20:

                required_cap_plate_thickness = .75

                required_ring_plate_thickness = 1

                additional_pipe_length_required = (
                                                              36 + required_cap_plate_thickness + 2 * required_ring_plate_thickness) / 12

            elif 20 <= upper_pipe_diameter < 26:

                required_cap_plate_thickness = 1

                required_ring_plate_thickness = 1

                additional_pipe_length_required = (
                                                              42 + required_cap_plate_thickness + 2 * required_ring_plate_thickness) / 12

            elif 26 <= upper_pipe_diameter < 32:

                required_cap_plate_thickness = 1

                required_ring_plate_thickness = 1.25

                additional_pipe_length_required = (
                                                              54 + required_cap_plate_thickness + 2 * required_ring_plate_thickness) / 12

            elif 32 <= upper_pipe_diameter < 48:

                required_cap_plate_thickness = 1.5

                required_ring_plate_thickness = 1.25

                additional_pipe_length_required = (
                                                              72 + required_cap_plate_thickness + 2 * required_ring_plate_thickness) / 12

            elif 48 <= upper_pipe_diameter < 60:

                required_cap_plate_thickness = 1.5

                required_ring_plate_thickness = 1.25

                additional_pipe_length_required = (
                                                              90 + required_cap_plate_thickness + 2 * required_ring_plate_thickness) / 12

            elif 60 <= upper_pipe_diameter < 72:

                required_cap_plate_thickness = 2

                required_ring_plate_thickness = 1.25

                additional_pipe_length_required = (
                                                              108 + required_cap_plate_thickness + 2 * required_ring_plate_thickness) / 12

            else:

                required_cap_plate_thickness = 2

                required_ring_plate_thickness = 1.25

                additional_pipe_length_required = (
                                                              150 + required_cap_plate_thickness + 2 * required_ring_plate_thickness) / 12

            required_cap_plate_diameter = lower_pipe_diameter + 1

            required_cap_plate_hole_diameter = upper_pipe_diameter

            required_ring_plate_diameter = lower_pipe_diameter - lower_pipe_thickness

            required_ring_plate_hole_diameter = upper_pipe_diameter

            ########################################################################################################################

            return required_cap_plate_thickness, required_cap_plate_diameter, required_cap_plate_hole_diameter, \
                required_ring_plate_thickness, required_ring_plate_diameter, required_ring_plate_hole_diameter, \
                additional_pipe_length_required,

        ########################################################################################################################

        def cabinet_moment(max_height, cabinet_height, cabinet_width):

            """Function to determine the moment on a cabinet caused by a wind load.  Kz accounts for the fact that pressure
            decreases as you move up in a fluid head.  Since we assume laminar flow for the calculations, the mass flow rate
            remains constant throughout the fluid.  Since pressure decreases and mass flow rate remains constant, the speed
            of the fluid must increase."""

            if max_height <= 15:  # If the max height of the sign is less than 15, use 15 for z.

                if exposure_type == "b" or exposure_type == "B":  # Determine if the exposure type is B.

                    kz = 2.01 * (15 / 1200) ** (2 / 7)  # Exposure type B kz formula with 15 substituted for z.

                elif exposure_type == "d" or exposure_type == "D":  # Determine if the exposure type is D.

                    kz = 2.01 * (15 / 700) ** (2 / 11.5)  # Exposure type D kz formula with 15 substituted for z.

                else:  # Else exposure type is C.

                    kz = 2.01 * (15 / 900) ** (2 / 9.5)  # Exposure C kz formula with 15 substituted for z.

            else:  # If the max height of the sign is greater than 15, use the max height for z.

                if exposure_type == "b" or exposure_type == "B":  # Determine if the exposure type is B.

                    kz = 2.01 * (max_height / 1200) ** (2 / 7)  # Exposure type B kz formula with max height for z.

                elif exposure_type == "d" or exposure_type == "D":  # Determine if the exposure type is D.

                    kz = 2.01 * (max_height / 700) ** (2 / 11.5)  # Exposure type D kz formula with max height for z.

                else:  # Else exposure type is C.

                    kz = 2.01 * (max_height / 900) ** (2 / 9.5)  # Exposure type C kz formula with max height for z.

            qz = .00256 * kz * 1 * .85 * wind_speed ** 2  # Calculate the stagnation pressure. qz = .00256 * kz * 1 * .85 * V^2

            area = cabinet_height * cabinet_width  # Calculate the area of the cabinet.

            force = qz * .85 * area * shape_coefficient * 2 / 3  # Calculate the force of the cabinet in lbsf. F = qz * .85 * area * shape_coefficient * 2 / 3

            moment = force * (
                        max_height - cabinet_height / 2)  # Calculate the moment on the cabinet in ft-lbsf. M = F * d. d = max height - cabinet height / 2

            return moment, force

        ########################################################################################################################

        """Calculate the moment on each of the cabinets."""

        cab_1_moment = cabinet_moment(overall_height, head_cabinet_height, head_cabinet_width)[0]
        cab_2_moment = cabinet_moment(cabinet_2_max_height, cabinet_2_height, cabinet_2_width)[0]
        cab_3_moment = cabinet_moment(cabinet_3_max_height, cabinet_3_height, cabinet_3_width)[0]
        cab_4_moment = cabinet_moment(cabinet_4_max_height, cabinet_4_height, cabinet_4_width)[0]

        """Calculate the force on each of the cabinets."""

        cab_1_force = cabinet_moment(overall_height, head_cabinet_height, head_cabinet_width)[1]
        cab_2_force = cabinet_moment(cabinet_2_max_height, cabinet_2_height, cabinet_2_width)[1]
        cab_3_force = cabinet_moment(cabinet_3_max_height, cabinet_3_height, cabinet_3_width)[1]
        cab_4_force = cabinet_moment(cabinet_4_max_height, cabinet_4_height, cabinet_4_width)[1]

        """Create lists to store variables to loop through."""

        cab_forces = [cab_2_force, cab_3_force, cab_4_force]
        cab_max_heights = [cabinet_2_max_height, cabinet_3_max_height, cabinet_4_max_height]
        cab_heights = [cabinet_2_height, cabinet_3_height, cabinet_4_height]

        all_cab_forces = [cab_1_force, cab_2_force, cab_3_force, cab_4_force]
        all_cab_max_heights = [overall_height, cabinet_2_max_height, cabinet_3_max_height, cabinet_4_max_height]
        all_cab_heights = [head_cabinet_height, cabinet_2_height, cabinet_3_height, cabinet_4_height]

        ########################################################################################################################

        def pipe_shape_coefficient(pipe_diameter, pipe_length):

            """Function calculates the shape coefficient of a length of pipe.  The shape coefficient of a pipe is a
            function of the pipe diameter and the pipe length."""

            pipe_length_inches = pipe_length * 12  # Convert the units of the pipe length from feet to inches.

            if pipe_length_inches / pipe_diameter <= 1:  # If the pipe length divided by the pipe diameter is less than or equal to 1, the shape coefficient of the pipe is .5.

                return .5

            elif pipe_length_inches / pipe_diameter <= 7:  # If the pipe length divided by the pipe diameter is greater than 1 and less than or equal to 7.  Do a linear interpolation to determine the shape coefficient of the pipe.

                return .5 + (1 / 60) * ((pipe_length_inches / pipe_diameter) - 1)

            elif pipe_length_inches / pipe_diameter <= 25:  # If the pipe length divided by the pipe diameter is greater than 7 and less than or equal to 25.  Do a linear interpolation to determine the shape coefficient of the pipe.

                return .6 + (1 / 180) * ((pipe_length_inches / pipe_diameter) - 7)

            else:  # If the pipe length divided by the pipe diameter is greater than 25, the shape coefficient of the pipe is .7.

                return .7

        ########################################################################################################################

        def pipe_moment(pipe_diameter, pipe_length, pipe_visible_length, pipe_upper_height):

            """Function calculated the moment on length of pipe."""

            if pipe_upper_height <= 15:  # If the max height of the pipe section is less than 15, use 15 for z.

                if exposure_type == "b" or exposure_type == "B":  # Determine if the exposure type is B.

                    kz = 2.01 * (15 / 1200) ** (2 / 7)  # Exposure type B kz formula with 15 substituted for z.

                elif exposure_type == "d" or exposure_type == "D":  # Determine if the exposure type is D.

                    kz = 2.01 * (15 / 700) ** (2 / 11.5)  # Exposure type D kz formula with 15 substituted for z.

                else:  # Else exposure type is C.

                    kz = 2.01 * (15 / 900) ** (2 / 9.5)  # Exposure C kz formula with 15 substituted for z.

            else:  # If the max height of the sign is greater than 15, use the max height for z.

                if exposure_type == "b" or exposure_type == "B":  # Determine if the exposure type is B.

                    kz = 2.01 * (pipe_upper_height / 1200) ** (
                            2 / 7)  # Exposure type B kz formula with max height for z.

                elif exposure_type == "d" or exposure_type == "D":  # Determine if the exposure type is D.

                    kz = 2.01 * (pipe_upper_height / 700) ** (
                            2 / 11.5)  # Exposure type D kz formula with max height for z.

                else:  # Else exposure type is C.

                    kz = 2.01 * (pipe_upper_height / 900) ** (
                            2 / 9.5)  # Exposure type C kz formula with max height for z.

            qz = .00256 * kz * 1 * .85 * wind_speed ** 2  # Calculate the stagnation pressure. qz = .00256 * kz * 1 * .85 * V^2

            area = pipe_diameter / 12 * pipe_visible_length  # Calculate the area of the pipe.  Only consider the visible length of the pipe.

            force = qz * .85 * area * pipe_shape_coefficient(pipe_diameter,
                                                             pipe_length) * 2 / 3  # Calculate the force of the pipe in lbsf. F = qz * .85 * area * shape_coefficient * 2 / 3

            moment = force * pipe_length / 2  # Calculate the moment on the cabinet in ft-lbsf. M = F * d. d = max height - cabinet height / 2

            return moment

        ########################################################################################################################

        """DETERMINE THE WIND LOAD FOR THE TOP PIPE."""

        possible_top_pipes = []  # Create an empty list to store the possible top pipes, ignoring the wind load on the pipe.

        """DETERMINE THE WIND LOAD ON THE TOP PIPE IF THERE ARE NO TRANSITIONS."""

        if pipe_1_transition_height == 0:  # If there are no pipe transitions, the pipe_1_transition_height will be equal to 0.

            top_pipe_total_length = overall_height  # If there are no pipe transitions, the total length of the top pipe will be the overall height of the sign.

            top_pipe_reduction_length = head_cabinet_height + cabinet_2_height + cabinet_3_height + cabinet_4_height  # If there are no pipe transitions, the reduction in visible pipe length is equal to the sum of the heights of all the cabinets.

            top_pipe_visible_length = top_pipe_total_length - top_pipe_reduction_length  # The visible length of the top pipe is the total length of the top pipe minus the top pipe reduction length.

            """Calculate the required section modulus of the top pipe with no transition, ignoring the wind load on the pipe."""

            top_pipe_required_section_modulus = (cab_1_moment + cab_2_moment + cab_3_moment + cab_4_moment) * 12 / (
                    pipe_yield_strength_PSI * .66) / number_of_pipes

            # print(top_pipe_required_section_modulus)

            for i in available_pipes:  # Loop through all the stock pipe options.

                if available_pipes.get(i)[
                    0] > top_pipe_required_section_modulus:  # If the section modulus of a pipe is greater than the required section modulus of the pipe, append that pipe to the list of possible top pipes.

                    possible_top_pipes.append(i)

            # print(possible_top_pipes[0])

            possible_top_pipes_with_pipe_wind_load = []  # Create an empty list to store the possible top pipes, including the wind load on the pipe.

            for i in available_pipes:  # Loop through all the stock pipe options.

                pipe_moment_top_pipe = pipe_moment(available_pipes.get(i)[1], top_pipe_total_length,
                                                   top_pipe_visible_length,
                                                   overall_height)  # Determine the moment due to the top pipe for each stock pipe option.

                total_moment = pipe_moment_top_pipe * number_of_pipes + cab_1_moment + cab_2_moment + cab_3_moment + cab_4_moment  # Add the moment due to the top pipe to the moments due to the cabinets.

                top_pipe_required_section_modulus = total_moment * 12 / (
                        pipe_yield_strength_PSI * .66) / number_of_pipes  # Determine the required section modulus of the pipe including the wind load on the pipe.

                # print(top_pipe_required_section_modulus)

                if top_pipe_required_section_modulus <= available_pipes.get(i)[
                    0]:  # Determine if the required section modulus is greater than the section modulus of the pipe.

                    possible_top_pipes_with_pipe_wind_load.append(
                        i)  # If the section modulus of the pipe is greater than the required section modulus, append that pipe to the list of possible pipes (including the force on the pipe).

            print("The optimal top pipe for this pylon is: " + possible_top_pipes_with_pipe_wind_load[0])

            total_moment_about_ground = pipe_moment(available_pipes.get(possible_top_pipes_with_pipe_wind_load[0])[1],
                                                    top_pipe_total_length, top_pipe_visible_length,
                                                    overall_height) * number_of_pipes + cab_1_moment + cab_2_moment + cab_3_moment + cab_4_moment

            total_force_about_ground = pipe_moment(available_pipes.get(possible_top_pipes_with_pipe_wind_load[0])[1],
                                                   top_pipe_total_length, top_pipe_visible_length, overall_height) / (
                                               top_pipe_total_length / 2) * number_of_pipes + cab_1_force + cab_2_force + cab_3_force + cab_4_force

            ########################################################################################################################

            """DETERMINE THE WIND LOAD ON THE TOP PIPE IF THERE ARE TRANSITIONS."""


        else:  # If pipe_1_transition_height is not equal to zero there are transitions.

            if pipe_2_transition_height == 0:  # If pipe_2_transition_height is equal to 0, the highest pipe transition is pipe_1_transition_height.

                max_transition_height = pipe_1_transition_height

            elif pipe_3_transition_height == 0:  # If pipe_3_transition_height is equal to 0, the highest pipe transition is pipe_2_transition_height.

                max_transition_height = pipe_2_transition_height

            elif pipe_4_transition_height == 0:  # If pipe_4_transition_height is equal to 0, the highest pipe transition is pipe_3_transition_height.

                max_transition_height = pipe_3_transition_height

            else:  # Else the highest pipe transition is pipe_4_transition_height.

                max_transition_height = pipe_4_transition_height

            top_pipe_total_length = overall_height - max_transition_height  # Determine the total length of the top pipe.

            top_pipe_reduction_length = 0  # Create a variable to store the length of pipe covered by the cabinets.

            total_moment_about_max_transition_height = 0  # Create a variable to store the moment about the max transition height.

            total_force_about_max_transition_height = 0  # Create a variable to store the forces above the max transition height.

            """Determine the reduction in visible pipe length, and the moment about the max transition caused by the head 
            cabinet."""

            if overall_height - head_cabinet_height <= max_transition_height:  # If the head cabinet covers the whole length of the top pipe the visible length of the pipe is 0.

                top_pipe_reduction_length += top_pipe_total_length

                total_force_about_max_transition_height = cab_1_force * (overall_height - max_transition_height) / \
                                                          head_cabinet_height

                total_moment_about_max_transition_height = cab_1_force * (overall_height - max_transition_height) / \
                                                           head_cabinet_height * (
                                                                   overall_height - max_transition_height) / 2

            else:  # If the head cabinet does not cover the entire length of the top pipe, reduce the visible length of the top pipe by the cabinet height.

                top_pipe_reduction_length += head_cabinet_height

                total_force_about_max_transition_height += cab_1_force

                total_moment_about_max_transition_height += cab_1_force * (overall_height - max_transition_height -
                                                                           head_cabinet_height / 2)

            for i in range(
                    len(cab_max_heights)):  # Determine the visible pipe reduction and moments caused about the cabinets for the top pipe.

                if cab_max_heights[i] > max_transition_height and cab_max_heights[i] - cab_heights[
                    i] >= max_transition_height:

                    top_pipe_reduction_length += cab_heights[i]

                    total_force_about_max_transition_height += cab_forces[i]

                    total_moment_about_max_transition_height += cab_forces[i] * (
                            cab_max_heights[i] - max_transition_height
                            - cab_heights[i] / 2)

                elif cab_max_heights[i] > max_transition_height > cab_max_heights[i] - cab_heights[i]:

                    top_pipe_reduction_length += cab_max_heights[i] - max_transition_height

                    total_force_about_max_transition_height += cab_forces[i] * (
                            cab_max_heights[i] - max_transition_height) / cab_heights[i]

                    total_moment_about_max_transition_height += cab_forces[i] * (
                            cab_max_heights[i] - max_transition_height) / cab_heights[i] * (cab_max_heights[
                                                                                                i] - max_transition_height) / 2

            else:

                pass

            top_pipe_visible_length = top_pipe_total_length - top_pipe_reduction_length

            """Calculate the required section modulus of the top pipe ignoring the wind load on the pipe."""

            top_pipe_required_section_modulus = total_moment_about_max_transition_height * 12 / (
                    pipe_yield_strength_PSI * .66) / number_of_pipes

            # print(top_pipe_required_section_modulus)

            for i in available_pipes:  # Loop through all the stock pipe options.

                if available_pipes.get(i)[
                    0] > top_pipe_required_section_modulus:  # If the section modulus of a pipe is greater than the required section modulus of the pipe, append that pipe to the list of possible top pipes.

                    possible_top_pipes.append(i)

            # print(possible_top_pipes[0])

            possible_top_pipes_with_pipe_wind_load = []  # Create an empty list to store the possible top pipes, including the wind load on the pipe.

            """Calculate the required section modulus of the top pipe with including the wind load on the pipe."""

            for i in available_pipes:  # Loop through all the stock pipe options.

                pipe_moment_top_pipe = pipe_moment(available_pipes.get(i)[1], top_pipe_total_length,
                                                   top_pipe_visible_length,
                                                   overall_height)  # Determine the moment due to the top pipe for each stock pipe option.

                total_moment = pipe_moment_top_pipe * number_of_pipes + total_moment_about_max_transition_height  # Add the moment due to the top pipe to the moments due to the cabinets.

                top_pipe_required_section_modulus = total_moment * 12 / (
                        pipe_yield_strength_PSI * .66) / number_of_pipes  # Determine the required section modulus of the pipe including the wind load on the pipe.

                # print(top_pipe_required_section_modulus)

                if top_pipe_required_section_modulus <= available_pipes.get(i)[
                    0]:  # Determine if the required section modulus is greater than the section modulus of the pipe.

                    possible_top_pipes_with_pipe_wind_load.append(
                        i)  # If the section modulus of the pipe is greater than the required section modulus, append that pipe to the list of possible pipes (including the force on the pipe).

            print("The optimal top pipe for this pylon is: " + possible_top_pipes_with_pipe_wind_load[0])

            total_force_about_max_transition_height += pipe_moment(
                available_pipes.get(possible_top_pipes_with_pipe_wind_load[0])[1], top_pipe_total_length,
                top_pipe_visible_length, overall_height) / (top_pipe_total_length / 2) * number_of_pipes

            total_moment_about_max_transition_height += pipe_moment(
                available_pipes.get(possible_top_pipes_with_pipe_wind_load[0])[1], top_pipe_total_length,
                top_pipe_visible_length, overall_height) * number_of_pipes

        ########################################################################################################################

        """DETERMINE THE WIND LOAD ON THE SECOND PIPE."""

        if pipe_1_transition_height == 0:

            pass

        else:

            if pipe_2_transition_height == 0:  # If pipe_2_transition_height is equal to 0, the highest pipe transition is pipe_1_transition_height.

                max_transition_height = pipe_1_transition_height

                second_highest_pipe_transition = 0

            elif pipe_3_transition_height == 0:  # If pipe_3_transition_height is equal to 0, the highest pipe transition is pipe_2_transition_height.

                max_transition_height = pipe_2_transition_height

                second_highest_pipe_transition = pipe_1_transition_height

            elif pipe_4_transition_height == 0:  # If pipe_4_transition_height is equal to 0, the highest pipe transition is pipe_3_transition_height.

                max_transition_height = pipe_3_transition_height

                second_highest_pipe_transition = pipe_2_transition_height

            else:  # Else the highest pipe transition is pipe_4_transition_height.

                max_transition_height = pipe_4_transition_height

                second_highest_pipe_transition = pipe_3_transition_height

            second_pipe_total_length = max_transition_height - second_highest_pipe_transition  # Determine the total length of the second pipe.

            second_pipe_reduction_length = 0  # Create a variable to store the length of pipe covered by the cabinets.

            total_force_about_second_transition_height = 0  # Create a varable to store the forces about the second transition height.

            total_moment_about_second_transition_height = 0  # Create a variable to store the moment about the second transition height.

            possible_second_pipes = []  # Create an empty list to store the possible top pipes, ignoring the wind load on the pipe.

            for i in range(len(all_cab_heights)):

                if all_cab_max_heights[i] >= max_transition_height and second_highest_pipe_transition >= \
                        all_cab_max_heights[i] - all_cab_heights[i]:

                    second_pipe_reduction_length += second_pipe_total_length

                    total_force_about_second_transition_height += all_cab_forces[i] * (
                            max_transition_height - second_highest_pipe_transition) / all_cab_heights[i]

                    total_moment_about_second_transition_height += all_cab_forces[i] * (
                            max_transition_height - second_highest_pipe_transition) / all_cab_heights[i] * (
                                                                           max_transition_height - second_highest_pipe_transition) / 2

                elif all_cab_max_heights[i] > max_transition_height and max_transition_height > all_cab_max_heights[i] - \
                        all_cab_heights[i] >= second_highest_pipe_transition:

                    second_pipe_reduction_length += max_transition_height - (
                            all_cab_max_heights[i] - all_cab_heights[i])

                    total_force_about_second_transition_height += all_cab_forces[i] * (
                            max_transition_height - (all_cab_max_heights[i] - all_cab_heights[i])) / \
                                                                  all_cab_heights[i]

                    total_moment_about_second_transition_height += all_cab_forces[i] * (
                            max_transition_height - (all_cab_max_heights[i] - all_cab_heights[i])) / \
                                                                   all_cab_heights[i] * (((all_cab_max_heights[i] -
                                                                                           all_cab_heights[
                                                                                               i]) - second_highest_pipe_transition) + (
                                                                                                 max_transition_height - (
                                                                                                 all_cab_max_heights[
                                                                                                     i] -
                                                                                                 all_cab_heights[
                                                                                                     i])) / 2)

                elif max_transition_height >= all_cab_max_heights[i] and all_cab_max_heights[i] - all_cab_heights[
                    i] >= second_highest_pipe_transition:

                    second_pipe_reduction_length += all_cab_heights[i]

                    total_force_about_second_transition_height += all_cab_forces[i]

                    total_moment_about_second_transition_height += all_cab_forces[i] * (
                            (all_cab_max_heights[i] - all_cab_heights[i]) - second_highest_pipe_transition +
                            all_cab_heights[i] / 2)

                elif max_transition_height > all_cab_max_heights[i] and second_highest_pipe_transition > \
                        all_cab_max_heights[i] - all_cab_heights[i] and all_cab_heights[i] != 0:

                    second_pipe_reduction_length += all_cab_max_heights[i] - second_highest_pipe_transition

                    total_force_about_second_transition_height += all_cab_forces[i] * (
                            all_cab_max_heights[i] - second_highest_pipe_transition) / all_cab_heights[i]

                    total_moment_about_second_transition_height += all_cab_forces[i] * (
                            all_cab_max_heights[i] - second_highest_pipe_transition) / all_cab_heights[i] * (
                                                                           all_cab_max_heights[
                                                                               i] - second_highest_pipe_transition) / 2

                else:

                    pass

            """Calculate the required section modulus of the second pipe ignoring the wind load on the pipe."""

            second_highest_pipe_visible_length = second_pipe_total_length - second_pipe_reduction_length
            # print(second_highest_pipe_visible_length)
            second_pipe_required_section_modulus = (
                                                           total_moment_about_second_transition_height + total_moment_about_max_transition_height
                                                           + total_force_about_max_transition_height * second_pipe_total_length) * 12 / (
                                                           pipe_yield_strength_PSI * .66) / number_of_pipes
            # print(total_moment_about_second_transition_height)
            # print(second_pipe_required_section_modulus)
            for i in available_pipes:  # Loop through all the stock pipe options.

                if available_pipes.get(i)[
                    0] > second_pipe_required_section_modulus:  # If the section modulus of a pipe is greater than the requried section modulus of the pipe, append that pipe to the list of possible top pipes.

                    possible_second_pipes.append(i)

            # print(possible_second_pipes[0])

            """Calculate the required section modulus of the second pipe including the wind load on the pipe."""

            possible_second_pipes_with_pipe_wind_load = []

            for i in available_pipes:

                pipe_moment_second_pipe = pipe_moment(available_pipes.get(i)[1], second_pipe_total_length,
                                                      second_highest_pipe_visible_length, max_transition_height)

                total_moment = pipe_moment_second_pipe * number_of_pipes + (
                        total_moment_about_second_transition_height + total_moment_about_max_transition_height
                        + total_force_about_max_transition_height * second_pipe_total_length)

                second_pipe_required_section_modulus = total_moment * 12 / (
                        pipe_yield_strength_PSI * .66) / number_of_pipes

                if second_pipe_required_section_modulus <= available_pipes.get(i)[0]:
                    possible_second_pipes_with_pipe_wind_load.append(i)

            print("The optimal second pipe for this pylon is: " + possible_second_pipes_with_pipe_wind_load[0])

            total_force_about_second_transition_height += total_force_about_max_transition_height + pipe_moment(
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[1], second_pipe_total_length,
                second_highest_pipe_visible_length, max_transition_height) / (
                                                                  second_pipe_total_length / 2) * number_of_pipes

            total_moment_about_second_transition_height += pipe_moment(
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[1], second_pipe_total_length,
                second_highest_pipe_visible_length, max_transition_height) * number_of_pipes + (
                                                                   total_moment_about_second_transition_height + total_moment_about_max_transition_height
                                                                   + total_force_about_max_transition_height * second_pipe_total_length)

        ########################################################################################################################

        """DETERMINE THE WIND LOAD ON THE THIRD PIPE."""

        if pipe_1_transition_height == 0 or pipe_2_transition_height == 0:

            pass

        else:

            if pipe_3_transition_height == 0:  # If pipe_3_transition_height is equal to 0, the second highest pipe transition is pipe_1_transition_height.

                second_highest_pipe_transition = pipe_1_transition_height

                third_highest_pipe_transition = 0

            elif pipe_4_transition_height == 0:  # If pipe_4_transition_height is equal to 0, the second highest pipe transition is pipe_2_transition_height.

                second_highest_pipe_transition = pipe_2_transition_height

                third_highest_pipe_transition = pipe_1_transition_height

            else:  # Else the second highest pipe transition is pipe_3_transition_height.

                second_highest_pipe_transition = pipe_3_transition_height

                third_highest_pipe_transition = pipe_2_transition_height

            third_pipe_total_length = second_highest_pipe_transition - third_highest_pipe_transition  # Determine the total length of the third pipe.

            third_pipe_reduction_length = 0  # Create a variable to store the length of pipe covered by the cabinets.

            total_force_about_third_transition_height = 0  # Create a vairbale to store the forces about the third transition height.

            total_moment_about_third_transition_height = 0  # Create a variable to store the moment about the third transition height.

            possible_third_pipes = []  # Create an empty list to store the possible top pipes, ignoring the wind load on the pipe.

            for i in range(len(all_cab_heights)):

                if all_cab_max_heights[i] >= second_highest_pipe_transition and third_highest_pipe_transition >= \
                        all_cab_max_heights[i] - all_cab_heights[i]:

                    third_pipe_reduction_length += third_pipe_total_length

                    total_force_about_third_transition_height += all_cab_forces[i] * (
                            second_highest_pipe_transition - third_highest_pipe_transition) / all_cab_heights[i]

                    total_moment_about_third_transition_height += all_cab_forces[i] * (
                            second_highest_pipe_transition - third_highest_pipe_transition) / all_cab_heights[i] * (
                                                                          second_highest_pipe_transition - third_highest_pipe_transition) / 2

                elif all_cab_max_heights[i] > second_highest_pipe_transition and second_highest_pipe_transition > \
                        all_cab_max_heights[i] - all_cab_heights[i] >= third_highest_pipe_transition:

                    third_pipe_reduction_length += second_highest_pipe_transition - (
                            all_cab_max_heights[i] - all_cab_heights[i])

                    total_force_about_third_transition_height += all_cab_forces[i] * (
                            second_highest_pipe_transition - (all_cab_max_heights[i] - all_cab_heights[i])) / \
                                                                 all_cab_heights[i]

                    total_moment_about_third_transition_height += all_cab_forces[i] * (
                            second_highest_pipe_transition - (all_cab_max_heights[i] - all_cab_heights[i])) / \
                                                                  all_cab_heights[i] * (((all_cab_max_heights[i] -
                                                                                          all_cab_heights[
                                                                                              i]) - third_highest_pipe_transition) + (
                                                                                                second_highest_pipe_transition - (
                                                                                                all_cab_max_heights[
                                                                                                    i] -
                                                                                                all_cab_heights[
                                                                                                    i])) / 2)

                elif second_highest_pipe_transition >= all_cab_max_heights[i] and all_cab_max_heights[i] - \
                        all_cab_heights[i] >= third_highest_pipe_transition:

                    third_pipe_reduction_length += all_cab_heights[i]

                    total_force_about_third_transition_height += all_cab_forces[i]

                    total_moment_about_third_transition_height += all_cab_forces[i] * (
                            (all_cab_max_heights[i] - all_cab_heights[i]) - third_highest_pipe_transition +
                            all_cab_heights[i] / 2)

                elif second_highest_pipe_transition > all_cab_max_heights[i] and third_highest_pipe_transition > \
                        all_cab_max_heights[i] - all_cab_heights[i] and all_cab_heights[i] != 0:

                    third_pipe_reduction_length += all_cab_max_heights[i] - third_highest_pipe_transition

                    total_force_about_third_transition_height += all_cab_forces[i] * (
                            all_cab_max_heights[i] - third_highest_pipe_transition) / all_cab_heights[i]

                    total_moment_about_third_transition_height += all_cab_forces[i] * (
                            all_cab_max_heights[i] - third_highest_pipe_transition) / all_cab_heights[i] * (
                                                                          all_cab_max_heights[
                                                                              i] - third_highest_pipe_transition) / 2

                else:

                    pass

            """Calculate the required section modulus of the third pipe ignoring the wind load on the pipe."""

            third_highest_pipe_visible_length = third_pipe_total_length - third_pipe_reduction_length

            third_pipe_required_section_modulus = (
                                                          total_moment_about_third_transition_height + total_moment_about_second_transition_height
                                                          + total_force_about_second_transition_height * third_pipe_total_length) * 12 / (
                                                          pipe_yield_strength_PSI * .66) / number_of_pipes

            for i in available_pipes:  # Loop through all the stock pipe options.

                if available_pipes.get(i)[
                    0] > third_pipe_required_section_modulus:  # If the section modulus of a pipe is greater than the requried section modulus of the pipe, append that pipe to the list of possible top pipes.

                    possible_third_pipes.append(i)

            # print(possible_third_pipes[0])

            """Calculate the required section modulus of the third pipe including the wind load on the pipe."""

            possible_third_pipes_with_pipe_wind_load = []

            for i in available_pipes:

                pipe_moment_third_pipe = pipe_moment(available_pipes.get(i)[1], third_pipe_total_length,
                                                     third_highest_pipe_visible_length, second_highest_pipe_transition)

                total_moment = pipe_moment_third_pipe * number_of_pipes + (
                        total_moment_about_third_transition_height + total_moment_about_second_transition_height
                        + total_force_about_second_transition_height * third_pipe_total_length)

                third_pipe_required_section_modulus = total_moment * 12 / (
                        pipe_yield_strength_PSI * .66) / number_of_pipes

                if third_pipe_required_section_modulus <= available_pipes.get(i)[0]:
                    possible_third_pipes_with_pipe_wind_load.append(i)

            print("The optimal third pipe for this pylon is: " + possible_third_pipes_with_pipe_wind_load[0])

            total_force_about_third_transition_height += total_force_about_second_transition_height + pipe_moment(
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[1], third_pipe_total_length,
                third_highest_pipe_visible_length, second_highest_pipe_transition) / (
                                                                 third_pipe_total_length / 2) * number_of_pipes

            total_moment_about_third_transition_height += pipe_moment(
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[1], third_pipe_total_length,
                third_highest_pipe_visible_length, second_highest_pipe_transition) * number_of_pipes + (
                                                                  total_moment_about_third_transition_height + total_moment_about_second_transition_height
                                                                  + total_force_about_second_transition_height * third_pipe_total_length)

        ########################################################################################################################

        """DETERMINE THE WIND LOAD ON THE FOURTH PIPE."""

        if pipe_1_transition_height == 0 or pipe_2_transition_height == 0 or pipe_3_transition_height == 0:

            pass

        else:

            if pipe_4_transition_height == 0:  # If pipe_4_transition_height is equal to 0, the third highest pipe transition is pipe_1_transition_height.

                third_highest_pipe_transition = pipe_1_transition_height

                fourth_highest_pipe_transition = 0

            else:  # Else the third highest pipe transition is pipe_2_transition_height.

                third_highest_pipe_transition = pipe_2_transition_height

                fourth_highest_pipe_transition = pipe_1_transition_height

            fourth_pipe_total_length = third_highest_pipe_transition - fourth_highest_pipe_transition  # Determine the total length of the fourth pipe.

            fourth_pipe_reduction_length = 0  # Create a variable to store the length of pipe covered by the cabinets.

            total_force_about_fourth_transition_height = 0  # Create a variable to store the forces about the fourth transition height.

            total_moment_about_fourth_transition_height = 0  # Create a variable to store the moment about the fourth transition height.

            possible_fourth_pipes = []  # Create an empty list to store the possible top pipes, ignoring the wind load on the pipe.

            for i in range(len(all_cab_heights)):

                if all_cab_max_heights[i] >= third_highest_pipe_transition and fourth_highest_pipe_transition >= \
                        all_cab_max_heights[i] - all_cab_heights[i]:

                    fourth_pipe_reduction_length += fourth_pipe_total_length

                    total_force_about_fourth_transition_height += all_cab_forces[i] * (
                            third_highest_pipe_transition - fourth_highest_pipe_transition) / all_cab_heights[i]

                    total_moment_about_fourth_transition_height += all_cab_forces[i] * (
                            third_highest_pipe_transition - fourth_highest_pipe_transition) / all_cab_heights[i] * (
                                                                           third_highest_pipe_transition - fourth_highest_pipe_transition) / 2

                elif all_cab_max_heights[i] > third_highest_pipe_transition and third_highest_pipe_transition > \
                        all_cab_max_heights[i] - all_cab_heights[i] >= fourth_highest_pipe_transition:

                    fourth_pipe_reduction_length += third_highest_pipe_transition - (
                            all_cab_max_heights[i] - all_cab_heights[i])

                    total_force_about_fourth_transition_height += all_cab_forces[i] * (
                            third_highest_pipe_transition - (all_cab_max_heights[i] - all_cab_heights[i])) / \
                                                                  all_cab_heights[i]

                    total_moment_about_fourth_transition_height += all_cab_forces[i] * (
                            third_highest_pipe_transition - (all_cab_max_heights[i] - all_cab_heights[i])) / \
                                                                   all_cab_heights[i] * (((all_cab_max_heights[i] -
                                                                                           all_cab_heights[
                                                                                               i]) - fourth_highest_pipe_transition) + (
                                                                                                 third_highest_pipe_transition - (
                                                                                                 all_cab_max_heights[
                                                                                                     i] -
                                                                                                 all_cab_heights[
                                                                                                     i])) / 2)

                elif third_highest_pipe_transition >= all_cab_max_heights[i] and all_cab_max_heights[i] - \
                        all_cab_heights[i] >= fourth_highest_pipe_transition:

                    fourth_pipe_reduction_length += all_cab_heights[i]

                    total_force_about_fourth_transition_height += all_cab_forces[i]

                    total_moment_about_fourth_transition_height += all_cab_forces[i] * (
                            (all_cab_max_heights[i] - all_cab_heights[i]) - fourth_highest_pipe_transition +
                            all_cab_heights[i] / 2)

                elif third_highest_pipe_transition > all_cab_max_heights[i] and fourth_highest_pipe_transition > \
                        all_cab_max_heights[i] - all_cab_heights[i] and all_cab_heights[i] != 0:

                    fourth_pipe_reduction_length += all_cab_max_heights[i] - fourth_highest_pipe_transition

                    total_force_about_fourth_transition_height += all_cab_forces[i] * (
                            all_cab_max_heights[i] - fourth_highest_pipe_transition) / all_cab_heights[i]

                    total_moment_about_fourth_transition_height += all_cab_forces[i] * (
                            all_cab_max_heights[i] - fourth_highest_pipe_transition) / all_cab_heights[i] * (
                                                                           all_cab_max_heights[
                                                                               i] - fourth_highest_pipe_transition) / 2

                else:

                    pass

            """Calculate the required section modulus of the fourth pipe ignoring the wind load on the pipe."""

            fourth_highest_pipe_visible_length = fourth_pipe_total_length - fourth_pipe_reduction_length

            fourth_pipe_required_section_modulus = (
                                                           total_moment_about_fourth_transition_height + total_moment_about_third_transition_height
                                                           + total_force_about_third_transition_height * fourth_pipe_total_length) * 12 / (
                                                           pipe_yield_strength_PSI * .66) / number_of_pipes

            for i in available_pipes:  # Loop through all the stock pipe options.

                if available_pipes.get(i)[
                    0] > fourth_pipe_required_section_modulus:  # If the section modulus of a pipe is greater than the requried section modulus of the pipe, append that pipe to the list of possible top pipes.

                    possible_fourth_pipes.append(i)

            # print(possible_fourth_pipes[0])

            """Calculate the required section modulus of the fourth pipe including the wind load on the pipe."""

            possible_fourth_pipes_with_pipe_wind_load = []

            for i in available_pipes:

                pipe_moment_fourth_pipe = pipe_moment(available_pipes.get(i)[1], fourth_pipe_total_length,
                                                      fourth_highest_pipe_visible_length, third_highest_pipe_transition)

                total_moment = pipe_moment_fourth_pipe * number_of_pipes + (
                        total_moment_about_fourth_transition_height + total_moment_about_third_transition_height
                        + total_force_about_third_transition_height * fourth_pipe_total_length)

                fourth_pipe_required_section_modulus = total_moment * 12 / (
                        pipe_yield_strength_PSI * .66) / number_of_pipes

                if fourth_pipe_required_section_modulus <= available_pipes.get(i)[0]:
                    possible_fourth_pipes_with_pipe_wind_load.append(i)

            print("The optimal fourth pipe for this pylon is: " + possible_fourth_pipes_with_pipe_wind_load[0])

            total_force_about_fourth_transition_height += total_force_about_third_transition_height + pipe_moment(
                available_pipes.get(possible_fourth_pipes_with_pipe_wind_load[0])[1], fourth_pipe_total_length,
                fourth_highest_pipe_visible_length, fourth_highest_pipe_transition) / (
                                                                  fourth_pipe_total_length / 2) * number_of_pipes

            total_moment_about_fourth_transition_height += pipe_moment(
                available_pipes.get(possible_fourth_pipes_with_pipe_wind_load[0])[1], fourth_pipe_total_length,
                fourth_highest_pipe_visible_length, third_highest_pipe_transition) * number_of_pipes + (
                                                                   total_moment_about_fourth_transition_height + total_moment_about_third_transition_height
                                                                   + total_force_about_third_transition_height * fourth_pipe_total_length)

        ########################################################################################################################

        """DETERMINE THE WIND LOAD ON THE FIFTH PIPE."""

        if pipe_1_transition_height == 0 or pipe_2_transition_height == 0 or pipe_3_transition_height == 0 or pipe_4_transition_height == 0:

            pass

        else:

            fourth_highest_pipe_transition = pipe_1_transition_height

            fifth_highest_pipe_transition = 0

            fifth_pipe_total_length = fourth_highest_pipe_transition - fifth_highest_pipe_transition  # Determine the total length of the fifth pipe.

            fifth_pipe_reduction_length = 0  # Create a variable to store the length of pipe covered by the cabinets.

            total_force_about_fifth_transition_height = 0  # Create a variabale to store the forces about the fifth transition height.

            total_moment_about_fifth_transition_height = 0  # Create a variable to store the moment about the fifth transition height.

            possible_fifth_pipes = []  # Create an empty list to store the possible top pipes, ignoring the wind load on the pipe.

            for i in range(len(all_cab_heights)):

                if all_cab_max_heights[i] >= fourth_highest_pipe_transition and fifth_highest_pipe_transition >= \
                        all_cab_max_heights[i] - all_cab_heights[i]:

                    fifth_pipe_reduction_length += fifth_pipe_total_length

                    total_force_about_fifth_transition_height += all_cab_forces[i] * (
                            fourth_highest_pipe_transition - fifth_highest_pipe_transition) / all_cab_heights[i]

                    total_moment_about_fifth_transition_height += all_cab_forces[i] * (
                            fourth_highest_pipe_transition - fifth_highest_pipe_transition) / all_cab_heights[i] * (
                                                                          fourth_highest_pipe_transition - fifth_highest_pipe_transition) / 2

                elif all_cab_max_heights[i] > fourth_highest_pipe_transition and fourth_highest_pipe_transition > \
                        all_cab_max_heights[i] - all_cab_heights[i] >= fifth_highest_pipe_transition:

                    fifth_pipe_reduction_length += fourth_highest_pipe_transition - (
                            all_cab_max_heights[i] - all_cab_heights[i])

                    total_force_about_fifth_transition_height += all_cab_forces[i] * (
                            fourth_highest_pipe_transition - (all_cab_max_heights[i] - all_cab_heights[i])) / \
                                                                 all_cab_heights[i]

                    total_moment_about_fifth_transition_height += all_cab_forces[i] * (
                            fourth_highest_pipe_transition - (all_cab_max_heights[i] - all_cab_heights[i])) / \
                                                                  all_cab_heights[i] * (((all_cab_max_heights[i] -
                                                                                          all_cab_heights[
                                                                                              i]) - fifth_highest_pipe_transition) + (
                                                                                                fourth_highest_pipe_transition - (
                                                                                                all_cab_max_heights[
                                                                                                    i] -
                                                                                                all_cab_heights[
                                                                                                    i])) / 2)

                elif fourth_highest_pipe_transition >= all_cab_max_heights[i] and all_cab_max_heights[i] - \
                        all_cab_heights[i] >= fifth_highest_pipe_transition:

                    fifth_pipe_reduction_length += all_cab_heights[i]

                    total_force_about_fifth_transition_height += all_cab_forces[i]

                    total_moment_about_fifth_transition_height += all_cab_forces[i] * (
                            (all_cab_max_heights[i] - all_cab_heights[i]) - fifth_highest_pipe_transition +
                            all_cab_heights[i] / 2)

                elif fourth_highest_pipe_transition > all_cab_max_heights[i] and fifth_highest_pipe_transition > \
                        all_cab_max_heights[i] - all_cab_heights[i] and all_cab_heights[i] != 0:

                    fifth_pipe_reduction_length += all_cab_max_heights[i] - fifth_highest_pipe_transition

                    total_force_about_fifth_transition_height += all_cab_forces[i] * (
                            all_cab_max_heights[i] - fifth_highest_pipe_transition) / all_cab_heights[i]

                    total_moment_about_fifth_transition_height += all_cab_forces[i] * (
                            all_cab_max_heights[i] - fifth_highest_pipe_transition) / all_cab_heights[i] * (
                                                                          all_cab_max_heights[
                                                                              i] - fifth_highest_pipe_transition) / 2

                else:

                    pass

            """Calculate the required section modulus of the fifth pipe ignoring the wind load on the pipe."""

            fifth_highest_pipe_visible_length = fifth_pipe_total_length - fifth_pipe_reduction_length

            fifth_pipe_required_section_modulus = (
                                                          total_moment_about_fifth_transition_height + total_moment_about_fourth_transition_height
                                                          + total_force_about_fourth_transition_height * fifth_pipe_total_length) * 12 / (
                                                          pipe_yield_strength_PSI * .66) / number_of_pipes

            for i in available_pipes:  # Loop through all the stock pipe options.

                if available_pipes.get(i)[
                    0] > fifth_pipe_required_section_modulus:  # If the section modulus of a pipe is greater than the requried section modulus of the pipe, append that pipe to the list of possible top pipes.

                    possible_fifth_pipes.append(i)

            # print(possible_fifth_pipes[0])

            """Calculate the required section modulus of the fifth pipe including the wind load on the pipe."""

            possible_fifth_pipes_with_pipe_wind_load = []

            for i in available_pipes:

                pipe_moment_fifth_pipe = pipe_moment(available_pipes.get(i)[1], fifth_pipe_total_length,
                                                     fifth_highest_pipe_visible_length, fourth_highest_pipe_transition)

                total_moment = pipe_moment_fifth_pipe * number_of_pipes + (
                        total_moment_about_fifth_transition_height + total_moment_about_fourth_transition_height
                        + total_force_about_fourth_transition_height * fifth_pipe_total_length)

                fifth_pipe_required_section_modulus = total_moment * 12 / (
                        pipe_yield_strength_PSI * .66) / number_of_pipes

                if fifth_pipe_required_section_modulus <= available_pipes.get(i)[0]:
                    possible_fifth_pipes_with_pipe_wind_load.append(i)

            print("The optimal fifth pipe for this pylon is: " + possible_fifth_pipes_with_pipe_wind_load[0])

            total_force_about_fifth_transition_height += total_force_about_fourth_transition_height + pipe_moment(
                available_pipes.get(possible_fifth_pipes_with_pipe_wind_load[0])[1], fifth_pipe_total_length,
                fifth_highest_pipe_visible_length, fifth_highest_pipe_transition) / (
                                                                 fifth_pipe_total_length / 2) * number_of_pipes

            total_moment_about_fifth_transition_height += pipe_moment(
                available_pipes.get(possible_fifth_pipes_with_pipe_wind_load[0])[1], fifth_pipe_total_length,
                fifth_highest_pipe_visible_length, fourth_highest_pipe_transition) * number_of_pipes + (
                                                                  total_moment_about_fifth_transition_height + total_moment_about_fourth_transition_height
                                                                  + total_force_about_fourth_transition_height * fifth_pipe_total_length)

        ########################################################################################################################

        """RETURN THE REQUIRED LENGTH OF PIPES TO BE PURCHASED AND THE DIMENSIONS OF THE PLATES."""

        if pipe_1_transition_height == 0:

            base_pipe_diameter = possible_top_pipes_with_pipe_wind_load[0]
            base_pipe_footage = overall_height * number_of_pipes


        elif pipe_2_transition_height == 0:

            values = pipe_rings_and_plate_calculator(available_pipes.get(possible_top_pipes_with_pipe_wind_load[0])[1],
                                                     available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[
                                                         1],
                                                     available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[
                                                         3])

            stack_pipe1_diameter = possible_top_pipes_with_pipe_wind_load[0]
            stack_pipe1_footage = (top_pipe_total_length + values[6]) * number_of_pipes

            base_pipe_diameter = possible_second_pipes_with_pipe_wind_load[0]
            base_pipe_footage = second_pipe_total_length * number_of_pipes


        elif pipe_3_transition_height == 0:

            values_1 = pipe_rings_and_plate_calculator(
                available_pipes.get(possible_top_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[3])

            stack_pipe2_diameter = possible_top_pipes_with_pipe_wind_load[0]
            stack_pipe2_footage = (top_pipe_total_length + values_1[6]) * number_of_pipes

            values_2 = pipe_rings_and_plate_calculator(
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[3])

            stack_pipe1_diameter = possible_second_pipes_with_pipe_wind_load[0]
            stack_pipe1_footage = (second_pipe_total_length + values_2[6]) * number_of_pipes

            base_pipe_diameter = possible_third_pipes_with_pipe_wind_load[0]
            base_pipe_footage = third_pipe_total_length * number_of_pipes

        elif pipe_4_transition_height == 0:

            values_1 = pipe_rings_and_plate_calculator(
                available_pipes.get(possible_top_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[3])

            stack_pipe3_diameter = possible_top_pipes_with_pipe_wind_load[0]
            stack_pipe3_footage = (top_pipe_total_length + values_1[6]) * number_of_pipes

            values_2 = pipe_rings_and_plate_calculator(
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[3])

            stack_pipe2_diameter = possible_second_pipes_with_pipe_wind_load[0]
            stack_pipe2_footage = (second_pipe_total_length + values_2[6]) * number_of_pipes

            values_3 = pipe_rings_and_plate_calculator(
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_fourth_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_fourth_pipes_with_pipe_wind_load[0])[3])

            stack_pipe1_diameter = possible_third_pipes_with_pipe_wind_load[0]
            stack_pipe1_footage = (third_pipe_total_length + values_3[6]) * number_of_pipes

            base_pipe_diameter = possible_fourth_pipes_with_pipe_wind_load[0]
            base_pipe_footage = fourth_pipe_total_length * number_of_pipes

        else:

            values_1 = pipe_rings_and_plate_calculator(
                available_pipes.get(possible_top_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[3])

            stack_pipe4_diameter = possible_top_pipes_with_pipe_wind_load[0]
            stack_pipe4_footage = (top_pipe_total_length + values_1[6]) * number_of_pipes

            values_2 = pipe_rings_and_plate_calculator(
                available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[3])

            stack_pipe3_diameter = possible_second_pipes_with_pipe_wind_load[0]
            stack_pipe3_footage = (second_pipe_total_length + values_2[6]) * number_of_pipes

            values_3 = pipe_rings_and_plate_calculator(
                available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_fourth_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_fourth_pipes_with_pipe_wind_load[0])[3])

            stack_pipe2_diameter = possible_third_pipes_with_pipe_wind_load[0]
            stack_pipe2_footage = (third_pipe_total_length + values_3[6]) * number_of_pipes

            values_4 = pipe_rings_and_plate_calculator(
                available_pipes.get(possible_fourth_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_fifth_pipes_with_pipe_wind_load[0])[1], \
                available_pipes.get(possible_fifth_pipes_with_pipe_wind_load[0])[3])

            stack_pipe1_diameter = possible_fourth_pipes_with_pipe_wind_load[0]
            stack_pipe1_footage = (fourth_pipe_total_length + values_4[6]) * number_of_pipes

            base_pipe_diameter = possible_fifth_pipes_with_pipe_wind_load[0]
            base_pipe_footage = fifth_pipe_total_length * number_of_pipes

        ########################################################################################################################

        """CALCULATE THE DIMENSIONS OF THE FOUNDATION.  THERE ARE THREE OPTIONS FOR FOOTINGS; PIER FOOTING, RECTANGULAR PIER
        FOOTING AND SPREAD FOOTING.  PIER FOOTINGS AND RECTANGULAR PIER FOOTINGS RELY ON THE HORIZONTAL BEARING PRESSURE OF
        SOIL TO RESIST OVERTURNING (THE FRICTION BETWEEN THE SOIL AND THE CONCRETE).  A SPREAD FOOTING USES THE WEIGHT OF 
        THE CONCRETE FOOTING TO PREVENT THE SIGN FROM OVERTURNING (BASICALLY ONLY USED WHEN THERE IS WATER OR ROCK THAT WILL
        BE IMMEDIATELY HIT OR THE RADIUS OF THE PIER IS GREATER THAN 8 FEET."""

        if pipe_1_transition_height == 0:

            total_moment = total_moment_about_ground

            total_force = total_force_about_ground

            moment_arm_length = total_moment / total_force

            smallest_allowed_pier_diameter = (12 + available_pipes.get(possible_top_pipes_with_pipe_wind_load[0])[
                1]) / 12

        elif pipe_2_transition_height == 0:

            total_moment = total_moment_about_second_transition_height

            total_force = total_force_about_second_transition_height

            moment_arm_length = total_moment / total_force

            smallest_allowed_pier_diameter = (12 + available_pipes.get(possible_second_pipes_with_pipe_wind_load[0])[
                1]) / 12

        elif pipe_3_transition_height == 0:

            total_moment = total_moment_about_third_transition_height

            total_force = total_force_about_third_transition_height

            moment_arm_length = total_moment / total_force

            smallest_allowed_pier_diameter = (12 + available_pipes.get(possible_third_pipes_with_pipe_wind_load[0])[
                1]) / 12

        elif pipe_4_transition_height == 0:

            total_moment = total_moment_about_fourth_transition_height

            total_force = total_force_about_fourth_transition_height

            moment_arm_length = total_moment / total_force

            smallest_allowed_pier_diameter = (12 + available_pipes.get(possible_fourth_pipes_with_pipe_wind_load[0])[
                1]) / 12

        else:

            total_moment = total_moment_about_fifth_transition_height

            total_force = total_force_about_fifth_transition_height

            moment_arm_length = total_moment / total_force

            smallest_allowed_pier_diameter = (12 + available_pipes.get(possible_fifth_pipes_with_pipe_wind_load[0])[
                1]) / 12

        ########################################################################################################################

        """CALCULATE THE ALLOWABLE DIMENSIONS OF THE FOUNDATION FOR A PIER FOOTER."""

        if foundation_type == 'pier footer':

            for i in common_pier_diameters:

                guess_depth = 10

                for counter in range(30):
                    estimated_soil_horizontal_load_bearing_capacity = 300 * guess_depth / 3

                    A = 2.34 * (total_force / number_of_pipes) / (estimated_soil_horizontal_load_bearing_capacity * i)

                    depth = (A / 2) * (1 + math.sqrt(1 + (4.36 * moment_arm_length) / A))

                    guess_depth = depth

                if i >= smallest_allowed_pier_diameter:
                    print("The required depth for a pier with diameter", i, "feet is", depth, "feet deep.")
                    allowable_diameters.append(i)
                    pier_depths.append(depth)

            """CALCULATE THE ALLOWABLE DIMENSIONS OF THE FOUNDATION FOR A RECTANGULAR PIER FOOTER."""

        elif foundation_type == 'rectangular pier':

            if length_of_rectangular_footer < smallest_allowed_pier_diameter / 12 or width_of_rectangular_footer < smallest_allowed_pier_diameter / 12:

                print('Error: Pole will not fit in footer.')

            else:

                guess_depth = 10

                for counter in range(30):
                    estimated_soil_horizontal_load_bearing_capacity = min(300 * guess_depth / 3, 1200)

                    A = 2.34 * total_force / (estimated_soil_horizontal_load_bearing_capacity * math.sqrt(
                        length_of_rectangular_footer ** 2 + width_of_rectangular_footer ** 2))

                    depth = (A / 2) * (1 + math.sqrt(1 + (4.36 * moment_arm_length) / A))

                    guess_depth = depth

            print("The required depth for a rectangular pier style footer with length", length_of_rectangular_footer,
                  "feet and width", width_of_rectangular_footer, "is", depth, "feet deep.")

            """CALCULATE THE ALLOWABLE DIMENSIONS OF THE FOUNDATION FOR A SPREAD FOOTER."""

        else:

            depth = total_moment * 4 / (150 * length_of_rectangular_footer * width_of_rectangular_footer ** 2)

            print("The required depth of a spread footer with length", length_of_rectangular_footer, "feet and width",
                  width_of_rectangular_footer, "is", depth, "feet.")

            print(150 * length_of_rectangular_footer * width_of_rectangular_footer * depth)

        return base_pipe_diameter, base_pipe_footage, stack_pipe1_diameter, stack_pipe1_footage, stack_pipe2_diameter, stack_pipe2_footage, stack_pipe3_diameter, stack_pipe3_footage, stack_pipe4_diameter, stack_pipe4_footage, depth

    ########################################################################################################################

    results = pylon_sign_engineering_calculator(overall_height, head_cabinet_height, head_cabinet_width, wind_speed,
                                                exposure_type, num_pipes, pipe_yield_strength, cab2_max_height,
                                                cab2_height, cab2_width, cab3_max_height,
                                                cab3_height, cab3_width, cab4_max_height, cab4_height, cab4_width,
                                                pipe1_transition_height, pipe2_transition_height,
                                                pipe3_transition_height, pipe4_transition_height, foundation_type,
                                                rect_footer_length, rect_footer_width)

    if foundation_type == "pier footer":
        options = [
            {"depth": pier_depths[i], "diameter": allowable_diameters[i]}
            for i in range(min(len(pier_depths), len(allowable_diameters)))
        ]
    else:
        options = []

    (base_pipe_diameter, base_pipe_footage, stack_pipe1_diameter, stack_pipe1_footage, stack_pipe2_diameter,
     stack_pipe2_footage, stack_pipe3_diameter, stack_pipe3_footage, stack_pipe4_diameter, stack_pipe4_footage,
     rectangular_footer_depth) = results

    pier_quantity = num_pipes

    # Save/update in DB
    # Save/update in DB
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    if foundation_type == "pier footer":
        # Pier footer → zero rectangular footer
        cursor.execute("""
            UPDATE component_pipe_and_foundation
            SET base_pipe_diameter=?, base_pipe_footage=?,
                stack_pipe1_diameter=?, stack_pipe1_footage=?,
                stack_pipe2_diameter=?, stack_pipe2_footage=?,
                stack_pipe3_diameter=?, stack_pipe3_footage=?,
                stack_pipe4_diameter=?, stack_pipe4_footage=?,
                rectangular_footer_length=0, rectangular_footer_width=0, rectangular_footer_depth=0,
                pier_quantity=?
            WHERE component_ID=?
        """, (
            base_pipe_diameter, base_pipe_footage,
            stack_pipe1_diameter, stack_pipe1_footage,
            stack_pipe2_diameter, stack_pipe2_footage,
            stack_pipe3_diameter, stack_pipe3_footage,
            stack_pipe4_diameter, stack_pipe4_footage,
            pier_quantity, component_id
        ))

        if cursor.rowcount == 0:
            cursor.execute("""
                INSERT INTO component_pipe_and_foundation (
                    component_ID,
                    base_pipe_diameter, base_pipe_footage,
                    stack_pipe1_diameter, stack_pipe1_footage,
                    stack_pipe2_diameter, stack_pipe2_footage,
                    stack_pipe3_diameter, stack_pipe3_footage,
                    stack_pipe4_diameter, stack_pipe4_footage,
                    rectangular_footer_length, rectangular_footer_width, rectangular_footer_depth,
                    pier_quantity
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, 0, ?)
            """, (
                component_id,
                base_pipe_diameter, base_pipe_footage,
                stack_pipe1_diameter, stack_pipe1_footage,
                stack_pipe2_diameter, stack_pipe2_footage,
                stack_pipe3_diameter, stack_pipe3_footage,
                stack_pipe4_diameter, stack_pipe4_footage,
                pier_quantity
            ))

        conn.commit()
        conn.close()

        return render_template_string(
            QUOTE_PIPE_FOUNDATION_OPTIONS,
            component_id=component_id,
            options=options
        )

    else:
        # Non-pier → zero pier fields; store rectangular length/width/depth
        pier_diameter = 0
        pier_depth = 0
        pier_quantity = 0

        # ✅ Add underground pipe length to base footage
        base_pipe_footage += rectangular_footer_depth * num_pipes

        cursor.execute("""
            UPDATE component_pipe_and_foundation
            SET base_pipe_diameter=?, base_pipe_footage=?,
                stack_pipe1_diameter=?, stack_pipe1_footage=?,
                stack_pipe2_diameter=?, stack_pipe2_footage=?,
                stack_pipe3_diameter=?, stack_pipe3_footage=?,
                stack_pipe4_diameter=?, stack_pipe4_footage=?,
                rectangular_footer_length=?, rectangular_footer_width=?, rectangular_footer_depth=?,
                pier_diameter=?, pier_depth=?, pier_quantity=?
            WHERE component_ID=?
        """, (
            base_pipe_diameter, base_pipe_footage,
            stack_pipe1_diameter, stack_pipe1_footage,
            stack_pipe2_diameter, stack_pipe2_footage,
            stack_pipe3_diameter, stack_pipe3_footage,
            stack_pipe4_diameter, stack_pipe4_footage,
            rect_footer_length, rect_footer_width, rectangular_footer_depth,
            pier_diameter, pier_depth, pier_quantity,
            component_id
        ))

        if cursor.rowcount == 0:
            cursor.execute("""
                INSERT INTO component_pipe_and_foundation (
                    component_ID,
                    base_pipe_diameter, base_pipe_footage,
                    stack_pipe1_diameter, stack_pipe1_footage,
                    stack_pipe2_diameter, stack_pipe2_footage,
                    stack_pipe3_diameter, stack_pipe3_footage,
                    stack_pipe4_diameter, stack_pipe4_footage,
                    rectangular_footer_length, rectangular_footer_width, rectangular_footer_depth,
                    pier_diameter, pier_depth, pier_quantity
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                component_id,
                base_pipe_diameter, base_pipe_footage,
                stack_pipe1_diameter, stack_pipe1_footage,
                stack_pipe2_diameter, stack_pipe2_footage,
                stack_pipe3_diameter, stack_pipe3_footage,
                stack_pipe4_diameter, stack_pipe4_footage,
                rect_footer_length, rect_footer_width, rectangular_footer_depth,
                pier_diameter, pier_depth, pier_quantity
            ))

        conn.commit()
        conn.close()

        # ✅ Go straight to costs (no options for non-pier)
        return redirect(url_for("pipe_foundation_costs", component_id=component_id))

########################################################################################################################

"""ROUTE TO SAVE FOUNDATION DIAMETER CHOICE"""

@app.route("/save_pipe_foundation_choice/<int:component_id>", methods=["POST"])
def save_pipe_foundation_choice(component_id):
    choice = request.form.get("choice")
    if not choice:
        return "No option selected", 400

    try:
        depth_str, diameter_str = choice.split("|")
        pier_depth = float(depth_str)
        pier_diameter = float(diameter_str)
    except Exception:
        return f"Invalid choice format: {choice}", 400

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Get current base_pipe_footage and pier_quantity (default to 0 if not found)
    cursor.execute("""
        SELECT base_pipe_footage, pier_quantity
        FROM component_pipe_and_foundation
        WHERE component_ID=?
    """, (component_id,))
    row = cursor.fetchone()
    if row:
        base_pipe_footage = float(row.base_pipe_footage or 0)
        pier_quantity = int(row.pier_quantity or 0)
    else:
        base_pipe_footage = 0
        pier_quantity = 0

    # Add underground portion
    new_base_pipe_footage = base_pipe_footage + (pier_depth * pier_quantity)

    # Upsert logic
    cursor.execute("""
        UPDATE component_pipe_and_foundation
        SET pier_depth=?, pier_diameter=?, base_pipe_footage=?
        WHERE component_ID=?
    """, (pier_depth, pier_diameter, new_base_pipe_footage, component_id))

    if cursor.rowcount == 0:
        cursor.execute("""
            INSERT INTO component_pipe_and_foundation (
                component_ID, pier_depth, pier_diameter,
                base_pipe_footage, digging_cost, concrete_cost,
                additional_footer_cost, pipe_cost
            )
            VALUES (?, ?, ?, ?, 0, 0, 0, 0)
        """, (component_id, pier_depth, pier_diameter, new_base_pipe_footage))

    conn.commit()
    conn.close()

    return redirect(url_for("pipe_foundation_costs", component_id=component_id))

########################################################################################################################

"""ROUTE TO SAVE FOUNDATION AND PIPE COSTS"""

@app.route("/pipe_foundation_costs/<int:component_id>", methods=["GET", "POST"])
def pipe_foundation_costs(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    cursor.execute("""
        SELECT pier_depth, pier_diameter, base_pipe_diameter, base_pipe_footage, stack_pipe1_diameter, 
               stack_pipe1_footage, stack_pipe2_diameter, stack_pipe2_footage, stack_pipe3_diameter, stack_pipe3_footage, 
               stack_pipe4_diameter, stack_pipe4_footage, pier_quantity, rectangular_footer_length, rectangular_footer_width,
               rectangular_footer_depth, digging_cost, concrete_cost, additional_footer_cost, pipe_cost
        FROM component_pipe_and_foundation
        WHERE component_ID = ?
    """, (component_id,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return "Pipe/Foundation data not found", 404

    (pier_depth, pier_diameter, base_pipe_diameter, base_pipe_footage, stack_pipe1_diameter, stack_pipe1_footage,
     stack_pipe2_diameter, stack_pipe2_footage, stack_pipe3_diameter, stack_pipe3_footage, stack_pipe4_diameter,
     stack_pipe4_footage, pier_quantity, rectangular_footer_length, rectangular_footer_width, rectangular_footer_depth,
     digging_cost, concrete_cost, additional_footer_cost, pipe_cost) = row

    if request.method == "POST":
        digging_cost = float(request.form.get("digging_cost", 0))
        concrete_cost = float(request.form.get("concrete_cost", 0))
        additional_footer_cost = float(request.form.get("additional_footer_cost", 0))
        pipe_cost = float(request.form.get("pipe_cost", 0))

        unit_cost = digging_cost + concrete_cost + additional_footer_cost + pipe_cost
        unit_price = unit_cost * 1.45

        conn = pyodbc.connect(CONN_STR)
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE component_pipe_and_foundation
            SET digging_cost=?, concrete_cost=?, additional_footer_cost=?, pipe_cost=?
            WHERE component_ID=?
        """, (digging_cost, concrete_cost, additional_footer_cost, pipe_cost, component_id))

        # ✅ update Components table too
        cursor.execute("""
            UPDATE Components
            SET unit_cost=?, unit_price=?
            WHERE component_ID=?
        """, (unit_cost, unit_price, component_id))

        conn.commit()
        conn.close()

        return redirect(url_for("quote_component", component_id=component_id, component_type_id=3))

    # ✅ fetch unit_cost and unit_price for display
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT unit_cost, unit_price
        FROM Components
        WHERE component_ID=?
    """, (component_id,))
    cost_row = cursor.fetchone()
    conn.close()

    unit_cost = cost_row.unit_cost if cost_row else 0
    unit_price = cost_row.unit_price if cost_row else 0

    # Pass current values to template
    context = dict(
        component_id=component_id,
        pier_depth=pier_depth,
        pier_diameter=pier_diameter,
        base_pipe_diameter=base_pipe_diameter,
        base_pipe_footage=base_pipe_footage,
        stack_pipe1_diameter=stack_pipe1_diameter,
        stack_pipe1_footage=stack_pipe1_footage,
        stack_pipe2_diameter=stack_pipe2_diameter,
        stack_pipe2_footage=stack_pipe2_footage,
        stack_pipe3_diameter=stack_pipe3_diameter,
        stack_pipe3_footage=stack_pipe3_footage,
        stack_pipe4_diameter=stack_pipe4_diameter,
        stack_pipe4_footage=stack_pipe4_footage,
        pier_quantity=pier_quantity,
        rectangular_footer_length=rectangular_footer_length,
        rectangular_footer_width=rectangular_footer_width,
        rectangular_footer_depth=rectangular_footer_depth,
        digging_cost=digging_cost or 0,
        concrete_cost=concrete_cost or 0,
        additional_footer_cost=additional_footer_cost or 0,
        pipe_cost=pipe_cost or 0,
        unit_cost=unit_cost,
        unit_price=unit_price,
    )
    return render_template_string(QUOTE_PIPE_FOUNDATION_COSTS, **context)

########################################################################################################################

"""ROUTE TO ADD MASONRY PRICING"""

@app.route("/add_masonry/<int:component_id>", methods=["POST"])
def add_masonry(component_id):
    description = request.form["masonry_description"]
    unit_cost = float(request.form["unit_cost"])
    qty = float(request.form["quantity"])

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO component_Masonry (component_ID, masonry_description, unit_cost, quantity)
        VALUES (?, ?, ?, ?)
    """, (component_id, description, unit_cost, qty))
    conn.commit()
    conn.close()

    update_masonry_component_totals(component_id)

    return redirect(url_for("quote_component", component_id=component_id, component_type_id=8))

########################################################################################################################

"""ROUTE TO ADD RENTAL EQUIPMENT PRICING"""

@app.route("/add_rental_equipment/<int:component_id>", methods=["POST"])
def add_rental_equipment(component_id):
    description = request.form["equipment_description"].strip()
    unit_cost = float(request.form.get("unit_cost", 0))
    qty = float(request.form.get("quantity", 0))

    # Hidden IDs (may or may not exist depending on where it was quoted)
    customer_id = int(request.form.get("customer_id", 0))
    opportunity_id = int(request.form.get("opportunity_id", 0))

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()
    cursor.execute("""
        INSERT INTO component_Rental_Equipment (component_ID, equipment_description, unit_cost, quantity)
        VALUES (?, ?, ?, ?)
    """, (component_id, description, unit_cost, qty))
    conn.commit()
    conn.close()

    # ✅ Update component totals
    update_rental_equipment_component_totals(component_id)

    # ✅ Redirect depending on context
    if customer_id:
        return redirect(url_for("quote_component",
                                component_id=component_id,
                                component_type_id=9,
                                customer_id=customer_id))
    else:
        return redirect(url_for("quote_component",
                                component_id=component_id,
                                component_type_id=9,
                                opportunity_id=opportunity_id))

########################################################################################################################

"""ROUTE TO UPDATE THE QUANTITIES OF COMPONENTS"""

@app.route("/update_component_quantities/<int:component_id>", methods=["POST"])
def update_component_quantities(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Grab lists in order
    material_ids = request.form.getlist("material_row_id[]")
    material_qtys = request.form.getlist("material_qty[]")
    labor_ids = request.form.getlist("labor_row_id[]")
    labor_qtys = request.form.getlist("labor_qty[]")

    updated_rows = 0
    deleted_rows = 0

    # ---------- MATERIALS ----------
    for rid, q in zip(material_ids, material_qtys):
        if not rid.strip():
            continue
        try:
            new_qty = float(q or 0)
        except ValueError:
            new_qty = 0.0

        if new_qty <= 0:
            print(f"Deleting material row ID {rid}")
            cursor.execute("""
                DELETE FROM [component_MFG_Materials]
                WHERE [ID] = ?
            """, (rid,))
            deleted_rows += cursor.rowcount
        else:
            print(f"Updating material row ID {rid} → {new_qty}")
            cursor.execute("""
                UPDATE [component_MFG_Materials]
                SET [quantity] = ?
                WHERE [ID] = ?
            """, (new_qty, rid))
            updated_rows += cursor.rowcount

    # ---------- LABOR ----------
    for rid, q in zip(labor_ids, labor_qtys):
        if not rid.strip():
            continue
        try:
            new_qty = float(q or 0)
        except ValueError:
            new_qty = 0.0

        if new_qty <= 0:
            print(f"Deleting labor row ID {rid}")
            cursor.execute("""
                DELETE FROM [component_MFG_Labor]
                WHERE [line_item_labor_ID] = ?
            """, (rid,))
            deleted_rows += cursor.rowcount
        else:
            print(f"Updating labor row ID {rid} → {new_qty}")
            cursor.execute("""
                UPDATE [component_MFG_Labor]
                SET [quantity] = ?
                WHERE [line_item_labor_ID] = ?
            """, (new_qty, rid))
            updated_rows += cursor.rowcount

    conn.commit()
    conn.close()

    # Recalculate totals for the component
    update_component_totals(component_id)

    # --- Retrieve query params to decide where to go back ---
    customer_id = request.args.get("customer_id", type=int)
    component_type_id = request.args.get("component_type_id", default=1, type=int)
    hide_back_button = request.args.get("hide_back_button", "false").lower() == "true"

    msg = []
    if updated_rows:
        msg.append(f"Updated {updated_rows} row(s)")
    if deleted_rows:
        msg.append(f"Deleted {deleted_rows} row(s)")
    if not msg:
        msg.append("No changes made")

    flash(" | ".join(msg), "success" if (updated_rows or deleted_rows) else "warning")

    # --- Redirect appropriately ---
    if customer_id:
        # Coming from customer details view
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=component_type_id,
            customer_id=customer_id,
            hide_back_button=True
        ))
    else:
        # Default: came from opportunity flow
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=component_type_id
        ))

########################################################################################################################

@app.route("/update_install_quantities/<int:component_id>", methods=["POST"])
def update_install_quantities(component_id):
    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Extract lists from POST form
    mat_ids = request.form.getlist("install_material_row_id[]")
    mat_qtys = request.form.getlist("install_material_qty[]")
    lab_ids = request.form.getlist("install_labor_row_id[]")
    lab_qtys = request.form.getlist("install_labor_qty[]")

    updated_rows = 0
    deleted_rows = 0

    # ==============================================================
    # 🔧 Installation Materials
    # ==============================================================
    for rid, q in zip(mat_ids, mat_qtys):
        rid = rid.strip()
        if not rid:
            continue

        try:
            new_qty = float(q or 0)
        except (TypeError, ValueError):
            new_qty = 0.0

        if new_qty <= 0:
            # Delete zero-qty materials
            cursor.execute("""
                DELETE FROM [Component_Install_Materials]
                WHERE [component_install_materials_ID] = ?
                  AND [component_ID] = ?
            """, (rid, component_id))
            deleted_rows += cursor.rowcount
        else:
            # Update existing material
            cursor.execute("""
                UPDATE [Component_Install_Materials]
                SET [quantity] = ?
                WHERE [component_install_materials_ID] = ?
                  AND [component_ID] = ?
            """, (new_qty, rid, component_id))
            updated_rows += cursor.rowcount

    # ==============================================================
    # 🧰 Installation Labor
    # ==============================================================
    for rid, q in zip(lab_ids, lab_qtys):
        rid = rid.strip()
        if not rid:
            continue

        try:
            new_qty = float(q or 0)
        except (TypeError, ValueError):
            new_qty = 0.0

        if new_qty <= 0:
            # Delete zero-qty labor rows
            cursor.execute("""
                DELETE FROM [Component_Install_Labor]
                WHERE [component_install_labor_ID] = ?
                  AND [component_ID] = ?
            """, (rid, component_id))
            deleted_rows += cursor.rowcount
        else:
            # Update existing labor rows
            cursor.execute("""
                UPDATE [Component_Install_Labor]
                SET [quantity] = ?
                WHERE [component_install_labor_ID] = ?
                  AND [component_ID] = ?
            """, (new_qty, rid, component_id))
            updated_rows += cursor.rowcount

    conn.commit()
    cursor.close()
    conn.close()

    # ==============================================================
    # 🧮 Recalculate Totals for Component
    # ==============================================================
    update_install_component_totals(component_id)

    # ==============================================================
    # 💬 User Feedback
    # ==============================================================
    msg = []
    if updated_rows:
        msg.append(f"Updated {updated_rows} row(s)")
    if deleted_rows:
        msg.append(f"Deleted {deleted_rows} row(s)")
    if not msg:
        msg.append("No changes made")

    flash(" | ".join(msg), "success" if (updated_rows or deleted_rows) else "warning")

    # ==============================================================
    # 🔁 Preserve navigation context (customer or opportunity)
    # ==============================================================
    # Read hidden form fields to know where we came from
    customer_id = request.form.get("customer_id", type=int)
    opportunity_id = request.form.get("opportunity_id", type=int)

    # Build redirect with appropriate context
    if customer_id:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=2,
            customer_id=customer_id
        ))
    elif opportunity_id:
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=2,
            opportunity_id=opportunity_id
        ))
    else:
        # Fallback (no parent context)
        return redirect(url_for(
            "quote_component",
            component_id=component_id,
            component_type_id=2
        ))

########################################################################################################################

"""ROUTE TO ADD LINE ITEMS AT THE CUSTOMER LEVEL"""

@app.route("/customer/<int:customer_id>/add_line_item", methods=["POST"])
def add_customer_line_item(customer_id):
    description = request.form.get("description")
    quantity = float(request.form.get("quantity") or 1)

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # get next available shared line_ID
    cursor.execute("SELECT MAX(line_ID) FROM Line_Items")
    max_line_items = cursor.fetchone()[0] or 0

    cursor.execute("SELECT MAX(line_ID) FROM Customer_Line_Items")
    max_customer_items = cursor.fetchone()[0] or 0

    next_line_id = max(int(max_line_items), int(max_customer_items)) + 1

    # get next sequence within this customer’s saved templates
    cursor.execute("SELECT MAX(line_item_sequence) FROM Customer_Line_Items WHERE customer_ID=?", (customer_id,))
    max_seq = cursor.fetchone()[0] or 0
    next_seq = max_seq + 1

    # insert with new unified line_ID
    cursor.execute("""
        INSERT INTO Customer_Line_Items (
            line_ID, customer_ID, line_item_description, quantity, unit_cost, unit_price, line_item_sequence
        )
        VALUES (?, ?, ?, ?, 0, 0, ?)
    """, (next_line_id, customer_id, description, quantity, next_seq))

    conn.commit()
    conn.close()

    return redirect(url_for("customer_detail_route", customer_id=customer_id))

########################################################################################################################

"""ROUTE TO UPDATE COMPONENTS AND LINE ITEMS AT THE CUSTOMER LEVEL"""

@app.route("/customer/<int:customer_id>/line_item/<int:line_id>/update_components", methods=["POST"])
def update_customer_line_item_and_components(customer_id, line_id):
    component_type_id = request.form.get("component_type_id")
    description = request.form.get("description")
    quantity = float(request.form.get("quantity") or 1)
    sequence_number = request.form.get("sequence_number")

    conn = pyodbc.connect(CONN_STR)
    cursor = conn.cursor()

    # Update base line item
    cursor.execute("""
        UPDATE Customer_Line_Items
        SET line_item_description=?, quantity=?, line_item_sequence=?
        WHERE line_ID=?
    """, (description, quantity, sequence_number, line_id))

    # Add new component if selected (MATCHES your working opportunity INSERT)
    if component_type_id:
        component_type_id = int(component_type_id)
        cursor.execute("""
            INSERT INTO Components
            (line_ID, component_type_ID, quantity, unit_cost, unit_price,
             factor1, factor2, factor3, factor4, factor5)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (line_id, component_type_id, 1, 0, 0, 0, 0, 0, 0, 0))

    conn.commit()
    conn.close()

    return redirect(url_for("customer_detail_route", customer_id=customer_id))

########################################################################################################################

"""ROUTE TO FIND INSTALL VENDORS"""

@app.route("/get_install_vendors/<int:opportunity_id>")
def get_install_vendors(opportunity_id):
    import pyodbc
    from flask import jsonify

    # --- Connect to Access ---
    conn = pyodbc.connect(
        r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
        r"DBQ=C:\\Users\\Brooks\\OneDrive\\Desktop\\Sign_App1 - Step 3.accdb;"
    )
    cursor = conn.cursor()

    # --- Get opportunity site info ---
    cursor.execute("SELECT site_state FROM Opportunities WHERE opportunity_id = ?", (opportunity_id,))
    row = cursor.fetchone()
    site_state = row[0].strip().upper() if row and row[0] else None

    # --- Define Canadian provinces ---
    canadian_provinces = {"AB", "BC", "NS", "ON", "QC"}

    # --- Select vendors by matching state/province ---
    cursor.execute("""
        SELECT vendor_name, email, city, state, zip, preferred
        FROM install_vendors
        WHERE state = ?
    """, (site_state,))

    rows = cursor.fetchall()
    conn.close()

    # --- Format JSON output ---
    vendors = []
    for r in rows:
        vendor_name, email, city, state, zip_code, preferred = r
        vendors.append({
            "vendor_name": vendor_name,
            "email": email,
            "city": city,
            "state": state,
            "zip": str(zip_code) if zip_code else "",
            "preferred": bool(preferred) if preferred is not None else False
        })

    # --- Sort preferred first ---
    vendors.sort(key=lambda x: x["preferred"], reverse=True)

    return jsonify(vendors)

########################################################################################################################

"""ROUTE TO ADD A LINE ITEM FROM THE CUSTOMER LEVEL TO THE OPPORTUNITY LEVEL"""

@app.route("/add_customer_quote_to_opportunity/<int:opportunity_id>/<int:customer_line_id>", methods=["POST"])
def add_customer_quote_to_opportunity(opportunity_id, customer_line_id):
    try:
        conn = pyodbc.connect(CONN_STR)
        cur = conn.cursor()

        # 1️⃣ Get source line item from Customer_Line_Items
        cur.execute("""
            SELECT customer_ID, line_item_description, quantity, unit_cost, unit_price
            FROM Customer_Line_Items
            WHERE line_ID=?
        """, (customer_line_id,))
        src = cur.fetchone()
        if not src:
            conn.close()
            return jsonify({"success": False, "message": "Customer line not found."}), 404

        # 2️⃣ Determine next sequence number
        cur.execute("""
            SELECT IIF(ISNULL(MAX(line_item_sequence)), 0, MAX(line_item_sequence))
            FROM Line_Items
            WHERE opportunity_ID=?
        """, (opportunity_id,))
        next_seq = (cur.fetchone()[0] or 0) + 10

        # 3️⃣ Determine next available line_ID
        cur.execute("""
            SELECT MAX(IIF(ISNULL(x.max_id),0,x.max_id))
            FROM (
                SELECT MAX(line_ID) AS max_id FROM Line_Items
                UNION ALL
                SELECT MAX(line_ID) AS max_id FROM Customer_Line_Items
            ) AS x
        """)
        next_line_id = (cur.fetchone()[0] or 0) + 1

        # 4️⃣ Insert new Line_Item
        cur.execute("""
            INSERT INTO Line_Items (
                line_ID, opportunity_ID, line_item_description,
                quantity, unit_cost, unit_price, line_item_sequence, activation_status
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, 'ACTIVE')
        """, (
            next_line_id,
            opportunity_id,
            src.line_item_description,
            src.quantity or 1,
            src.unit_cost or 0,
            src.unit_price or 0,
            next_seq
        ))
        new_line_id = next_line_id

        # 5️⃣ Copy Components
        cur.execute("""
            SELECT component_ID, component_type_ID, quantity, unit_cost, unit_price,
                   factor1, factor2, factor3, factor4, factor5, factor6, factor7, factor8,
                   factor9, factor10, factor11, factor12, factor13, factor14, factor15,
                   factor16, factor17, factor18, factor19, factor20, factor21, factor22,
                   factor23, factor24, factor25, factor26, factor27
            FROM Components
            WHERE line_ID=?
        """, (customer_line_id,))
        src_components = cur.fetchall()

        old_to_new_component = {}

        for c in src_components:
            cur.execute("""
                INSERT INTO Components (
                    line_ID, component_type_ID, quantity, unit_cost, unit_price,
                    factor1, factor2, factor3, factor4, factor5, factor6, factor7, factor8,
                    factor9, factor10, factor11, factor12, factor13, factor14, factor15,
                    factor16, factor17, factor18, factor19, factor20, factor21, factor22,
                    factor23, factor24, factor25, factor26, factor27
                )
                VALUES (
                    ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?, ?, ?,
                    ?, ?, ?, ?, ?
                )
            """, (
                new_line_id, c.component_type_ID, c.quantity or 0, c.unit_cost or 0, c.unit_price or 0,
                c.factor1, c.factor2, c.factor3, c.factor4, c.factor5, c.factor6, c.factor7, c.factor8,
                c.factor9, c.factor10, c.factor11, c.factor12, c.factor13, c.factor14, c.factor15,
                c.factor16, c.factor17, c.factor18, c.factor19, c.factor20, c.factor21, c.factor22,
                c.factor23, c.factor24, c.factor25, c.factor26, c.factor27
            ))
            cur.execute("SELECT @@IDENTITY")
            new_cid = int(cur.fetchone()[0])
            old_to_new_component[c.component_ID] = new_cid

        # 6️⃣ Copy related component tables
        if old_to_new_component:
            old_ids_tuple = tuple(old_to_new_component.keys())
            q_marks = ",".join(["?"] * len(old_ids_tuple))

            # --- MFG Materials ---
            cur.execute(f"""
                SELECT component_ID, material_ID, quantity
                FROM component_MFG_Materials
                WHERE component_ID IN ({q_marks})
            """, old_ids_tuple)
            for m in cur.fetchall():
                cur.execute("""
                    INSERT INTO component_MFG_Materials (component_ID, material_ID, quantity)
                    VALUES (?, ?, ?)
                """, (old_to_new_component[m.component_ID], m.material_ID, m.quantity or 0))

            # --- MFG Labor ---
            cur.execute(f"""
                SELECT component_ID, labor_ID, quantity
                FROM component_MFG_Labor
                WHERE component_ID IN ({q_marks})
            """, old_ids_tuple)
            for l in cur.fetchall():
                cur.execute("""
                    INSERT INTO component_MFG_Labor (component_ID, labor_ID, quantity)
                    VALUES (?, ?, ?)
                """, (old_to_new_component[l.component_ID], l.labor_ID, l.quantity or 0))

            # --- Install Materials ---
            cur.execute(f"""
                SELECT component_ID, material_description, material_unit, unit_cost, quantity
                FROM component_Install_Materials
                WHERE component_ID IN ({q_marks})
            """, old_ids_tuple)
            for im in cur.fetchall():
                cur.execute("""
                    INSERT INTO component_Install_Materials (component_ID, material_description, material_unit, unit_cost, quantity)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    old_to_new_component[im.component_ID],
                    im.material_description, im.material_unit,
                    im.unit_cost or 0, im.quantity or 0
                ))

            # --- Install Labor ---
            cur.execute(f"""
                SELECT component_ID, install_labor_ID, quantity
                FROM component_Install_Labor
                WHERE component_ID IN ({q_marks})
            """, old_ids_tuple)
            for il in cur.fetchall():
                cur.execute("""
                    INSERT INTO component_Install_Labor (component_ID, install_labor_ID, quantity)
                    VALUES (?, ?, ?)
                """, (
                    old_to_new_component[il.component_ID],
                    il.install_labor_ID, il.quantity or 0
                ))

            # --- EMC Units ---
            cur.execute(f"""
                SELECT component_ID, EMC_description, unit_cost, quantity
                FROM component_EMC
                WHERE component_ID IN ({q_marks})
            """, old_ids_tuple)
            for e in cur.fetchall():
                cur.execute("""
                    INSERT INTO component_EMC (component_ID, EMC_description, unit_cost, quantity)
                    VALUES (?, ?, ?, ?)
                """, (
                    old_to_new_component[e.component_ID],
                    e.EMC_description, e.unit_cost or 0, e.quantity or 0
                ))

            # --- Rental Equipment ---
            cur.execute(f"""
                SELECT component_ID, equipment_description, unit_cost, quantity
                FROM component_Rental_Equipment
                WHERE component_ID IN ({q_marks})
            """, old_ids_tuple)
            for re in cur.fetchall():
                cur.execute("""
                    INSERT INTO component_Rental_Equipment (component_ID, equipment_description, unit_cost, quantity)
                    VALUES (?, ?, ?, ?)
                """, (
                    old_to_new_component[re.component_ID],
                    re.equipment_description, re.unit_cost or 0, re.quantity or 0
                ))

        conn.commit()

        # 7️⃣ Recalculate totals for each new component
        for old_cid, new_cid in old_to_new_component.items():
            cur.execute("SELECT component_type_ID FROM Components WHERE component_ID=?", (new_cid,))
            ctype = cur.fetchone()[0]

            # 🚫 Skip recalculating Manual Price Entry components (ID = 10)
            if ctype == 10:
                continue

            if ctype == 1:
                update_component_totals(new_cid)
            elif ctype == 2:
                update_install_component_totals(new_cid)
            elif ctype == 4:
                update_emc_component_totals(new_cid)
            elif ctype == 9:
                update_rental_equipment_component_totals(new_cid)
            else:
                update_component_totals(new_cid)

        # 8️⃣ Roll up totals to the new Line_Item
        cur.execute("""
            SELECT
                SUM(IIF(ISNULL(unit_price),0,unit_price) * IIF(ISNULL(quantity),0,quantity)) AS total_price,
                SUM(IIF(ISNULL(unit_cost),0,unit_cost) * IIF(ISNULL(quantity),0,quantity)) AS total_cost
            FROM Components
            WHERE line_ID=?
        """, (new_line_id,))
        totals = cur.fetchone()
        total_price = float(totals.total_price or 0)
        total_cost = float(totals.total_cost or 0)

        cur.execute("""
            UPDATE Line_Items
            SET unit_price=?, unit_cost=?
            WHERE line_ID=?
        """, (total_price, total_cost, new_line_id))

        conn.commit()
        conn.close()

        print(f"[SUCCESS] Created new line item {new_line_id} with copied components and recalculated totals.")
        return jsonify({"success": True, "new_line_id": new_line_id})

    except Exception as e:
        print("❌ ERROR ADDING QUOTE:", str(e))
        import traceback; traceback.print_exc()
        try:
            conn.close()
        except:
            pass
        return jsonify({"success": False, "message": str(e)})

########################################################################################################################

"""ROUTE TO DISPLAY ALL THE LINE ITEMS ON THE CUSTOMER LEVEL ON AN OPPORTUNITY"""

@app.route("/get_customer_quotes/<int:opportunity_id>")
def get_customer_quotes(opportunity_id):
    conn = pyodbc.connect(CONN_STR)
    cur = conn.cursor()

    # Find the customer for this opportunity
    cur.execute("SELECT customer_ID FROM Opportunities WHERE opportunity_ID=?", (opportunity_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return jsonify([])

    customer_id = row.customer_ID

    # Pull all saved line items for that customer
    cur.execute("""
        SELECT line_ID, line_item_description, quantity, unit_price
        FROM Customer_Line_Items
        WHERE customer_ID=?
        ORDER BY line_item_sequence, line_ID
    """, (customer_id,))
    items = [{
        "line_ID": r.line_ID,
        "line_item_description": r.line_item_description or "",
        "quantity": float(r.quantity or 0),
        "unit_price": float(r.unit_price or 0)
    } for r in cur.fetchall()]

    conn.close()
    return jsonify(items)

########################################################################################################################

"""ROUTE TO ADD STANDARD LINE ITEMS TO OPPORTUNITIES LIKE SURVEYS AND PERMIT PROCUREMENT"""

@app.route("/add_standard_line_item/<int:opportunity_id>/<int:standard_id>", methods=["POST"])
def add_standard_line_item(opportunity_id, standard_id):
    try:
        conn = pyodbc.connect(CONN_STR)
        cur = conn.cursor()

        # Fetch from Standard_Line_Items
        cur.execute("""
            SELECT line_item_description, quantity, unit_price
            FROM Standard_Line_Items
            WHERE ID = ?
        """, (standard_id,))
        std = cur.fetchone()
        if not std:
            conn.close()
            return jsonify({"success": False, "message": "Standard item not found."}), 404

        # Determine next sequence number
        cur.execute("""
            SELECT IIF(ISNULL(MAX(line_item_sequence)), 0, MAX(line_item_sequence))
            FROM Line_Items
            WHERE opportunity_ID = ?
        """, (opportunity_id,))
        next_seq = (cur.fetchone()[0] or 0) + 10

        # Determine next available line_ID
        cur.execute("""
            SELECT MAX(IIF(ISNULL(x.max_id),0,x.max_id))
            FROM (
                SELECT MAX(line_ID) AS max_id FROM Line_Items
                UNION ALL
                SELECT MAX(line_ID) AS max_id FROM Customer_Line_Items
            ) AS x
        """)
        next_line_id = (cur.fetchone()[0] or 0) + 1

        # Insert into Line_Items (no component)
        cur.execute("""
            INSERT INTO Line_Items (
                line_ID, opportunity_ID, line_item_description, 
                quantity, unit_cost, unit_price, 
                line_item_sequence, activation_status
            )
            VALUES (?, ?, ?, ?, 0, ?, ?, 'ACTIVE')
        """, (
            next_line_id,
            opportunity_id,
            std.line_item_description,
            std.quantity or 1,
            std.unit_price or 0,
            next_seq
        ))

        conn.commit()
        conn.close()
        return jsonify({"success": True, "message": "Standard line item added."})

    except Exception as e:
        print("❌ ERROR ADDING STANDARD LINE ITEM:", str(e))
        import traceback; traceback.print_exc()
        try: conn.close()
        except: pass
        return jsonify({"success": False, "message": str(e)})

########################################################################################################################

@app.route("/login", methods=["GET", "POST"])
def login_route():
    if request.method == "POST":
        email = request.form.get("email").strip()
        password = request.form.get("password").strip()

        conn = pyodbc.connect(CONN_STR)
        cur = conn.cursor()
        cur.execute("""
            SELECT employee_ID, employee_type
            FROM Employees
            WHERE employee_email = ? AND password = ?
        """, (email, password))
        row = cur.fetchone()
        conn.close()

        if row:
            session["employee_ID"] = row.employee_ID
            session["employee_type"] = row.employee_type
            session["email"] = email

            return redirect(url_for("index"))  # go to home screen
        else:
            return render_template_string(LOGIN_TEMPLATE, error="Invalid email or password.")

    return render_template_string(LOGIN_TEMPLATE)

########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################
########################################################################################################################

if __name__ == "__main__":
    app.run(debug=True)